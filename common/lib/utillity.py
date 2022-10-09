import unittest
import re, keyboard
import os
import sys
import pyautogui
from datetime import datetime
from threading import Thread
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from common.locators.loginpage_locators import LoginPageLocators
import time
import psutil
from selenium.common.exceptions import NoSuchElementException, NoAlertPresentException,StaleElementReferenceException,TimeoutException
import warnings, shlex, subprocess,pywinauto
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.common import alert
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common import keys
from selenium.webdriver.support.color import Color
from configparser import ConfigParser, NoOptionError
from selenium.webdriver.support.ui import Select
from common.lib.core_utility import CoreUtillityMethods as core_utilobj
from common.lib.javascript import JavaScript
from common.lib import root_path
from common.lib.global_variables import Global_variables
from common.locators.dialog_locators import OpenMasterFileDialog
from common.lib.sikuli import root_path as sikuli_root
if sys.platform == 'linux':
    from pykeyboard import PyKeyboard
    pykeyboard = PyKeyboard()
    from pymouse import PyMouse
    mouse_=PyMouse()
    from PIL import Image
    import pyscreenshot as ImageGrab
else:
    import win32com.client
    import uiautomation as automation
    from uisoup import uisoup
    from pywinauto.application import Application
    from PIL import Image, ImageGrab

class UtillityMethods(object):
    
    windows=[]
    
    def __init__(self, driver):
        self.driver = driver
        self.shortwait = WebDriverWait(self.driver, 50)
        self.mediumwait = WebDriverWait(self.driver, 100)
        self.longwait = WebDriverWait(self.driver, 150)
        self.browser = UtillityMethods.parseinitfile(self, 'browser')
    
    def kill_browser_process(self):
        '''
        This function will kill browser, browser driver and python shell process.
        
        :Usage kill_browser_process()
        '''
        if sys.platform == 'linux':
            process_list = ['firefox', 'chrome']
            for p_name in process_list:
                cmd = 'pkill {0}'.format(p_name)
                os.system(cmd)
        else:
            # Iterate over all running process
            for proc in psutil.process_iter():
                try:
                    # Get process name from process object.
                    processName = proc.name()
                    if processName in ['chromedriver.exe', 'geckodriver.exe', 'IEDriverServer.exe', 'MicrosoftWebDriver.exe', 'chrome.exe', 'firefox.exe', 'MicrosoftEdge.exe', 'iexplore.exe']:
                        proc.kill()
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    pass
    
#             process_list = ['firefox', 'geckodriver', 'chrome', 'chromedriver', 'iexplore', 'IEDriverServer', 'MicrosoftEdge', 'MicrosoftWebDriver']
#             for process_name in process_list:
#                 cmd="TASKKILL /f  /FI \"IMAGENAME eq {0}*\"".format(process_name)
#                 loop_condition = True
#                 max_cont_while_loop=60
#                 while loop_condition:
#                     if max_cont_while_loop == 0:
#                         break
#                     stdout = subprocess.Popen(shlex.split(cmd), stdout=subprocess.PIPE)
#                     output = stdout.communicate()[0]
#                     if 'notasksrunning' in str(str(output.decode("utf-8")).lower().replace(' ', '')):
#                         loop_condition = False
#                     time.sleep(1)
#                     max_cont_while_loop -= 1
        
    def find_affected_testcase(self):
        """
        This function is not meant to use by any test script. 
        """  
        import glob
        os.chdir("D:\\workspace\\Sameer\\selenium\\S7385\\scripts")
        for pname in glob.glob("*.py"):
            fObj = open(pname, "r")
            line = fObj.readline()
            while line:
                if bool(re.match('.*frame.*', line)):
                    print(pname)
                    break
                else:
                    line = fObj.readline()
            
    def _validate_page(self, locator):
        self.longwait.until(EC.visibility_of_element_located(locator))
        
    def get_user_name(self):
        '''
        Description: This method will get all username and password from config file and store in global variable used_username_in_current_session as dict form.
        '''
        try:
            init_file = 'config.init'
            fileObj = open(init_file, "r")
            line = fileObj.readline().replace('\n','')
            while line:
                if 'mr' in line:
                    username = line.split(' ')
                    line = fileObj.readline().replace('\n','')
                    password_ = line.split(' ')
                    try:
                        Global_variables.used_username_in_current_session[username[1]] = password_[1]
                    except IndexError:
                        Global_variables.used_username_in_current_session[username[1]] = ''
                line = fileObj.readline().replace('\n','')
            fileObj.close()
        except Exception as e:
            print(Exception(e))
    
    def validate_and_get_webdriver_object(self, css_locator, webdriver_object_name, parent_object=None, timeout_=190, pause_time=1):
        '''
        This function is used to validate the webdriver object
        css_locator: css will be provided by User(#TableChart_1)
        webdriver_object_name: The meaningful and related name of the object will be provided by User.(Preview chart)
        '''
        try:
            if parent_object != None:
#                 UtillityMethods.synchronize_until_element_is_visible_within_parent_object(self, parent_object, css_locator, timeout_, pause_time=pause_time)
                return parent_object.find_element_by_css_selector(css_locator)
            else:
#                 UtillityMethods.synchronize_until_element_is_visible(self, css_locator, timeout_, pause_time)
                return self.driver.find_element_by_css_selector(css_locator)
        except NoSuchElementException:
            display_msg = "{0} is currently not available in the page. The Provided CSS attribute ['{1}'] might not be correct.".format(webdriver_object_name, css_locator)
            raise AttributeError(display_msg)
    
    def validate_and_get_webdriver_objects(self, css_locator, webdriver_objects_reference_name, parent_object=None, timeout_=190, pause_time=1):
        '''
        This function is used to verify the list of webdriver elements in the page
        css_locator: css will be provided by User((#TableChart_1)
        webdriver_objects_reference_name : The meaningful and related name of the objects will be provided by User.(Preview chart)
        '''
        if parent_object != None:
#             UtillityMethods.synchronize_until_element_is_visible_within_parent_object(self, parent_object, css_locator, timeout_, pause_time=pause_time)
            resp = parent_object.find_elements_by_css_selector(css_locator)
        else:
#             UtillityMethods.synchronize_until_element_is_visible(self, css_locator, timeout_, pause_time)
            resp = self.driver.find_elements_by_css_selector(css_locator)
        if len(resp) == 0:
            display_msg = "{0} is currently not available in the page. The Provided CSS attribute ['{1}'] might not be correct.".format(webdriver_objects_reference_name, css_locator)
            raise AttributeError(display_msg)
        else:
            return resp
    
    def validate_and_get_webdriver_object_using_locator(self, locator, webdriver_object_name, parent_object=None):
        '''
        This function is used to validate the webdriver object
        css_locator: css will be provided by User(#TableChart_1)
        webdriver_object_name: The meaningful and related name of the object will be provided by User.(Preview chart)
        '''
        try:
            if parent_object != None:
                return parent_object.find_element(*locator)
            else:
                return self.driver.find_element(*locator)
        except NoSuchElementException:
            display_msg = "{0} is currently not available in the page. The Provided CSS attribute ['{1}'] might not be correct.".format(webdriver_object_name, locator)
            raise AttributeError(display_msg)
    
    def validate_and_get_webdriver_objects_using_locator(self, locator, webdriver_objects_reference_name, parent_object=None):
        '''
        This function is used to verify the list of webdriver elements in the page
        css_locator: css will be provided by User((#TableChart_1)
        webdriver_objects_reference_name : The meaningful and related name of the objects will be provided by User.(Preview chart)
        '''
        if parent_object != None:
            resp = parent_object.find_elements(*locator)
        else:
            resp = self.driver.find_elements(*locator)
        if len(resp) == 0:
            display_msg = "{0} is currently not available in the page. The Provided CSS attribute ['{1}'] might not be correct.".format(webdriver_objects_reference_name, locator)
            raise AttributeError(display_msg)
        else:
            return resp
          
    def toLog(self, *args):
        fileObj = open("D:\\log.txt", "a")
        for arg in args:
            fileObj.write(arg)
            fileObj.write("\n")
        fileObj.close()
    
    
    def parseinitfile(self, key):
        init_file = 'config.init'
        config_pair = {}
        try:
            fileObj = open(init_file, "r")
            line = fileObj.readline()
            while line:
                lineObjbj = re.match(r'(\S*)\s(.*)', line)
                keyName = lineObjbj.group(1)
                config_pair[keyName] = lineObjbj.group(2)
                line = fileObj.readline()
            fileObj.close()
        except IOError:
            exit()
        if key in config_pair:
            return (config_pair[key])
        else:
            return ('Key not found')

    def verify_js_alert(self, expected_alert_msg, msg):
        """
        Syntax: utillobj.verify_js_alert('Warning: Original sort could not be determined', 'Step 10: Verify Alert')
        """
        time.sleep(1)
        try:
            als = self.driver.switch_to.alert
#             als = alert.Alert(self.driver)
            actual_alert_msg=als.text
            als.accept()
            UtillityMethods.asequal(self, True, True, msg+"- alert is available")
            UtillityMethods.asin(self, expected_alert_msg, actual_alert_msg, msg+"- alert message is displayed")
        except:
            print('Alert Prompt is not displaying.')
            UtillityMethods.asequal(self, True, False, msg)
        time.sleep(2)
        UtillityMethods.switch_to_default_content(self)
        
    def color_picker(self, color, rgb_type='rgb'):
        """
        Usage: color_picker(self,'comment', 'rgba')
        params:color='green'
        params:rgb_type='rgb' or 'rgba' or 'transparent'
        return:rgb(0, 0, 0) or rgba(0, 0, 0)
        """
        color_file=os.path.join(root_path.ROOT_PATH, 'color.data')
        section = 'transparent' if color == 'transparent' else 'DEFAULT'
        color_pair = {}
        parser = ConfigParser()
        parser.optionxform=str
        parser.read(color_file)
        try:
            color_pair[color] = parser[section][color]
            if rgb_type=='rgb':
                return (rgb_type + color_pair[color])
            elif rgb_type=='transparent':
                return(color_pair[color])
            else:
                return(rgb_type + re.sub('\)', ', 1)', color_pair[color]))
        except (KeyError, NoOptionError) as e:
            print("Specified Color is not available in color.data. " + str(e))
            return

    def wf_login(self, mrid=None, mrpass=None, add_home_path=True):
        '''
        Desc: This function is used to login to WF page. Here user can pass id password if they wish, otherwise it will read from suite.init file.
        '''
        loginid = core_utilobj.parseinitfile(self, 'mrid') if mrid==None else core_utilobj.parseinitfile(self, mrid)
        loginpwd = core_utilobj.parseinitfile(self, 'mrpass') if mrpass==None else core_utilobj.parseinitfile(self, mrpass)
        node = core_utilobj.parseinitfile(self, 'nodeid')
        port = core_utilobj.parseinitfile(self, 'httpport')
        context = core_utilobj.parseinitfile(self, 'wfcontext')
        setup_url ='http://' + node + ':' + port + context + '/home8206' if add_home_path else 'http://' + node + ':' + port + context + '/'
        refused_to_connect_msg = "refused to connect"
        cannot_reach_this_page = "can't reach this page"
        api_url_response_text= self.driver.find_element_by_css_selector("body").text.replace(' ', '').lower()
        if refused_to_connect_msg.replace(' ', '') in api_url_response_text:
            raise EnvironmentError("WebFOCUS Client {0} is down the reason is '{1}'.".format(setup_url, refused_to_connect_msg))
        if cannot_reach_this_page.replace(' ', '') in api_url_response_text:
            raise EnvironmentError("WebFOCUS Client {0} is down the reason is '{1}'.".format(setup_url, cannot_reach_this_page))
        UtillityMethods.synchronize_with_number_of_element(self, 'input[id=SignonUserName]', 1, 270)
        usename= self.driver.find_element(*LoginPageLocators.uname)
        usename.click()
        time.sleep(2)
        usename.send_keys(loginid)
        time.sleep(1)
        if loginpwd!=None:
            passwd =self.driver.find_element(*LoginPageLocators.pword)
            passwd.send_keys(loginpwd)
            time.sleep(1)
        sign_in =self.driver.find_element(*LoginPageLocators.submit)
        sign_in.click()
        time.sleep(5)
        UtillityMethods.synchronize_with_number_of_element(self, 'body', 1, 160)
        invalid_username_password_msg = "Invalidusernameorpassword"
        html_body_text= self.driver.find_element_by_css_selector("body").text.replace(' ', '')
        if invalid_username_password_msg in html_body_text:
            raise NameError("Invalid user name or password error thrown for the user -{0}".format(loginid))
        access_to_item_denied_msg = "Accesstoitemdenied"
        if access_to_item_denied_msg in html_body_text:
            raise UserWarning("Access to item denied- Project Folder or Master file doesn't exist")
        error_hint="Itemdoesnotexist"
        resource_not_found_err_msg="Received Item Does not Exist. Means Either  Repository folder OR the required fex is not available in the set up."
        if error_hint in html_body_text:
            raise LookupError(resource_not_found_err_msg)
        access_to_item_denied_msg = "Accesstoitemdenied"
        project_not_found_err_msg="Received Access to item denied Means project_id is not available"
        if access_to_item_denied_msg in html_body_text:
            raise LookupError(project_not_found_err_msg)
        core_utilobj.update_current_working_area_browser_specification(self)
        Global_variables.used_username_in_current_session[loginid] = loginpwd
        
    def wf_logout(self):
        
        setup_url = UtillityMethods.get_setup_url(self)
        logout = lambda : self.driver.get("" + setup_url.replace('home8206','') + "service/wf_security_logout.jsp ")
        if Global_variables.browser_name == 'edge' :
            click_func = lambda : UtillityMethods.click_on_edge_browser_popup_button_using_uiautomation(self)
            logout_process = Thread(target=logout)
            click_process =  Thread(target=click_func)
            logout_process.start()
            time.sleep(2)
            click_process.start()
            click_process.join()
            logout_process.join()
        else :
            logout()
            time.sleep(4)
            try:
                als = alert.Alert(self.driver)
                als.accept()
            except NoAlertPresentException:
                pass
        
    def get_setup_url(self):
        '''
        Desc: This function will return the set up URL.
        '''
        node = core_utilobj.parseinitfile(self, 'nodeid')
        port = core_utilobj.parseinitfile(self, 'httpport')
        context = core_utilobj.parseinitfile(self, 'wfcontext')
        port = '' if port == '' else ':' + port 
        setup_url = "http://{0}{1}{2}/home8206".format(node, port, context)
        return(setup_url)
    
    def verify_current_url(self, alias_url_path, msg):
        '''
        This will verify current setup Url
        @param alias_url_path: This url is after alias location
        @param msg: 'Step 9'
        :Usage verify_current_url('portal/P292_S19901/G513445/Portal_for_Context_Menu_Testing', 'Step 9')
        '''
        setup_url = UtillityMethods.get_setup_url(self).replace('home8206','')
        expected_url = setup_url + alias_url_path
        current_url = self.driver.current_url
        UtillityMethods.asequal(self, expected_url, current_url, msg)
    
    def verify_current_tab_name(self, tab_name, msg):
        '''
        This will verify current window tab name
        @param alias_url_path: This url is after alias location
        :Usage verify_tab_name('portal', 'Step 9')
        '''
        current_tab_name = self.driver.title
        UtillityMethods.asequal(self, tab_name, current_tab_name, msg)

    def synchronize_with_number_of_element(self, element_css, expected_number, expire_time, pause_time=1):
        '''
        This function synchronize with expected number of element in current browser.
        :param parent_css:-element css need to pass
        :param expected_number:-0, 1, 2,........
        :param expire_time:-1, 5,........
        :param pause_time=0.2,1...
        :Usage synchronize_with_number_of_element("[id*='RangeValuesBox'] [id*='From'] > input", 1, 190)
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                raise ValueError('The Provided CSS - [ {0} ] for synchronization mismatched'.format(element_css))
            try:
                temp_obj = self.driver.find_elements_by_css_selector(element_css)
                check_obj_exist = temp_obj[0]
                del check_obj_exist
            except IndexError:
                time.sleep(pause_time)
                continue
            except TypeError:
                time.sleep(pause_time)
                continue
            if len(temp_obj) == int(expected_number):
                run_=False
                break
            else:
                time.sleep(pause_time)
        time.sleep(pause_time)
        
    def synchronize_with_number_of_element_within_parent_object(self, parent_object, element_css, expected_number, expire_time, pause_time=1):
        '''
        This function synchronize with expected number of element in current browser.
        :param parent_css:-element css need to pass
        :param expected_number:-0, 1, 2,........
        :param expire_time:-1, 5,........
        :param pause_time=0.2,1...
        :Usage synchronize_with_number_of_element("[id*='RangeValuesBox'] [id*='From'] > input", 1, 190)
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                raise ValueError('The Provided CSS - [ {0} ] for synchronization mismatched'.format(element_css))
            try:
                temp_obj = parent_object.find_elements_by_css_selector(element_css)
                check_obj_exist = temp_obj[0]
                del check_obj_exist
            except IndexError:
                time.sleep(pause_time)
                continue
            except TypeError:
                time.sleep(pause_time)
                continue
            if len(temp_obj) == int(expected_number):
                run_=False
                break
            else:
                time.sleep(pause_time)
        time.sleep(pause_time)
    
    def synchronize_with_visble_text(self, element_css, visble_element_text, expire_time, pause_time=1, text_option='dom_visible_text', condition_type='asin'):
        '''
        This function synchronize with expected visible text of element in current browser.
        :param parent_css:-element css need to pass
        :param visble_element_text:-visible text need to pass
        :param expire_time:-1, 5,........
        :param pause_time=0.2,1...
        :usage utillityobject.synchronize_with_visble_text(self, "[id*='RangeValuesBox'] [id*='From'] > input", '2014', 190, text_option='text_value')
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                raise ValueError('The Provided CSS - [ {0} ] for synchronization mismatched'.format(element_css))
            try:
                temp_str_value_elem=self.driver.find_element_by_css_selector(element_css)
                temp_str_value=UtillityMethods.get_attribute_value(self, temp_str_value_elem, text_option)
            except NoSuchElementException:
                time.sleep(pause_time)
                continue
            except StaleElementReferenceException:
                time.sleep(pause_time)
                continue
            str_value = re.sub(' ','',temp_str_value[text_option]).replace('\n','')
            if condition_type is 'asin':
                if str(visble_element_text.replace(' ','')) in str_value:
                    run_=False
                    break
                else:
                    time.sleep(pause_time)
            elif condition_type is 'asnotin':
                if str(visble_element_text.replace(' ','')) not in str_value:
                    run_=False
                    break
                else:
                    time.sleep(pause_time)
        time.sleep(pause_time)
    
    def synchronize_with_visble_text_within_parent_object(self, parent_object, element_css, visble_element_text, expire_time, pause_time=1, text_option='dom_visible_text'):
        '''
        This function synchronize with expected visible text of element in current browser.
        :param parent_css:-element css need to pass
        :param visble_element_text:-visible text need to pass
        :param expire_time:-1, 5,........
        :param pause_time=0.2,1...
        :usage utillityobject.synchronize_with_visble_text(self, "[id*='RangeValuesBox'] [id*='From'] > input", '2014', 190, text_option='text_value')
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                raise ValueError('The Provided CSS - [ {0} ] for synchronization mismatched'.format(element_css))
            try:
                temp_str_value_elem=parent_object.find_element_by_css_selector(element_css)
                temp_str_value=UtillityMethods.get_attribute_value(self, temp_str_value_elem, text_option)
            except NoSuchElementException:
                time.sleep(pause_time)
                continue
            except StaleElementReferenceException:
                time.sleep(pause_time)
                continue
            str_value = re.sub(' ','',temp_str_value[text_option]).replace('\n','')
            if str(visble_element_text.replace(' ','')) in str_value:
                run_=False
                break
            else:
                time.sleep(pause_time)
        time.sleep(pause_time)
    
    def synchronize_until_element_disappear(self, element_css, expire_time, pause_time=1):
        '''
        This function is used to check the length of element reduced.
        :Usage synchronize_until_element_disappear("[id*='RangeValuesBox']", 9)
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                raise ValueError('The Provided CSS - [ {0} ] for synchronization mismatched'.format(element_css))
            try:
                temp_obj = self.driver.find_element_by_css_selector(element_css).is_displayed()
            except NoSuchElementException:
                time.sleep(pause_time)
                temp_obj = False
            if temp_obj == False:
                run_=False
                break
            else:
                time.sleep(pause_time)
                continue
        time.sleep(pause_time)
    
    def synchronize_until_element_disappear_within_parent_object(self, parent_object, element_css, expire_time, pause_time=1):
        '''
        This function is used to check the length of element reduced.
        :Usage synchronize_until_element_disappear("[id*='RangeValuesBox']", 9)
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                raise ValueError('The Provided CSS - [ {0} ] for synchronization mismatched'.format(element_css))
            try:
                temp_obj = parent_object.find_element_by_css_selector(element_css).is_displayed()
            except NoSuchElementException:
                time.sleep(pause_time)
                temp_obj = False
            if temp_obj == False:
                run_=False
                break
            else:
                time.sleep(pause_time)
                continue
        time.sleep(pause_time)
    
    def synchronize_until_element_is_visible(self, element_css, expire_time, pause_time=1):
        '''
        This function is used to synchronize until element is visible.
        
        :Usage synchronize_until_element_disappear("[id*='RangeValuesBox']", 9)
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                raise ValueError('The Provided CSS - [ {0} ] for synchronization mismatched'.format(element_css))
            try:
                temp_obj = self.driver.find_element_by_css_selector(element_css).is_displayed()
            except NoSuchElementException:
                time.sleep(pause_time)
                temp_obj = False
            if temp_obj == True:
                run_=False
                break
            else:
                time.sleep(pause_time)
                continue
        time.sleep(pause_time)
    
    def synchronize_until_element_is_visible_within_parent_object(self, parent_object, element_css, expire_time, pause_time=1):
        '''
        This function is used to synchronize until element is visible.
        
        :Usage synchronize_until_element_disappear("[id*='RangeValuesBox']", 9)
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                raise ValueError('The Provided CSS - [ {0} ] for synchronization mismatched'.format(element_css))
            try:
                temp_obj = parent_object.find_element_by_css_selector(element_css).is_displayed()
            except NoSuchElementException:
                time.sleep(pause_time)
                temp_obj = False
            if temp_obj == True:
                run_=False
                break
            else:
                time.sleep(pause_time)
                continue
        time.sleep(pause_time)
    
    def wait_for_page_loads(self, time_out, sleep_interval=0.5, pause_time=1):
        """
        Webdriver will wait until complete the page load
        """
        JavaScript.wait_for_page_loads(self, time_out, sleep_interval, pause_time)
                
    def verify_element_color_using_get_attribute(self, element_css, color, msg, attribute='fill'):
        '''
        Desc : This function is to verify element color using element attribute.
        User can send css of the element, color name from color data and  message, parameter's values should be given when call the function
        '''
        if attribute == 'fill':
            if Global_variables.browser_name in ['ie', 'edge']:
                temp_obj=((self.driver.find_element_by_css_selector(element_css).get_attribute(attribute))[:-9]+")")[4:]
            else:
                temp_obj=((self.driver.find_element_by_css_selector(element_css).get_attribute(attribute))[:-10]+")")[4:]
            actual_color = "rgb"+temp_obj
            expected_color=UtillityMethods.color_picker(self, color, 'rgb')
        UtillityMethods.asequal(self, actual_color, expected_color, msg)
    
    def verify_element_color_using_css_property(self, element_css, color, msg, attribute='stroke', element_obj=None):
        '''
        Desc: This function is to verify element color using value of css property example 'stroke'
        '''
        if element_obj == None :
            actual_color = Color.from_string(self.driver.find_element_by_css_selector(element_css).value_of_css_property(attribute)).rgba
        else :
            actual_color = Color.from_string(element_obj.value_of_css_property(attribute)).rgba
        expected_color=UtillityMethods.color_picker(self, color, 'rgba')
        UtillityMethods.asequal(self, expected_color, actual_color, msg)
    
    def get_element_attribute(self, elem, attrib):
        '''
        Desc:- This function is to get a specified attribute value of an element.
        ''' 
        return(elem.get_attribute(attrib))
        
    def get_element_css_propery(self, elem, attrib):
        '''
        Desc:- This function is to get a specified attribute value of an element.
        ''' 
        return(elem.value_of_css_property(attrib))
    
    def marker_enable(self, to_move_location_element):
        self.driver.execute_script("arguments[0].style.display='inline';",to_move_location_element)
        time.sleep(5)
        element_coordinate=core_utilobj.get_web_element_coordinate(self, to_move_location_element)
        x=element_coordinate['x']
        y=element_coordinate['y']
        self.driver.execute_script("arguments[0].style.display='';",to_move_location_element)
        return(x, y)
        
    def verify_web_element_visible(self, web_element):
        '''
        Desc:- This function is for synchronization of object with certain criteria.
        :param wait_time:- 0, 1, 2, 3.. 
        '''
        pass
    
    def verify_object_highlighted(self, elem, checked_class='checked', highlight_color='pattens_blue_1', step_num='1'):
        '''
        Desc:- Verify the object is highlighted either by checking class attribute or by the highlighted background color
        '''
        if checked_class=='checked':
            status=UtillityMethods.get_element_attribute(self, elem, 'class').lower()
            UtillityMethods.asin(self, 'checked', status, "Step"+str(step_num)+": Verify object is highlighted")
        if highlight_color=='pattens_blue_1':
            actual_color = Color.from_string(UtillityMethods.get_element_css_propery(self, elem,'background')).rgba
            expected_color=UtillityMethods.color_picker(self, highlight_color, 'rgba')
            UtillityMethods.asequal(self, actual_color, expected_color, "Step"+str(step_num)+": Verify highlighted object's background color")
    
    def set_text_to_textbox_using_keybord(self, text_string, text_box_css=None, text_box_elem=None,pause_time=1, type_speed=1, click_type='single'):
        '''
        Desc:- This function is used to type the text string to the text box.
        '''
        if text_box_elem!=None:
            textbox_elem = text_box_elem
        elif text_box_css!=None:
            textbox_elem = self.driver.find_element_by_css_selector(text_box_css)
        if click_type =='single': 
            core_utilobj.python_left_click(self, textbox_elem)
        elif click_type == 'double': 
            core_utilobj.python_doubble_click(self, textbox_elem)
        time.sleep(pause_time)
        if sys.platform == 'linux':
            pykeyboard.press_key(pykeyboard.control_key)
            pyautogui.press('a')
            pykeyboard.release_key(pykeyboard.control_key)
            if Global_variables.browser_name in ['edge']:
                pykeyboard.tap_key(pykeyboard.backspace_key)
            else:
                pykeyboard.tap_key(pykeyboard.delete_key)
            time.sleep(pause_time)
            pykeyboard.type_string(str(text_string), interval=int(type_speed))
        else:
            keyboard.press('ctrl')
            keyboard.press('a')
            keyboard.release('a')
            keyboard.release('ctrl')
            time.sleep(pause_time)
            if Global_variables.browser_name in ['edge']:
                keyboard.press('BACKSPACE')
            else:
                keyboard.send('del')
            time.sleep(pause_time)
            keyboard.write(str(text_string), delay=int(type_speed))
        time.sleep(pause_time)
    
    def set_color_in_color_picker_dialog(self, parent_css, color, close_dialog_button=None):
        '''
        Desc:- This function is used to type the text string to the text box.
        '''
        color_picker_dialog_css = parent_css + " div[class*='window-active']"
        close_dialog_buttons={'ok':'#BiColorPickerOkBtn',
                            'apply': '',
                            'cancel':''}
        ok_btn_css = color_picker_dialog_css + " #BiColorPickerOkBtn"
        UtillityMethods.synchronize_with_number_of_element(self, ok_btn_css, 1, 30)
        rgb_color=UtillityMethods.color_picker(self, color)
        color_obj=re.match(r'rgb\((\d+)\,\s(\d+)\,\s(\d+)\)', rgb_color)
        red=self.driver.find_element_by_css_selector(color_picker_dialog_css + " input#RedTextField")
        green=self.driver.find_element_by_css_selector(color_picker_dialog_css + " input#GreenTextField")
        blue=self.driver.find_element_by_css_selector(color_picker_dialog_css + " input#BlueTextField")
        UtillityMethods.set_text_to_textbox_using_keybord(self, color_obj.group(1), text_box_elem=red)
        UtillityMethods.set_text_to_textbox_using_keybord(self, color_obj.group(2), text_box_elem=green)
        UtillityMethods.set_text_to_textbox_using_keybord(self, color_obj.group(3), text_box_elem=blue)
        if close_dialog_button != None:
            self.driver.find_element_by_css_selector(close_dialog_buttons[close_dialog_button]).click()
        time.sleep(2)
    
    def select_combo_box_item(self, combobox_item, combobox_dropdown_elem=None, combobox_dropdown_css=None):
        '''
        Desc:- This function is used to select an item from the combobox. 
        '''
        if combobox_dropdown_elem!=None:
            combodropdown_elem = combobox_dropdown_elem
        elif combobox_dropdown_css!=None:
            combodropdown_elem = self.driver.find_element_by_css_selector(combobox_dropdown_css)
        core_utilobj.left_click(self, combodropdown_elem)
        time.sleep(1)
        menu_items=self.driver.find_elements_by_css_selector("div[id^='BiPopup'][style*='inherit'] div[id^='BiComboBoxItem']")
        actual_popup_list=[el.text.strip() for el in menu_items]
        menu_items[actual_popup_list.index(combobox_item)].click()
        time.sleep(1)
    
    def verify_combo_box_item(self, expected_combobox_list, combobox_dropdown_elem=None, combobox_dropdown_css=None, msg='Step X'):
        '''
        Desc:- This function is used to verify the items in combobox. 
        '''
        if combobox_dropdown_elem!=None:
            combodropdown_elem = combobox_dropdown_elem
        elif combobox_dropdown_css!=None:
            combodropdown_elem = self.driver.find_element_by_css_selector(combobox_dropdown_css)
        core_utilobj.left_click(self, combodropdown_elem)
        time.sleep(1)
        custom_msg=msg+": Verify the combo box item list."
        menu_items=self.driver.find_elements_by_css_selector("div[id^='BiPopup'][style*='inherit'] div[id^='BiComboBoxItem']")
        actual_popup_list=[el.text.strip() for el in menu_items  if bool(re.match('\S+', el.text.strip()))]
        UtillityMethods.as_List_equal(self, expected_combobox_list, actual_popup_list, custom_msg)
        
    def select_bipopup_list_item(self, list_item_name, bipopup_item_css="table tr", action_chain_click=False,**kwargs):
        '''
        Desc:- This function is for selecting a single item from a BiPopup list items.
        :param list_item_name:- 'Add to Query' or 'Delete' or 'Horizontal Axis' 
        :param bipopup_item_css:- need to pass the css if there is any other type of css
        ''' 
        bipopups=self.driver.find_elements_by_css_selector("div[id^='BiPopup'][style*='inherit']")
        menu_items=bipopups[len(bipopups)-1].find_elements_by_css_selector(bipopup_item_css)
        if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
            core_utilobj.python_move_to_element(self, bipopups[-1], element_location='top_middle', yoffset=3)
        actual_popup_list=[el.text.strip() for el in menu_items]
        if 'element_location' in kwargs:
            location=kwargs['element_location']
        else:
            location='middle_right'
        if Global_variables.browser_name in ['ie', 'edge']:
            core_utilobj.python_left_click(self, menu_items[actual_popup_list.index(list_item_name)], element_location=location, xoffset=-9)
        else:
            core_utilobj.left_click(self, menu_items[actual_popup_list.index(list_item_name)], element_location=location, xoffset=-9, action_chain_click=action_chain_click)
            
    def verify_bipopup_list_item(self, expected_popup_list, msg, verify_type='all', bipopup_item_css="table tr", comparison_value='asequal'):
        '''
        Desc:- This function is for verifying items from a BiPopup list.
        :param expected_popup_list:- need to pass the list to be verified
        :param msg:- need to pass the Step no [e.g.. Step 04a]
        :param verify_type:- 'all' or 'ticked' [by default - all]
        :param bipopup_item_css:- need to pass the css if there is any other type of css [by default - it is table tr]
        '''
        bipopups=self.driver.find_elements_by_css_selector("div[id^='BiPopup'][style*='inherit']")
        menu_items=bipopups[len(bipopups)-1].find_elements_by_css_selector(bipopup_item_css)
        if verify_type == 'all':
            actual_popup_list=[el.text.strip() for el in menu_items  if bool(re.match('\S+', el.text.strip())) and el.is_displayed()] #el.is_displayed() - checking whether element is visible. Because, Edge driver return invisible element text too
            UtillityMethods.verify_list_values(self, expected_popup_list, actual_popup_list, msg + ": Verify all items in in BiPopup.", assert_type=comparison_value)
        elif verify_type == 'ticked':
            actual_ticked_list=[el.text.strip() for el in menu_items if bool(re.match('.*checked$', el.find_element_by_css_selector("td:nth-child(1)").get_attribute("class")))]
            UtillityMethods.verify_list_values(self, expected_popup_list, actual_ticked_list, msg + ": Verify ticked or selected items in in BiPopup.", assert_type=comparison_value)
        else:
            print("Verify_type mismatched")
    
    def verify_element_visiblty(self, element=None, element_css=None, visible=True, msg='Step X'):
        '''
        Desc:- This function is to verify whether the element is visible.
        '''
        try:
            if element!=None:
                status = element.is_displayed()
            elif element_css!=None:
                status = self.driver.find_element_by_css_selector(element_css).is_displayed()
        except NoSuchElementException:
            status = False
        UtillityMethods.asequal(self, status, visible, msg)
    
    def verify_element_disable(self, element=None, attribute_='class'):
        '''
        Desc:- This function is to verify whether the element is disabled.
        '''
        if element is None:
            raise AttributeError("Element is None.")
        try:
            element_class_value = UtillityMethods.get_element_attribute(self, element, attribute_)
            temp_value = element_class_value[0]
        except TypeError:
            del temp_value    
            raise TypeError("This element of '"+str(attribute_)+"' value is null.")
        status=bool(re.match('.*button-disabled', element_class_value))
        return (status)
    
    def click_any_bibutton_in_dialog(self, dialog_css, btn_name):
        '''
        Desc: This is used to click on any button whose ID start with 'BiButton'.
        :param: dialog: "div[id^='BiDialog']"
        :param: btn_name: "OK"
        :Usage: utillobj.click_dialog_button("div[id^='BiDialog']", "OK")
        '''        
        dialog_css = dialog_css + " div[id^='BiButton']"
        dialog_btns=self.driver.find_elements_by_css_selector(dialog_css)
        btn_text_list=[el.text.strip() for el in dialog_btns]
        dialog_btns[btn_text_list.index(btn_name)].click()
        time.sleep(5)
        
    def verify_dialogs_and_popups_caption(self, caption_css, text_to_verify, msg='Step X'):
        '''
        Desc:- This function is to verify the text written inside the caption.
        '''        
        UtillityMethods.verify_element_visiblty(self, element_css=caption_css, msg='Step X: Verify required Caption element is visible.')
        caption_elem=self.driver.find_element_by_css_selector(caption_css)
        UtillityMethods.asin(self, text_to_verify, caption_elem.text, msg)
        
    def verify_dialogs_and_popups_text(self, text_css, text_to_verify, msg='Step X'):
        '''
        Desc:- This function is to verify the text written inside the popup.
        '''        
        UtillityMethods.verify_element_visiblty(self, element_css=text_css, msg='Step X: Verify required text element is visible in the popup.')
        text_elem=self.driver.find_element_by_css_selector(text_css)
        UtillityMethods.asin(self, text_to_verify, text_elem.text, msg)
            
    def verify_dialogs_and_popups_text_in_element(self, popup_element, text_to_verify, msg='Step X'):
        '''
        Desc:- This function is to verify the text written inside any element within the popup.
        '''        
        UtillityMethods.verify_element_visiblty(self, element=popup_element, msg='Step X: Verify required element is visible in the popup.')
        UtillityMethods.asin(self, text_to_verify, popup_element.text, msg)
    
            
    def take_browser_snapshot(self, file_name, image_type='actual', left=0, right=0, top=0, bottom=0):
        '''
        Desc: This action is used to take browser screenshot, and save to '.png' format.
        :param file_name: file for saving
        :param image_type: where you want to save your image in directory
        :usage take_monitor_screenshot('full_monitor_screenshot', image_type='actual')
        '''
        location='actual_images' if image_type=='actual' else 'failure_captures' if image_type=='fail' else 'images'
        browser_specification = core_utilobj.get_current_browser_specification(self)
        bleft=left
        btop=top if top != 0 else browser_specification['browser_height']
        bottom= bottom if top != 0 else browser_specification['outer_height']
        file_path=os.getcwd() + "\\" + location + "\\" + file_name + ".png"
        im=ImageGrab.grab()
        im.save(file_path)
        resolution=pyautogui.size()
        bright=resolution[0]-right
        bbottom=resolution[1]-bottom
        bbox = (bleft, btop, bright, bbottom)
        base_image = Image.open(file_path)
        cropped_image = base_image.crop(bbox)
        cropped_image.save(file_path)
    
    def take_monitor_snapshot(self, file_name, image_type='actual', left=0, right=0, top=0, bottom=0):
        """
        :param file_name: file for saving
        :param image_type: where you want to save your image in directory
        :param left: how much you want to reduce the size from left in output of your image
        :param top: how much you want to reduce the size from top in output of your image
        :param right: how much you want to reduce the size from right in output of your image
        :param bottom: how much you want to reduce the size from bottom in output of your image
        :usage take_monitor_screenshot('full_monitor_screenshot', image_type='actual', left=10, top=25, right=10, bottom=25)
        """  
        location='actual_images' if image_type=='actual' else 'failure_captures' if image_type=='fail' else 'images'      
        file_path=os.getcwd() + "\\" + location + "\\" + file_name + ".png"
        im=ImageGrab.grab()
        im.save(file_path)
        resolution=pyautogui.size()
        bleft = left
        btop = top
        bright=resolution[0]-right
        bbottom=resolution[1]-bottom
        bbox = (bleft, btop, bright, bbottom)
        base_image = Image.open(file_path)
        cropped_image = base_image.crop(bbox)
        cropped_image.save(file_path)
          
    def take_snapshot(self, element, file_name, image_type='actual', left=0, right=0, top=0, bottom=0):
        x=Global_variables.current_working_area_browser_x
        y=Global_variables.current_working_area_browser_y
        core_utilobj.move_mouse_to_offset(self, x_offset=x, y_offset=y)
        time.sleep(1)
        location='actual_images' if image_type=='actual' else 'images'
        file_path=os.getcwd() + "\\" + location + "\\" + file_name + ".png"
        bounding_box = (
            element.location['x'] + left, 
            element.location['y'] + top,  
            (element.location['x'] + element.size['width']) + right,
            (element.location['y'] + element.size['height']) + bottom  
        )
        warnings.simplefilter("ignore", ResourceWarning)
        return UtillityMethods.bounding_box_snapshot(self, bounding_box, file_path)

    def bounding_box_snapshot(self, bounding_box, filename):
        self.driver.save_screenshot(filename)
        base_image = Image.open(filename)
        cropped_image = base_image.crop(bounding_box)
        base_image = base_image.resize(cropped_image.size)
        base_image.paste(cropped_image, (0, 0))
        base_image.save(filename)
        return base_image  
    
    def beautifulsoup_object_creation(self):
        from bs4 import BeautifulSoup
        page_source_text=self.driver.page_source
        soup = BeautifulSoup(page_source_text, 'html.parser')
        return (soup)
    
    def verify_table_data(self, table_css, file_name, number_of_rows=None):
        '''
        This function is used to verify run time data set like table structure.
        :param table_css="table[summary='Summary'] > tbody > tr"
        :param: file_name="test1.xlsx"
        :Usage: verify_table_data("table[summary='Summary']", "test1.xlsx")
        :@author AAkhan
        '''
        soup = UtillityMethods.beautifulsoup_object_creation(self)
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = soup.select(table_css)
        if number_of_rows is not None:
            table_row_lenght = int(number_of_rows)
        else:
            table_row_lenght = len(table_rows)
        for r in range(0, table_row_lenght):
            total_col = table_rows[r].find_all('td')
            for c in range(len(total_col)):
                browser_value=str(total_col[c].get_text(strip=True).replace(' ',''))
                if s1.cell(row=r + 1, column=c + 1).value != None and s1.cell(row=r + 1, column=c + 1).value != ' ':
                    excel_value=s1.cell(row=r + 1, column=c + 1).value.replace(' ','')
                else:
                    excel_value=None
                if (excel_value != None and browser_value != '') or (excel_value != None) or (browser_value != ''):
                    if excel_value == browser_value:
                        status=[0]
                        continue
                    else:
                        status=[r+1, c+1, 'expected_value', excel_value, 'actual_value', browser_value]
                        return (status)
        return (status)
    
    def create_table_data(self, table_css, file_name):
        '''
        This function is used to create run time data set like table structure.
        :param table_css="table[summary='Summary'] > tbody > tr"
        :param: file_name="test1.xlsx"
        :Usage: create_table_data("table[summary='Summary']", "test1.xlsx")
        :@author AAkhan
        '''
        soup = UtillityMethods.beautifulsoup_object_creation(self)
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = soup.select(table_css)
        for r in range(0, len(table_rows)):
            total_col = table_rows[r].find_all('td')
            for c in range(len(total_col)):
                s.cell(row=r + 1, column=c + 1).value = total_col[c].get_text(strip=True)
        wb.save(os.getcwd() + "\data\\" + file_name)
        
    def create_table_data_start_end_rowcolumn(self, file_name, table_css="table[summary='Summary']", start_rownum=None, end_rownum=None, start_colnum=None, end_colnum=None):
        '''
        Desc This function is used to create run time data set using starting row, end , starting column and end column.
        :Usage: create_table_data("test1.xlsx", "table[summary='Summary']", start_rownum=3, end_rownum=10, start_colnum=1, end_colnum=5)
        '''
        soup = UtillityMethods.beautifulsoup_object_creation(self)
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = soup.select(table_css +' > tbody > tr')
        start_row_num=start_rownum if start_rownum!=None else 0
        no_of_rows=end_rownum if end_rownum!=None else len(table_rows)
        for r in range(start_row_num, no_of_rows):
            total_col = table_rows[r].find_all('td')
            start_col_num=start_colnum if start_colnum!=None else 0
            no_of_cols=end_colnum if end_colnum!=None else len(total_col)
            for c in range(start_col_num, no_of_cols):
                s.cell(row=r + 1, column=c + 1).value = total_col[c].get_text(strip=True)
        wb.save(os.getcwd() + "\data\\" + file_name)
        
    def compare_table_data_using_start_end_rowcolumn(self, file_name, table_css, start_rownum=None, end_rownum=None, start_colnum=None, end_colnum=None):
        '''
        Desc :This function is used to verify run time data set like table structure.
        :Usage: verify_table_data("test1.xlsx", "Stepx X:Verify report data", table[summary='Summary']", start_rownum=2, end_rownum=10, start_colnum=2, end_colnum=5)
        '''
        soup = UtillityMethods.beautifulsoup_object_creation(self)
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = soup.select(table_css+' > tbody > tr')
        start_row_num=start_rownum if start_rownum!=None else 0
        no_of_rows = int(end_rownum) if end_rownum is not None else len(table_rows)
        for r in range(start_row_num, no_of_rows):
            total_col = table_rows[r].find_all('td')
            start_col_num=start_colnum if start_colnum!=None else 0
            no_of_cols=end_colnum if end_colnum!=None else len(total_col)
            for c in range(start_col_num, no_of_cols):
                browser_value=str(total_col[c].get_text(strip=True).replace(' ',''))
                if s1.cell(row=r + 1, column=c + 1).value != None and s1.cell(row=r + 1, column=c + 1).value != ' ':
                    excel_value=s1.cell(row=r + 1, column=c + 1).value.replace(' ','')
                else:
                    excel_value=None
                if (excel_value != None and browser_value != '') or (excel_value != None) or (browser_value != ''):
                    if excel_value == browser_value:
                        status=[0]
                        continue
                    else:
                        status=[r+1,c+1]
                        return (status)
        return (status)
    
    def verify_table_data_using_start_end_rowcolumn(self, file_name, table_css, msg, start_rownum=None, end_rownum=None, start_colnum=None, end_colnum=None):
        '''
        Desc : Verify the compared table data using start and end rowcolumn
        :Usage: verify_table_data_using_start_end_rowcolumn
        '''
        x=UtillityMethods.compare_table_data_using_start_end_rowcolumn(self, file_name, table_css, start_rownum=start_rownum, end_rownum=end_rownum, start_colnum=start_colnum, end_colnum=end_colnum)
        UtillityMethods.asequal(self, len(x), 1, msg+ ' Row,Column -'+ str(x)) 
    
    def scroll_down_to_search_element(self, item_name, tabel_parent_css, tabel_row_css, scroll_css="div.bi-tree-view-body", elem_location=0):
        '''
        Desc:- Search element using java-script scroll attribute.
        :param item_name='What element need to search.'
        :param tabel_parent_css='Scroll parent element css-value.'
        :param tabel_row_css='element row css-value.'
        :param scroll_css='Actual scroll css-value.'
        :param elem_location='current element location need to scroll inside java-script.'
        :usage  scroll_down_to_search_element('Page_designer1', "#paneIbfsExplorer_exList", "#paneIbfsExplorer_exList > div.bi-tree-view-body-content > table > tbody > tr")
        '''
        tabel_scroll_css = tabel_parent_css + " " + scroll_css
        run_=True
        old_list=[]
        elem_location = elem_location
        scroll_elem=self.driver.find_element_by_css_selector(tabel_scroll_css)
        while run_:
            elems = self.driver.find_elements_by_css_selector(tabel_row_css)
            current_list = [elem.text.strip() for elem in elems if elem.text.strip() != '']
            if old_list == current_list:
                run_ = False
                break
            if item_name not in current_list:
                elem_location += self.driver.execute_script("return arguments[0].offsetTop;", elems[len(current_list)-1])
                self.driver.execute_script("return arguments[0].scrollTop="+str(elem_location), scroll_elem)
                time.sleep(1)
            else:
                run_ = False
                break
            old_list=current_list.copy()
        return (elem_location)
        
    def multiselect_item_using_scroll_operation(self, parent_css, item_name_list):
        """
        Desc:-Select file from IBFS explorer dialog.
        :param file_name_list='list of Files need to select.'
        :param file_type='File type need to select.'
        :Usage select_file_from_ibfs_explorer(['Page_designer1'])
        """
        table_row_css = parent_css+" > div.bi-tree-view-body-content > table > tbody > tr"
        selected_element_css = table_row_css+ "[class*='selected']"
        keyboard.press('ctrl')
        elem_location=0
        for name in item_name_list:
            elem_location += UtillityMethods.scroll_down_to_search_element(self, name, parent_css, table_row_css, elem_location=elem_location)
            time.sleep(1)
            try:
                elems = self.driver.find_elements_by_css_selector(table_row_css)
                required_tr_elem = elems[[name in elem for elem in [elem.text.strip() for elem in elems if elem.text.strip() != '']].index(True)]
            except ValueError:
                raise ValueError("File name '" + name + "' not exist.")
            core_utilobj.python_left_click(self, required_tr_elem, element_location='bottom_left', xoffset=10)
            UtillityMethods.synchronize_with_visble_text(self, selected_element_css, name.replace(' ',''), 10)
        keyboard.release('ctrl')
        
    def scroll_and_select_combobox_item_(self, combobox_btn_css, combobox_item, **kwargs):
        """
        Desc: scroll and select combo box item
        This method used to scroll and click on combobox item.
        Example Usage : scroll_and_select_combobox_item(self, "#dlgSecurityUserEdit #SecurityUserEditDialog_cbDefGroup div[id^='BiButton']", 'P116')
        """
        Error_Msg="[{0}] item not listed in comboxbox".format(combobox_item)
        combobox_items_css="div[id^='BiPopup'][style*='inherit'] div[id^='BiComboBoxItem']"
        combobox_scrollable_css="div[id^='BiPopup'][style*='inherit']>div[class*='bi-combo-box-list']"
        combo_btn=self.driver.find_element_by_css_selector(combobox_btn_css)
        if Global_variables.browser_name == 'firefox':
            action1 = ActionChains(self.driver)
            action1.move_to_element(combo_btn).click(combo_btn).perform()
            del action1
        else:
            core_utilobj.left_click(self, combo_btn)
        UtillityMethods.synchronize_with_number_of_element(self, "div[id^='BiPopup'][style*='inherit']", 1, expire_time=4)
        combobox_items_obj=self.driver.find_elements_by_css_selector(combobox_items_css)
        combobox_item_index=JavaScript.find_element_index_by_text(self, combobox_items_obj, combobox_item)
        if combobox_item_index == None :
            raise NoSuchElementException(Error_Msg)
        combobox_item_obj = combobox_items_obj[combobox_item_index]
        scroll_offset=JavaScript.get_scroll_offsetTop(self, combobox_item_obj)
        JavaScript.scroll_element(self, combobox_scrollable_css, scroll_offset-20, wait_time=1)
        combobox_item_obj.click()
      
    
    """************************OLD FUNCTIONS*********************************************"""
            
    def login_wf(self, mrid, mrpass, sync_css = None, add_home_path=True):#Need to Delete
        '''
        This function is used to login in webfocus
        '''
        loginid = UtillityMethods.parseinitfile(self,mrid)
        loginpwd = UtillityMethods.parseinitfile(self,mrpass)
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url ='http://' + node + ':' + port + context + '/home8206' if add_home_path else 'http://' + node + ':' + port + context + '/'
        refused_to_connect_msg = "refused to connect"
        cannot_reach_this_page = " reach this page"
        api_url_response_text= self.driver.find_element_by_css_selector("body").text.replace(' ', '').lower()
        if refused_to_connect_msg.replace(' ', '') in api_url_response_text:
            raise EnvironmentError("WebFOCUS Client {0} is down the reason is '{1}'.".format(setup_url, refused_to_connect_msg))
        if cannot_reach_this_page.replace(' ', '') in api_url_response_text:
            raise EnvironmentError("WebFOCUS Client {0} is down the reason is 'Cannot {1}'.".format(setup_url, cannot_reach_this_page))
        UtillityMethods.synchronize_with_number_of_element(self, 'input[id=SignonUserName]', 1, 60)
        usename= self.driver.find_element(*LoginPageLocators.uname)
        usename.click()
        time.sleep(2)
        usename.send_keys(loginid)
        time.sleep(1)
        if loginpwd!=None:
            passwd =self.driver.find_element(*LoginPageLocators.pword)
            passwd.send_keys(loginpwd)
            time.sleep(1)
        sign_in =self.driver.find_element(*LoginPageLocators.submit)
        sign_in.click()
        time.sleep(5)
        UtillityMethods.synchronize_with_number_of_element(self, 'body', 1, 180)
        invalid_username_password_msg = "Invalidusernameorpassword"
        html_body_text= self.driver.find_element_by_css_selector("body").text.replace(' ', '')
        if invalid_username_password_msg in html_body_text:
            raise NameError("Invalid user name or password error thrown for the user -{0}".format(loginid))
        access_to_item_denied_msg = "Accesstoitemdenied"
        if access_to_item_denied_msg in html_body_text:
            raise UserWarning("Access to item denied- Project Folder or Master file doesn't exist")
        #if sync_css != None :
        #    UtillityMethods.synchronize_with_number_of_element(self, sync_css, 1, 120)
        error_hint="Itemdoesnotexist"
        resource_not_found_err_msg="Received Item Does not Exist. Means Either  Repository folder OR the required fex is not available in the set up."
        if error_hint in html_body_text:
            raise LookupError(resource_not_found_err_msg)
        access_to_item_denied_msg = "Accesstoitemdenied"
        project_not_found_err_msg="Received Access to item denied Means project_id is not available"
        if access_to_item_denied_msg in html_body_text:
            raise LookupError(project_not_found_err_msg)
        core_utilobj.update_current_working_area_browser_specification(self)
        Global_variables.used_username_in_current_session[loginid] = loginpwd
        
    def infoassist_api_login(self, tool, master, folder, mrid, mrpass):#Need to Delete
        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        project = UtillityMethods.parseinitfile(self, 'project_id')
        suite = UtillityMethods.parseinitfile(self, 'suite_id')
        folder = project + '/' + suite
        api_url = setup_url + 'ia?tool=' + tool + '&master=' + master + '&item=IBFS%3A%2FWFC%2FRepository%2F' + folder
        self.driver.get(api_url.replace('home8206',''))
        UtillityMethods.login_wf(self, mrid,mrpass, sync_css = "#resultArea", add_home_path=False)
        time.sleep(30)
        file_not_found_msg = "THEDESCRIPTIONCANNOTBEFOUNDFORFILE"
        file_not_found_text =self.driver.find_element_by_css_selector("html body").text.replace(' ', '')
        if file_not_found_msg in file_not_found_text:
                raise UserWarning("Master file {0} doesn't exist in the specified folder".format(master))
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def infoassist_api_login_(self, tool, master, mrid, mrpass):#Need to Delete
        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        project = UtillityMethods.parseinitfile(self, 'project_id')
        suite = UtillityMethods.parseinitfile(self, 'suite_id')
        group_id=UtillityMethods.parseinitfile(self, 'group_id')
        folder = project + '_' + suite + '_' + group_id
        api_url = setup_url + 'ia?tool=' + tool + '&master=' + master + '&item=IBFS%3A%2FWFC%2FRepository%2F' + folder
        self.driver.get(api_url.replace('home8206',''))
        UtillityMethods.login_wf(self, mrid,mrpass, sync_css = "#resultArea", add_home_path=False)
        time.sleep(30)
        file_not_found_msg = "THEDESCRIPTIONCANNOTBEFOUNDFORFILE"
        file_not_found_text =self.driver.find_element_by_css_selector("html body").text.replace(' ', '')
        if file_not_found_msg in file_not_found_text:
                raise UserWarning("Master file {0} doesn't exist in the specified folder".format(master))
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def invoke_infoassist_api_login(self, tool, master, folder, mrid, mrpass):#Need to Delete
        """
        :param tool: report, document
        :param folder: 'P292_S10863/G408271' 
        :Usage invoke_infoassist_api_login('report','baseapp/dimensions/wf_retail_product','P292_S10863/G408271', 'mrid', 'mrpass')
        """
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        project = UtillityMethods.parseinitfile(self, 'project_id')
        suite = UtillityMethods.parseinitfile(self, 'suite_id')
        group_id=UtillityMethods.parseinitfile(self, 'group_id')
        folder = project + '_' + suite + '/' + group_id
        api_url = setup_url + 'ia?tool=' + tool + '&master=' + master + '&item=IBFS%3A%2FWFC%2FRepository%2F' + folder
        self.driver.get(api_url.replace('home8206',''))
        UtillityMethods.login_wf(self, mrid,mrpass, sync_css = "#resultArea", add_home_path=False)
        time.sleep(30)
        file_not_found_msg = "THEDESCRIPTIONCANNOTBEFOUNDFORFILE"
        file_not_found_text =self.driver.find_element_by_css_selector("html body").text.replace(' ', '')
        if file_not_found_msg in file_not_found_text:
                raise UserWarning("Master file {0} doesn't exist in the specified folder".format(master))
        core_utilobj.update_current_working_area_browser_specification(self)
    
    def invoke_ia_tool_without_master_file_using_api(self, tool, mrid='mrid', mrpass='mrpass', folder=None, login=True):
        """
        Description : Invoke IA tool(chart, report) without master file using api url
        :param : login = True if login is need to invoke api else login=False
        :Usage : invoke_ia_tool_without_master_file_using_api('chart')
        """
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        if folder == None :
            project = UtillityMethods.parseinitfile(self, 'project_id')
            suite = UtillityMethods.parseinitfile(self, 'suite_id')
            group_id=UtillityMethods.parseinitfile(self, 'group_id')
            folder = project + '_' + suite + '/' + group_id
        api_url = setup_url + 'ia?tool=' + tool + '&item=IBFS:/WFC/Repository/' + folder
        self.driver.get(api_url.replace('home8206',''))
        if login == True :
            UtillityMethods.login_wf(self, mrid,mrpass, sync_css = "#resultArea", add_home_path=False)
        UtillityMethods.wait_for_page_loads(self, 60)
        file_not_found_msg = "accesstoitemdenied"
        file_not_found_text =self.driver.find_element_by_css_selector("html body").text.replace(' ', '').lower()
        if file_not_found_msg in file_not_found_text:
                raise UserWarning(file_not_found_text)
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def infoassist_api_edit(self, fex, tool, folder, **kwargs):#Need to Delete
        """
        :param fex:  AR-RP-001a.fex
        :param tool: report, chart, 
        :param folder: S7072 
        :param kwargs['mrid']='mrid'
        :param kwargs['mrpass']='mrpass'
        :Usage infoassist_api_edit('C2222586', 'report', 'S10006', mrid='mrid', mrpass='mrpass')
        __author = Gobizen 
        """
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        project_id=UtillityMethods.parseinitfile(self, 'project_id')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        folder = UtillityMethods.parseinitfile(self, 'suite_id')
        api_url = setup_url+ 'ia?&item=IBFS%3A%2FWFC%2FRepository/'+project_id+'/'+folder+'/'+fex+'.fex&tool='+tool
        self.driver.get(api_url.replace('home8206',''))
        if 'mrid' in kwargs:
            UtillityMethods.login_wf(self, kwargs['mrid'], kwargs['mrpass'], sync_css = "#resultArea", add_home_path=False)
        time.sleep(30)
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def infoassist_api_edit_(self, fex, tool, folder, **kwargs):#Need to Delete
        """
        :param fex:  C2222586
        :param tool: report, chart, 
        :param folder: P292_S10863/G408271 
        :Usage infoassist_api_edit('C2222586', 'report', 'P292_S10863/G408271', mrid='mrid', mrpass='mrpass')
        """
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        project_id = UtillityMethods.parseinitfile(self, 'project_id')
        suite = UtillityMethods.parseinitfile(self, 'suite_id')
        group_id=UtillityMethods.parseinitfile(self, 'group_id')
        folder = project_id + '_' + suite + '/' + group_id
        api_url = setup_url+ 'ia?&item=IBFS%3A%2FWFC%2FRepository/'+folder+'/'+fex+'.fex&tool='+tool
        self.driver.get(api_url.replace('home8206',''))
        if 'mrid' in kwargs:
            UtillityMethods.login_wf(self, kwargs['mrid'], kwargs['mrpass'], sync_css = "#resultArea", add_home_path=False)
        time.sleep(30)
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def infoassist_api_login_with_masterfile_promt(self, tool, mrid, mrpass):
        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        project = UtillityMethods.parseinitfile(self, 'project_id')
        suite = UtillityMethods.parseinitfile(self, 'suite_id')
        folder = project + '/' + suite
        api_url = setup_url + 'ia?tool=' + tool + '&item=IBFS%3A%2FWFC%2FRepository%2F' + folder
        self.driver.get(api_url.replace('home8206',''))
        UtillityMethods.login_wf(self, mrid, mrpass, add_home_path=False)
        time.sleep(30)
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def active_run_fex_api_login(self, fex, folder, mrid, mrpass):
        """
        Usage: active_run_fex_api_login('Active_report.fex','S7067','mrid','mrpass')
        """        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        project_id=UtillityMethods.parseinitfile(self, 'project_id')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        self.driver.get(setup_url)
        UtillityMethods.login_wf(self, mrid, mrpass, sync_css="div[class*='left-main-panel-content-button']")        
        api_url = setup_url.replace('home8206','') + 'run.bip?BIP_REQUEST_TYPE=BIP_RUN&BIP_folder=IBFS%253A%252FWFC%252FRepository%252F' + project_id + "/" + folder + '&BIP_item=' + fex
        self.driver.get(api_url)
        time.sleep(30)
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def active_run_fex_api_login_(self, fex, folder, mrid, mrpass):
        """
        Usage: active_run_fex_api_login_('Active_report.fex','G427910','mrid','mrpass')
        """        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        project_id = UtillityMethods.parseinitfile(self, 'project_id')
        suite = UtillityMethods.parseinitfile(self, 'suite_id')
        folder = UtillityMethods.parseinitfile(self, 'group_id')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        self.driver.get(setup_url)
        UtillityMethods.login_wf(self, mrid,mrpass, sync_css="div[class*='left-main-panel-content-button']")     
        api_url = setup_url + 'run.bip?BIP_REQUEST_TYPE=BIP_RUN&BIP_folder=IBFS%253A%252FWFC%252FRepository%252F' + project_id + "_" + suite + "/" + folder + '&BIP_item=' + fex
        self.driver.get(api_url.replace('home8206',''))
        time.sleep(30)
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def run_htm_api_login(self, html, mrid, mrpass):
        """
        Usage: run_htm_api_login('C12345.htm','Public','mrid','mrpass')
        """   
        from common.pages import wf_mainpage   
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        self.driver.get(setup_url)
        UtillityMethods.login_wf(self, mrid, mrpass) 
        time.sleep(2)
        project = UtillityMethods.parseinitfile(self, 'project_id')
        suite = UtillityMethods.parseinitfile(self, 'suite_id')
        folder = project + '/' + suite
        p=project + '->' + suite + '->' + html 
        wf_mainpage_obj= wf_mainpage.Wf_Mainpage(self.driver)
        if wf_mainpage_obj.get_repository_item_availability(p):
            api_url = setup_url + 'run.bip?BIP_REQUEST_TYPE=BIP_RUN&BIP_folder=IBFS%253A%252FWFC%252FRepository%252F' + folder + '&BIP_item=' + html+'.htm'
            self.driver.get(api_url.replace('home8206',''))
            time.sleep(30)
        else:
            unittest_obj= unittest.TestCase(self.driver)
            unittest_obj.fail("Html File does not exists")
            Global_variables.asert_failure_count += 1
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def run_reporting_server_api_login(self, html, folder, mrid, mrpass):
        """
        Usage: run_htm_api_login('C12345.htm','Public','mrid','mrpass')
        Author: Jesmin
        """        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        self.driver.get(setup_url)
        UtillityMethods.login_wf(self, mrid, mrpass)        
        api_url = setup_url + 'run.bip?BIP_REQUEST_TYPE=BIP_RUN&BIP_folder=IBFS:/EDA/EDASERVE/' + folder + '&BIP_item=' + html
        self.driver.get(api_url.replace('home8206',''))
        time.sleep(30)
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def invoke_bip_portal(self, portal_name, mrid, mrpass, mode='run'):
        """
        Usage: invoke_bip_portal('BIP_xxx_Portal123_V4', 'mrid', 'mrpass','edit')
        """        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        time.sleep(2)
        project = UtillityMethods.parseinitfile(self, 'project_id')
        suite = UtillityMethods.parseinitfile(self, 'suite_id')
        folder = project + '/' + suite
        setup_url = 'http://' + node + ':' + port + context
        if mode=='run' :
            api_url = setup_url + '/portal/'+folder+'/'+portal_name
        elif mode=='edit' :
            api_url = setup_url + '/portal/PortalDesigner?path='+folder+'/'+portal_name
        else:
            raise ("Wrong url link call")
        self.driver.get(api_url)
        UtillityMethods.login_wf(self, mrid, mrpass)        
        time.sleep(30) 
        core_utilobj.update_current_working_area_browser_specification(self)
    
    def invoke_webfocu(self, mrid, mrpass):
        """
        Usage: invoke_webfocu(self, 'mrid', 'mrpass')
        This function invoke wf welcome page
        """
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        self.driver.get(setup_url)
        UtillityMethods.login_wf(self, mrid, mrpass)
    
    def invoke_legacyhomepage(self, mrid, mrpass):
        """
        Usage: invoke_legacyhomepage(self, 'mrid', 'mrpass')
        This function invoke wf legacyhomepage page
        """
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home8206'
        self.driver.get(setup_url)
        UtillityMethods.login_wf(self, mrid, mrpass)
        time.sleep(3)
        UtillityMethods.navigate_to_legacyhomepage(self)
        
    def navigate_to_legacyhomepage(self):
        current_setup = UtillityMethods.get_setup_url(self)
        current_setup = current_setup.replace('home8206','')
        setup_url = '{0}legacyhome'.format(current_setup)
        self.driver.get(setup_url)
        core_utilobj.update_current_working_area_browser_specification(self)
        
    def infoassist_api_logout(self):
        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/'
        logout = lambda : self.driver.get("" + setup_url + "service/wf_security_logout.jsp ") 
        if Global_variables.browser_name == 'edge' :
            click_func = lambda :  UtillityMethods.click_on_edge_browser_popup_button_using_uiautomation(self)
            logout_process = Thread(target=logout)
            click_process = Thread(target=click_func)
            logout_process.start()
            click_process.start()
            click_process.join()
            logout_process.join()
        else :
            logout()
            time.sleep(2)
            try:
                als = alert.Alert(self.driver)
                als.accept()
                if Global_variables.saved_case_flag:
                    raise AttributeError('Alert window found')
            except NoAlertPresentException:
                pass
        
    def create_failure_log(self, msg):
        try:
            Global_variables.verification_failure_msg_list.append(msg)
#             fo=open(os.getcwd()+"\\failure_captures\\"+UtillityMethods.current_test_case+".log", 'a', encoding="utf-8")
#             fo.write(msg+"\n");
#             fo.close()
        except:
#             print("Unable to append in failure capture message '{0}' to log file",format(msg))
            print("Unable to append in failure capture message '{0}' to verification_failure_msg_list",format(msg))
    
    def as_List_equal(self,*params):
        Flag=False
        try:
            testobj = unittest.TestCase()
            testobj.assertListEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
            Global_variables.verification_steps.append(params[2] + " - PASSED.")
            Flag=True
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            msg = m.group(1)+" - FAILED."+'List Equal comparison : value1=['+str(params[0])+'], value2=['+str(params[1])+']'
            print(msg)
            Global_variables.verification_steps.append(msg)
            Global_variables.asert_failure_count += 1
            UtillityMethods.create_failure_log(self, msg)
        finally:
            suite_type=UtillityMethods.parseinitfile(self, 'suite_type')
            try:
                try:
                    step_number=re.search(r'\d+.\d+', params[2]).group()
                except AttributeError:
                    step_number=re.search(r'\d+', params[2]).group()
            except AttributeError:
                step_number=re.search(r'[x|X]', params[2]).group()
            if suite_type == 'smoke':
                file_name=Global_variables.current_test_case
                file_path=os.getcwd() + "\\" + file_name + "_" + step_number + ".png"
                self.driver.save_screenshot(file_path)
            else:
                pass
        UtillityMethods.verification_screenshot_capture(self, step_number, Flag)
    
    def asequal(self, *params):
        Flag=False
        try:
            testobj = unittest.TestCase()
            testobj.assertEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
            Global_variables.verification_steps.append(params[2] + " - PASSED.")
            Flag=True
        except AssertionError as e:
            m = re.match(r'.*([S|s]tep.*)...', str(e.args))
            msg = m.group(1)+" - FAILED."+'Equal comparison : value1=['+str(params[0])+'], value2=['+str(params[1])+']'
            print(msg)
            Global_variables.verification_steps.append(msg)
            Global_variables.asert_failure_count += 1
            UtillityMethods.create_failure_log(self, msg)
        finally:
            suite_type=UtillityMethods.parseinitfile(self, 'suite_type')
            try:
                try:
                    step_number=re.search(r'\d+.\d+', params[2]).group()
                except AttributeError:
                    step_number=re.search(r'\d+', params[2]).group()
            except AttributeError:
                step_number=re.search(r'[x|X]', params[2]).group()
            if suite_type == 'smoke':
                file_name=Global_variables.current_test_case
                file_path=os.getcwd() + "\\" + file_name + "_" + step_number + ".png"
                self.driver.save_screenshot(file_path)
            else:
                pass
        UtillityMethods.verification_screenshot_capture(self, step_number, Flag)
        
    def as_not_equal(self, *params):
        Flag=False
        try:
            testobj = unittest.TestCase()
            testobj.assertNotEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
            Global_variables.verification_steps.append(params[2] + " - PASSED.")
            Flag=True
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            msg = m.group(1)+" - FAILED."+'Not Equal comparison : value1=['+str(params[0])+'], value2=['+str(params[1])+']'
            print(msg)
            Global_variables.verification_steps.append(msg)
            Global_variables.asert_failure_count += 1
            UtillityMethods.create_failure_log(self, msg)
        finally:
            suite_type=UtillityMethods.parseinitfile(self, 'suite_type')
            try:
                try:
                    step_number=re.search(r'\d+.\d+', params[2]).group()
                except AttributeError:
                    step_number=re.search(r'\d+', params[2]).group()
            except AttributeError:
                step_number=re.search(r'[x|X]', params[2]).group()
            if suite_type == 'smoke':
                file_name=Global_variables.current_test_case
                file_path=os.getcwd() + "\\" + file_name + "_" + step_number + ".png"
                self.driver.save_screenshot(file_path)
            else:
                pass
        UtillityMethods.verification_screenshot_capture(self, step_number, Flag)
    
    def asin(self, *params):
        Flag=False
        try:
            #testobj = unittest.TestCase()
            assert params[0] in params[1], params[2]
            print(params[2] + " - PASSED.")
            Global_variables.verification_steps.append(params[2] + " - PASSED.")
            Flag=True
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            msg = m.group(1)+" - FAILED."+'ASIN comparison : value1=['+str(params[0])+'], value2=['+str(params[1])+']'
            print(msg)
            Global_variables.verification_steps.append(msg)
            Global_variables.asert_failure_count += 1
            UtillityMethods.create_failure_log(self, msg)
        finally:
            suite_type=UtillityMethods.parseinitfile(self, 'suite_type')
            try:
                try:
                    step_number=re.search(r'\d+.\d+', params[2]).group()
                except AttributeError:
                    step_number=re.search(r'\d+', params[2]).group()
            except AttributeError:
                step_number=re.search(r'[x|X]', params[2]).group()
            if suite_type == 'smoke':
                file_name=Global_variables.current_test_case
                file_path=os.getcwd() + "\\" + file_name + "_" + step_number + ".png"
                self.driver.save_screenshot(file_path)
            else:
                pass
        UtillityMethods.verification_screenshot_capture(self, step_number, Flag)
            
    def as_notin(self, *params):
        Flag=False
        try:
            #testobj = unittest.TestCase()
            assert params[0] not in params[1], params[2]
            print(params[2] + " - PASSED.")
            Global_variables.verification_steps.append(params[2] + " - PASSED.")
            Flag=True
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            msg = m.group(1)+" - FAILED."+'Equal comparison : value1=['+str(params[0])+'], value2=['+str(params[1])+']'
            print(msg)
            Global_variables.verification_steps.append(msg)
            Global_variables.asert_failure_count += 1
            UtillityMethods.create_failure_log(self, msg)
        finally:
            suite_type=UtillityMethods.parseinitfile(self, 'suite_type')
            try:
                try:
                    step_number=re.search(r'\d+.\d+', params[2]).group()
                except AttributeError:
                    step_number=re.search(r'\d+', params[2]).group()
            except AttributeError:
                step_number=re.search(r'[x|X]', params[2]).group()
            if suite_type == 'smoke':
                file_name=Global_variables.current_test_case
                file_path=os.getcwd() + "\\" + file_name + "_" + step_number + ".png"
                self.driver.save_screenshot(file_path)
            else:
                pass
        UtillityMethods.verification_screenshot_capture(self, step_number, Flag)
    
    def as_GE(self, *params):
        Flag=False
        try:
            testobj = unittest.TestCase()
            testobj.assertGreaterEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
            Global_variables.verification_steps.append(params[2] + " - PASSED.")
            Flag=True
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            msg = m.group(1)+" - FAILED."+'As GEqual to comparison : value1=['+str(params[0])+'], value2=['+str(params[1])+']'
            print(msg)
            Global_variables.verification_steps.append(msg)
            Global_variables.asert_failure_count += 1
            UtillityMethods.create_failure_log(self, msg)
        finally:
            suite_type=UtillityMethods.parseinitfile(self, 'suite_type')
            try:
                try:
                    step_number=re.search(r'\d+.\d+', params[2]).group()
                except AttributeError:
                    step_number=re.search(r'\d+', params[2]).group()
            except AttributeError:
                step_number=re.search(r'[x|X]', params[2]).group()
            if suite_type == 'smoke':
                file_name=Global_variables.current_test_case
                file_path=os.getcwd() + "\\" + file_name + "_" + step_number + ".png"
                self.driver.save_screenshot(file_path)
            else:
                pass
        UtillityMethods.verification_screenshot_capture(self, step_number, Flag)

    def as_LE(self, *params):
        Flag=False
        try:
            testobj = unittest.TestCase()
            testobj.assertLessEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
            Global_variables.verification_steps.append(params[2] + " - PASSED.")
            Flag=True
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            msg = m.group(1)+" - FAILED."+'As LEqual to comparison : value1=['+str(params[0])+'], value2=['+str(params[1])+']'
            print(msg)
            Global_variables.verification_steps.append(msg)
            Global_variables.asert_failure_count += 1
            UtillityMethods.create_failure_log(self, msg)
        finally:
            suite_type=UtillityMethods.parseinitfile(self, 'suite_type')
            try:
                try:
                    step_number=re.search(r'\d+.\d+', params[2]).group()
                except AttributeError:
                    step_number=re.search(r'\d+', params[2]).group()
            except AttributeError:
                step_number=re.search(r'[x|X]', params[2]).group()
            if suite_type == 'smoke':
                file_name=Global_variables.current_test_case
                file_path=os.getcwd() + "\\" + file_name + "_" + step_number + ".png"
                self.driver.save_screenshot(file_path)
            else:
                pass
        UtillityMethods.verification_screenshot_capture(self, step_number, Flag)
    
    def verification_screenshot_capture(self, step_number, Flag):
        '''
        Description: This will capture screenshot and save it current working directory.
        Flag: Based on this argument test case name will go(Ex: if Flag is True, then case is C0011, Else Flag is Flase, then case is test_C0011)
        Usage : verification_screenshot_capture('09.01', True)
        '''
        case = Global_variables.current_test_case
        if not os.path.isdir(case):
            os.makedirs(case)
        case_id = case if Flag else 'test_'+str(case)
        file_name = case_id + '_' + step_number
        try:
            path_location = os.path.join(os.getcwd(),case)
            file_location = os.path.join(path_location, file_name + '.png')
            self.driver.save_screenshot(file_location)
            Global_variables.verification_test_case_name.append(file_name)
        except Exception as e:
            print(Exception(e))
            print('Exception in save screenshot of verification : ' + Global_variables.current_test_case)
        
    def generate_verification_html_file(self, script_exception=False):
        '''
        Description: This will generate a html file and save it as case_id.html.
        Usage : generate_verification_html_file()
        '''
        case = Global_variables.current_test_case
        if not os.path.isdir(case):
            os.makedirs(case)
        path_location = os.path.join(os.getcwd(), case)
        file_name = os.path.join(path_location, 'index.html')
        with open(file_name, 'w') as html:
            html_start = """
                        <!DOCTYPE html>
                        <html>
                        <meta name="viewport" content="width=device-width, initial-scale=1.0">
                        <style>
                        * {
                        box-sizing: border-box;
                        }
                        p{
                            color: green;
                        }
                        .pErr{
                            color: red;
                        }
                        img{
                        width: 100%;
                        }
                        #time{
                            color: black;
                        }
                        </style>
                        <body>
                        """
            html_end = """
                        </body>
                        </html>
                       """
            html.write(html_start)
            html.write('<div><h1>' + case + '</h1></div>')
            html.write('<div><h3>Test case credentials: ' + str(Global_variables.used_username_in_current_session) + '</h3></div>')
            if len(Global_variables.verification_test_case_name) == len(Global_variables.verification_steps):
                for case_name, step in zip(Global_variables.verification_test_case_name, Global_variables.verification_steps):
                    if 'test_' not in case_name:
                        p = '<div><p id="' + case_name + '">' + step + '</p></div>'
                        img = '<div><img src = "' + case_name + '.png" alt="' + case_name + ' image not able to found"/></div>'
                    else:
                        p = '<div><p class="pErr">' + step + '</p></div>'
                        img = '<div><img src = "' + case_name + '.png" alt="' + case_name + ' image not able to found"/></div>'
                    html.write(p)
                    html.write(img)
            if script_exception:
                p = '<div><p class="pErr">Unexpected failure occurs before reaching next verification point...</p></div>'
                img = '<div><img src = "test_' + case + '.png"  alt="' + case + ' image not able to found"/></div>'
                html.write(p)
                html.write(img)
            html.write('<p id="time">' + str(datetime.now().strftime('%m/%d/%Y %H:%M:%S')) + '<p>')    
            html.write(html_end)
            html.close()
        
    def capture_screenshot(self, step_number, testrail_step_doc, expected_image_verify=False):
        '''
        Description: This will capture screenshot and save it current working directory.
        Flag: Based on this argument test case name will go(Ex: if Flag is True, then case is C0011, Else Flag is Flase, then case is test_C0011)
        Usage : verification_screenshot_capture('09.01', True)
        '''
        try:
            Global_variables.testrail_steps_doc.append(testrail_step_doc)
            step_num = str(step_number).replace(' ','')
            file_name = Global_variables.current_test_case+"_expected_image_"+step_num if expected_image_verify else Global_variables.current_test_case+"_actual_image_"+step_num
            path_location = os.path.join(os.getcwd(),'actual_images')
            if not os.path.isdir(path_location):
                os.makedirs(path_location)
            file_location = os.path.join(path_location, Global_variables.current_test_case+"_actual_image_"+step_num)
            self.driver.save_screenshot(file_location + '.png')
            Global_variables.test_case_name_to_log.append(file_name)
        except Exception as e:
            print(Exception(e))
            print('Exception in save screenshot of verification : ' + Global_variables.current_test_case)
        
    def generate_html_file(self):
        '''
        Description: This will generate a html file and save it as case_id.html.
        Usage : generate_html_file()
        '''
        case=Global_variables.current_test_case
        path_location = os.path.join(os.getcwd(),'actual_images')
        if not os.path.isdir(path_location):
            os.makedirs(path_location)
        file_location = os.path.join(path_location, Global_variables.current_test_case+".html")
        with open(file_location, 'w') as html:
            html_start = """
                <!DOCTYPE html>
                <html>
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <style>
                * {
                box-sizing: border-box;
                }
                img, th{
                width: 50%;
                border: 1px solid gray;
                }
                table{
                width:100%;
                }
                #time{
                color: black;
                }
                </style>
                <body>
                """
            html_end = """</body></html>"""
            html.write(html_start)
            html.write('<div><h1>' + case + '</h1></div>')
            html.write('<div><h3>Test case credentials: ' + str(Global_variables.used_username_in_current_session) + '</h3></div>')
            if len(Global_variables.test_case_name_to_log) == len(Global_variables.testrail_steps_doc):
                try:
                    for case_name, doc in zip(Global_variables.test_case_name_to_log, Global_variables.testrail_steps_doc):
                        row1='<div><hr/></div>'
                        row2="<div>"+doc+"</div>"
                        if '_expected_image_' in case_name:
                            row3="<div><table><tr><th>Expected Image</th><th>Actual Image</th></tr><table></div>"
                            row4="<div><img src='" + case_name + ".png' alt='" + case_name + " image not able to found'/><img src='" + case_name.replace('expected_image', 'actual_image') + ".png' alt='" + case_name.replace('expected_image', 'actual_image') + " image not able to found'/></div>"
                        else:
                            row3="<div><table><tr><th>Actual Image</th></tr><table></div>"
                            row4="<div><img src='" +case_name + ".png' style='width:100%;' alt='" + case_name + " image not able to found'/></div>"
                        html.write(row1)
                        html.write(row2)
                        html.write(row3)
                        html.write(row4)
                except Exception as e:
                    html.write(Exception(e))
            row5="<div id='time'>"+str(datetime.now().strftime('%m/%d/%Y %H:%M:%S'))+"<div>"
            html.write(row5)
            html.write(html_end)
    
    def verify_list_values(self, expected_list_value, actual_list_value, msg, assert_type='asequal'):
        """
        Description : This method will verify list values with different assert types
        :Arg - assert_type = 'asequal' Verify whether given actual and expected list values are same
        :Arg - assert_type = 'asin' Verify whether given expected list values are available in given actual list values
        :Arg - assert_type = 'asnotin' Verify whether given expected list values are not available in given actual list values
        :Example1 : verify_list_values([1, 2, 3, 4], [1, 2, 3, 4], 'Step 01 : Verify two list equal', assert_type='asequal')
        :Example2 : verify_list_values([3, 4], [1, 2, 3, 4], 'Step 01 : Verify 3 and 4 available in [1, 2, 3, 4]', assert_type='asin')
        :Example3 : verify_list_values([3], [1, 2, 3, 4], 'Step 01 : Verify 3 not available in [1, 2, 3, 4]', assert_type='asnotin')
        """
        
        if assert_type == 'asequal' :
            UtillityMethods.as_List_equal(self, expected_list_value, actual_list_value, msg)
        elif assert_type == 'asin' :
            token=True
            for expected_value in expected_list_value :
                if expected_value not in actual_list_value :
                    token = False
                    break
            if token == False :
                UtillityMethods.asin(self, expected_value, actual_list_value, msg)
            else : 
                UtillityMethods.asequal(self, True, token, msg)
        elif assert_type == 'asnotin' :
            token=True
            for expected_value in expected_list_value :
                if expected_value in actual_list_value :
                    token = False
                    break
            if token == False :
                UtillityMethods.as_notin(self, expected_value, actual_list_value, msg)
            else : 
                UtillityMethods.asequal(self, True, token, msg)
        else :
            type_error = "Currently '{0}' assert type not using"
            raise TypeError(type_error)

    def compare_data_set(self, table_id, tr_id, file_name, color='default', **kwargs):
        #file_name1="act_"+file_name
        #UtillityMethods.create_data_set(self, table_id, tr_id, file_name1, color, **kwargs)
        #status = UtillityMethods.compare_excel_sheet(self, file_name, file_name1, 'Sheet') 
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr[id*='I0']"
            conditional_css = 'rgb'
        elif tr_id =='bgcolor':
            bgcolor = str(UtillityMethods.color_picker(self, color, 'rgb'))
            css="[id='{0}'] tr[id*='I0']"
            conditional_css = 'background-color: ' + bgcolor
        else:
            css="[id='{0}'] tr[id^='{1}']"
        #wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        #wb = Workbook()
        #status=[]
        #sheet = wb.get_active_sheet
        #s = wb.get_sheet_by_name('Sheet')
        #s1 = wb1.get_sheet_by_name('Sheet')
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=None
        table_css = css.format(table_id, tr_id)
        if tr_id not in ['notrgb', 'bgcolor']:
            status = UtillityMethods.verify_table_data(self, table_css, file_name, number_of_rows=no_of_rows)
        else:
            soup = UtillityMethods.beautifulsoup_object_creation(self)
            wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
            status=[]
            s1 = wb1.get_sheet_by_name('Sheet')
            table_rows = soup.select(table_css)
            if no_of_rows is not None:
                table_row_lenght = int(no_of_rows)
            else:
                table_row_lenght = len(table_rows)
            if tr_id == 'notrgb':
#                 row_number=1
                for table_row in range(0, table_row_lenght):
#                 for table_row in table_rows:
                    table_row_data = str(table_rows[table_row].get('style')).replace(';', '').replace(' ', '')
                    if conditional_css not in table_row_data:
                        total_col = table_rows[table_row].find_all('td')
                        for c in range(len(total_col)):
                            browser_value=str(total_col[c].get_text(strip=True))
                            if s1.cell(row=table_row+1, column=c + 1).value != None and s1.cell(row=table_row+1, column=c + 1).value != ' ':
                                excel_value=s1.cell(row=table_row+1, column=c + 1).value
                            else:
                                excel_value=None
                            if (excel_value !=None and browser_value != '') or (excel_value != None) or (browser_value != ''):
                                if excel_value == browser_value:
                                    status=[0]
                                else:
                                    status=[table_row+1,c+1]
                                    return (status)
#                         row_number += 1
            elif tr_id =='bgcolor':
#                 row_number=1
                for table_row in range(0, table_row_lenght):
#                 for table_row in table_rows:
                    table_row_data = str(table_rows[table_row].get('style')).replace(';', '').replace(' ', '')
                    if conditional_css.replace(' ', '') == table_row_data:
                        total_col = table_rows[table_row].find_all('td')
                        for c in range(len(total_col)):
                            browser_value=str(total_col[c].get_text(strip=True))
                            if s1.cell(row=table_row+1, column=c + 1).value != None and s1.cell(row=table_row+1, column=c + 1).value != ' ':
                                excel_value=s1.cell(row=table_row+1, column=c + 1).value
                            else:
                                excel_value=None
                            if (excel_value != None and browser_value != '') or (excel_value != None) or (browser_value != ''):
                                if excel_value == browser_value:
                                    status=[0]
                                else:
                                    status=[table_row+1,c+1]
                                    return (status)
#                         row_number += 1
        return (status)

#         for r in range(0, len(table_rows)):
#             total_col = table_rows[r].find_all('td')
#             for c in range(len(total_col)):
#                 value=str(total_col[c].get_text(strip=True))
#                 print(value)
#                 if s1.cell(row=r + 1, column=c + 1).value != None and value != '':
#                     if s1.cell(row=r + 1, column=c + 1).value.replace(' ','') == value.replace(' ',''):
#                         status=[0]
#                         continue
#                     else:
#                         status=[r+1,c]
#                         return (status)
#         return (status)
#         table_rows = self.driver.find_elements_by_css_selector(css.format(table_id, tr_id))
#         if 'desired_no_of_rows' in kwargs:
#             no_of_rows=kwargs['desired_no_of_rows']
#         else:
#             no_of_rows=len(table_rows)
#         for r in range(0, no_of_rows):
#         #for r in range(0, len(table_rows)):
#             columns = table_rows[r].find_elements_by_css_selector("td")
#             for c in range(0, len(columns)):
#                 if s1.cell(row=r + 1, column=c + 1).value != None:
#                     if s1.cell(row=r + 1, column=c + 1).value.strip() == str(columns[c].text.strip()):
#                         status=[0]
#                         continue
#                     else:
#                         status=[r+1,c]
#                         return (status)
#         return (status)

    def create_data_set(self,table_id, tr_id, file_name, color='default', **kwargs):
        """
        Usage: utillobj.create_data_set('ITableData0','I0r','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','rgb','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','notrgb','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','bgcolor','C2140681.xlsx', color='green')
        """
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr[id*='I0']"
            conditional_css = 'rgb'
        elif tr_id =='bgcolor':
            bgcolor = str(UtillityMethods.color_picker(self, color, 'rgb'))
            css="[id='{0}'] tr[id*='I0']"
            conditional_css = 'background-color: ' + bgcolor
        else:
            css="[id='{0}'] tr[id^='{1}']"
        #wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        #wb = Workbook()
        #status=[]
        #sheet = wb.get_active_sheet
        #s = wb.get_sheet_by_name('Sheet')
        #s1 = wb1.get_sheet_by_name('Sheet')
        table_css = css.format(table_id, tr_id)
        if tr_id not in ['notrgb', 'bgcolor']:
            UtillityMethods.create_table_data(self, table_css, file_name)
        else:
            soup = UtillityMethods.beautifulsoup_object_creation(self)
            wb = Workbook()
            s = wb.get_sheet_by_name('Sheet')
            table_rows = soup.select(table_css)
            row_number=1
            if tr_id == 'notrgb':
                for table_row in table_rows:
                    table_row_data = str(table_row.get('style'))
                    if conditional_css not in table_row_data:
                        total_col = table_row.find_all('td')
                        for c in range(len(total_col)):
                            s.cell(row=row_number, column=c + 1).value = total_col[c].get_text(strip=True)
                        row_number += 1
            elif tr_id =='bgcolor':
                for table_row in table_rows:
                    table_row_data = str(table_row.get('style')).replace(';', '').replace(' ', '')
                    if conditional_css.replace(' ', '') == table_row_data:
                        total_col = table_row.find_all('td')
                        for c in range(len(total_col)):
                            s.cell(row=row_number, column=c + 1).value = total_col[c].get_text(strip=True)
                        row_number += 1
            wb.save(os.getcwd() + "\data\\" + file_name)
        '''
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr:not([style*='rgb'])[id*='I0']"
        elif tr_id=='bgcolor':
            bgcolor=self.color_picker(color, 'rgb')
            css="[id='{0}'] tr[style*='background-color: " + bgcolor + "']"
        else:
            css="[id='{0}'] tr[id^='{1}']"
        #from openpyxl import Workbook
#         wb = Workbook()
        #sheet = wb.get_active_sheet
#         s = wb.get_sheet_by_name('Sheet')
        table_css = css.format(table_id, tr_id)
        UtillityMethods.create_table_data(self, table_css, file_name)
#         table_rows = self.driver.find_elements_by_css_selector(css.format(table_id, tr_id))
#         if 'desired_no_of_rows' in kwargs:
#             no_of_rows=kwargs['desired_no_of_rows']
#         else:
#             no_of_rows=len(table_rows)
#         for r in range(0, no_of_rows):
#         #for r in range(0, len(table_rows)):
#             columns = table_rows[r].find_elements_by_css_selector("td")
#             for c in range(0, len(columns)):
#                 s.cell(row=r + 1, column=c + 1).value = str(columns[c].text.strip())
#         wb.save(os.getcwd() + "\data\\" + file_name)
'''

    def verify_data_set(self,table_id, tr_id, file_name, msg, color='default', **kwargs):
        """
        Usage: utillobj.verify_data_set('ITableData0','I0r','C2140681.xlsx',"Step 10: verify data set")
        Usage: utillobj.verify_data_set('ITableData0','rgb','C2140681.xlsx',"Step 10: verify data set")
        Usage: utillobj.verify_data_set('ITableData0','notrgb','C2140681.xlsx',"Step 10: verify data set")
        Usage: utillobj.verify_data_set('ITableData0','bgcolor','C2140681.xlsx',"Step 10: verify data set", color='green')
        """
        x=UtillityMethods.compare_data_set(self, table_id, tr_id, file_name, color=color, **kwargs)
        UtillityMethods.asequal(self, len(x), 1, msg+ ' Row,Column -'+ str(x))
    '''******************verify data set previous verification***************************'''

    def compare_data_set_old(self, table_id, tr_id, file_name, color='default', **kwargs):
        #file_name1="act_"+file_name
        #UtillityMethods.create_data_set(self, table_id, tr_id, file_name1, color, **kwargs)
        #status = UtillityMethods.compare_excel_sheet(self, file_name, file_name1, 'Sheet') 
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr:not([style*='rgb'])[id*='I0']"
        elif tr_id=='bgcolor':
            bgcolor=self.color_picker(color, 'rgb')
            css="[id='{0}'] tr[style*='background-color: " + bgcolor + "']"
        else:
            css="[id='{0}'] tr[id ^='{1}']"
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        #wb = Workbook()
        status=[]
        #sheet = wb.get_active_sheet
        #s = wb.get_sheet_by_name('Sheet')
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(css.format(table_id, tr_id))
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        for r in range(0, no_of_rows):
        #for r in range(0, len(table_rows)):
            columns = table_rows[r].find_elements_by_css_selector("td")
            for c in range(0, len(columns)):
                if s1.cell(row=r + 1, column=c + 1).value != None:
                    if s1.cell(row=r + 1, column=c + 1).value.strip() == str(columns[c].text.strip()):
                        status=[0]
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)

    def create_data_set_old(self,table_id,tr_id,file_name, color='default', **kwargs):
        """
        Usage: utillobj.create_data_set('ITableData0','I0r','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','rgb','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','notrgb','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','bgcolor','C2140681.xlsx', color='green')
        """
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr:not([style*='rgb'])[id*='I0']"
        elif tr_id=='bgcolor':
            bgcolor=self.color_picker(color, 'rgb')
            css="[id='{0}'] tr[style*='background-color: " + bgcolor + "']"
        else:
            css="[id='{0}'] tr[id ^='{1}']"
        #from openpyxl import Workbook
        wb = Workbook()
        #sheet = wb.get_active_sheet
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(css.format(table_id, tr_id))
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        for r in range(0, no_of_rows):
        #for r in range(0, len(table_rows)):
            columns = table_rows[r].find_elements_by_css_selector("td")
            for c in range(0, len(columns)):
                s.cell(row=r + 1, column=c + 1).value = str(columns[c].text.strip())
        wb.save(os.getcwd() + "\data\\" + file_name)

    def verify_data_set_old(self,table_id, tr_id, file_name,msg, color='default', **kwargs):
        """
        Usage: utillobj.verify_data_set('ITableData0','I0r','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','rgb','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','notrgb','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','bgcolor','C2140681.xlsx',"Step 10: fail data set", color='green')
        """
        x= self.compare_data_set_old(table_id,tr_id,file_name, color)
        self.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))
        
    '''******************verify data set previous verification***************************'''


    def compare_popup_data_set(self, popup_id, table_id, file_name, **kwargs):
        """
        Usage: utillobj.compare_popup_data_set('wall2','FiltSel2','C2140681.xlsx')
        """
        #file_name1="act_"+file_name
        #UtillityMethods.create_popup_data_set(self, popup_id, table_id, file_name, **kwargs)
        #status = UtillityMethods.compare_excel_sheet(self, file_name, file_name1, 'Sheet')
        
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        #wb = Workbook()
        status=[]
        #sheet = wb.get_active_sheet
        #s = wb.get_sheet_by_name('Sheet')
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector("[id='{0}'] table[id='{1}'] tr".format(popup_id, table_id))
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        for r in range(0, no_of_rows):
            columns = table_rows[r].find_elements_by_css_selector("td")
            for c in range(0, len(columns)):
#                 tmp=s1.cell(row=r + 1, column=c + 1).value.strip() strip not working for excel
                tmp=s1.cell(row=r + 1, column=c + 1).value
                if (tmp):
                    if tmp.replace(' ','').lower().replace('\u00A0', ' ').strip() == str(columns[c].text.replace(' ','').lower().replace('\u00A0', ' ').strip()):
                        status=[0]
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)


    def create_popup_data_set(self, popup_id, table_id, file_name, **kwargs):
            """
            Usage: utillobj.create_popup_data_set('wall2','FiltSel2','C2140681.xlsx')
            """
            #from openpyxl import Workbook
            wb = Workbook()
            #sheet = wb.get_active_sheet
            s = wb.get_sheet_by_name('Sheet')
            table_rows = self.driver.find_elements_by_css_selector("[id='{0}'] table[id ='{1}'] tr".format(popup_id, table_id))
            if 'desired_no_of_rows' in kwargs:
                no_of_rows=kwargs['desired_no_of_rows']
            else:
                no_of_rows=len(table_rows)
            for r in range(0, no_of_rows):
                columns = table_rows[r].find_elements_by_css_selector("td")
                for c in range(0, len(columns)):
                    s.cell(row=r + 1, column=c + 1).value = str(columns[c].text.strip())
            wb.save(os.getcwd() + "\data\\" + file_name)
    
    def verify_popup_data_set(self, popup_id, table_id, file_name, msg):
            """
            Usage: utillobj.verify_popup_data_set('wall2','FiltSel2','C2140681.xlsx',"Step 10: fail data set")
            """
            x= self.compare_popup_data_set(popup_id, table_id, file_name)
            self.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))
            
    def compare_pivot_data_set(self, table_id, file_name, **kwargs):
        """
        Usage: utillobj.compare_pivot_data_set('piv1', C2140681_Ds01.xlsx)
        """
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        rows_css = "#" + table_id + " > tbody > tr:nth-child(2) > td >table > tbody > tr"
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        for r in range(0, no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                if (s1.cell(row=r + 1, column=c + 1).value):
                    value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text.strip()
                    if s1.cell(row=r + 1, column=c + 1).value.replace(' ','').lower() == str(value).replace(' ','').lower():
                        status=[0]
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)


    def create_pivot_data_set(self, table_id, file_name):
        """
        Syntax: utillobj.create_pivot_data_set('piv1', 'C2140681_Ds01.xlsx')
        """
        rows_css = "#" + table_id + " > tbody > tr:nth-child(2) > td >table > tbody > tr"
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        for r in range(len(table_rows)):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text.strip()
                s.cell(row=r + 1, column=c + 1).value = str(value)
        wb.save(os.getcwd() + "\data\\" + file_name)

    
    def verify_pivot_data_set(self, table_id, file_name, msg, **kwargs):
        """
        Usage: utillobj.verify_pivot_data_set('piv1', 'C2140681_Ds01.xlsx,"Step 10: fail data set")
        """
        x= self.compare_pivot_data_set(table_id, file_name, **kwargs)
        self.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))
        
    def ie_window_handle(self, wndnum, suite_type, **kwargs):
        a=self.driver.title
#         temp_obj=a[0:10].replace(' ','').replace('-','')
        c=kwargs['window_title'] if 'window_title' in kwargs else a[0:10].replace(' ','').replace('-','')+" - Internet Explorer"
        for x in range(0, 5):
            self.driver.implicitly_wait=0
            try:
                val=pywinauto.findwindows.find_window(title_re=c)
                print("IE crash not happend.")
                time.sleep(5)
                return(0)                     
            except:
                self.driver.implicitly_wait=0
                try:
                    val1=pywinauto.findwindows.find_window(title_re='Internet Explorer')
                    time.sleep(2)
                    _path = os.path.join(root_path.ROOT_PATH, 'IE_crash_close_program.exe')
                    os.system(_path)
                    time.sleep(1)
                    return(1)
                except:
                    pass                            
            time.sleep(5)
            
    def switch_to_window(self,wndnum, pause=15, **kwargs):
        """
        Usage: switch_to_window(1)
        """
        time.sleep(5)
        try :
            current_window_title = self.driver.title
        except :
            pass
        suite_type = UtillityMethods.parseinitfile(self, 'suite_type')
        status=0
        num = kwargs['win_num'] if 'win_num' in kwargs else 0
        if wndnum > num and suite_type == 'visualization' and Global_variables.browser_name == 'ie':
            status=UtillityMethods.ie_window_handle(self, wndnum, suite_type, **kwargs)
        print("status",status)
        if suite_type == 'visualization' and Global_variables.browser_name == 'ie' and wndnum ==1 and status > 0:
            self.toLog("Inside Crash")
            Global_variables.ie_crash_wndnum=0
            unittest.TestCase.fail(self, "IE crash failure")
            self.toLog("break 2nd - if IE status 1 then mark as fail")         
        UtillityMethods.switch_to_window_handle_test(self, wndnum, **kwargs)
        time.sleep(pause)
        self.driver.maximize_window()
        window_widht = self.driver.execute_script("return window.innerWidth|| document.documentElement.clientWidth|| document.body.clientWidth || document.body.scrollWidth;")
        if window_widht < self.driver.execute_script("return screen.availWidth;"):
            self.driver.maximize_window()
        time.sleep(pause/2)
        br=UtillityMethods.get_browser_height(self)
        get_outer_height = 0
        if Global_variables.browser_name == 'ie' and wndnum > 0:
            get_outer_height = br['browser_height'] - br['outer_height']
        Global_variables.current_working_area_browser_x=br['browser_width']
        Global_variables.current_working_area_browser_y=br['browser_height'] - get_outer_height
        try :
            if Global_variables.browser_name =='edge' and wndnum == 1 and suite_type == 'visualization' :
                core_utilobj.bring_edge_window_to_foregound(self, current_window_title)
        except :
            pass

    def switch_to_main_window(self, **kwargs):
        """
        Usage : switch_to_main_window()
        Close all open window and switch to main window, this is useful when you want to return back to main window and close other opened window
        @author: AAKhan
        """
        total_window=len(self.driver.window_handles)
        i = total_window - 1
        if total_window > 1:
            timeout=0
            while i:
                self.driver.switch_to.window(self.driver.window_handles[i])
                time.sleep(3)
                self.driver.close()
                time.sleep(3)
                i-=1
                if i == 0:
                    break
                if timeout == 250:
                    print('No window found')
                    break
                timeout+=1
        self.driver.switch_to.window(self.driver.window_handles[i])
        time.sleep(3)
        br=UtillityMethods.get_browser_height(self)
        Global_variables.current_working_area_browser_x=br['browser_width']
        Global_variables.current_working_area_browser_y=br['browser_height']
        
    def kill_process(self, process_name):
        """
        :param process_name: give executable file you want to kill(check in task manager to correct process name)
        :usage pkill('notepad')
        """
        try:
            killed = os.system("taskkill /im " + process_name+".exe /f")
        except Exception:
            killed = 0
        return killed
    
    def take_browser_screenshot(self, file_name, image_type='actual', **kwargs):#Need to delete
        """
        This action is used to take browser screenshot, and save to '.png' format.
        :param file_name: file for saving
        :param image_type: where you want to save your image in directory
        :usage take_monitor_screenshot('full_monitor_screenshot', image_type='actual')
        """
        location='actual_images' if image_type=='actual' else 'failure_captures' if image_type=='fail' else 'images'
        browser = UtillityMethods.get_browser_height(self)
        left_rs=kwargs['left'] if 'left' in kwargs else 0
        top_rs = kwargs['top'] if 'top' in kwargs else browser['browser_height']
        right = kwargs['right'] if 'right' in kwargs else 0
        bottom = kwargs['bottom'] if 'bottom' in kwargs else browser['outer_height']
        file_path=os.getcwd() + "\\" + location + "\\" + file_name + ".png"
        im=ImageGrab.grab()
        im.save(file_path)
        resolution=pyautogui.size()
        right_rs=resolution[0]-right
        bottom_rs=resolution[1]-bottom
        bbox = (left_rs, top_rs, right_rs, bottom_rs)
        base_image = Image.open(file_path)
        cropped_image = base_image.crop(bbox)
        cropped_image.save(file_path)
    
    def take_monitor_screenshot(self, file_name, image_type='actual', left=0, top=0, right=0, bottom=0):#Need to delete
        """
        :param file_name: file for saving
        :param image_type: where you want to save your image in directory
        :param left: how much you want to reduce the size from left in output of your image
        :param top: how much you want to reduce the size from top in output of your image
        :param right: how much you want to reduce the size from right in output of your image
        :param bottom: how much you want to reduce the size from bottom in output of your image
        :usage take_monitor_screenshot('full_monitor_screenshot', image_type='actual', left=10, top=25, right=10, bottom=25)
        """        
        if image_type=='actual' :
            location='actual_images'
        elif image_type=='fail' :
            location='failure_captures'
        else:
            location='images'
        file_path=os.getcwd() + "\\" + location + "\\" + file_name + ".png"
        im=ImageGrab.grab()
        im.save(file_path)
        resolution=pyautogui.size()
        left_rs = left
        top_rs = top
        right_rs=resolution[0]-right
        bottom_rs=resolution[1]-bottom
        bbox = (left_rs, top_rs, right_rs, bottom_rs)
        base_image = Image.open(file_path)
        cropped_image = base_image.crop(bbox)
        cropped_image.save(file_path)
        
          
    def take_screenshot(self, element, file_name, image_type='actual', x=0, y=0, w=0, h=0):#Need to delete
        time.sleep(15)
        if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
            UtillityMethods.click_type_using_pyautogui(self, element, move=True)
        else:
            action = ActionChains(self.driver)
            action.move_to_element_with_offset(element, 0, 0).perform()
            del action
        time.sleep(3)
        location='actual_images' if image_type=='actual' else 'images'
        file_path=os.getcwd() + "\\" + location + "\\" + file_name + ".png"
        bounding_box = (
            element.location['x']+x,  # left
            element.location['y']+y,  # upper
            (element.location['x'] + element.size['width'])+w,  # right
            (element.location['y'] + element.size['height'])+h  # bottom
        )
        warnings.simplefilter("ignore", ResourceWarning)
        return self.bounding_box_screenshot(bounding_box, file_path)

    def bounding_box_screenshot(self, bounding_box, filename):#Need to delete
        self.driver.save_screenshot(filename)
        base_image = Image.open(filename)
        cropped_image = base_image.crop(bounding_box)
        base_image = base_image.resize(cropped_image.size)
        base_image.paste(cropped_image, (0, 0))
        base_image.save(filename)
        return base_image
    
    def autoit_print(self,msg):
        if Global_variables.browser_name in ['firefox']:
            file = os.getcwd() + "\common\\lib\\firefoxPrint.exe"
            proc = subprocess.Popen(file, stdout=subprocess.PIPE)
            (out, err) = proc.communicate()
            #val = (str(out.decode('utf-8')))
            #self.asin('28E1', val, msg)                
            
        else:
            file = os.getcwd() + "\common\\lib\\chromePrint.exe"
            proc = subprocess.Popen(file, stdout=subprocess.PIPE)
            (out, err) = proc.communicate()
            #val = (str(out.decode('utf-8')))
            #self.asin('28E1', val, msg)  
            
    def verify_printer_window(self,msg):        
        file = os.getcwd() + "\common\\lib\\printer_window.exe"
        proc = subprocess.Popen(file, stdout=subprocess.PIPE)
        (out, err) = proc.communicate()
        val = (str(out.decode('utf-8')))
        self.asin('success', val, msg)     
    
    def verify_export_save_as(self,actual_file): 
        """
        Usage: utillobj.verify_export_save_as('C2053849_Ds01')
        """     
        if os.path.isfile(os.getcwd()+"\data\\"+actual_file+".htm"):
            os.remove(os.getcwd()+"\data\\"+actual_file+".htm")
        procesobj = subprocess.Popen(os.getcwd()+"\common\\lib\\export_save_as.exe \data\\"+actual_file)
        procesobj.wait()
        time.sleep(3)
        del procesobj
           
    def check_htm_file_exist(self,actual_file,msg):
        """
        Usage: utillobj.check_htm_file_exist("C2053849_Ds01.htm","Step 02: Verify HTML file")
        """
        file = os.path.isfile(os.getcwd()+"\data\\"+actual_file)
        self.asequal(file,True,msg)
        
    def save_window(self,actual_file, **kwargs):
        """
        Usage: utillobj.save_window('C2053851_actual_2')
            or
        Usage: save_window('C2053851_actual_1_'+browser, pyautogui_save=True)
        """
        if os.path.isfile(os.getcwd()+"\\data\\"+actual_file+".xls"):
            os.remove(os.getcwd()+"\\data\\"+actual_file+".xls")
        if os.path.isfile(os.getcwd()+"\\data\\"+actual_file+".xlsx"):
            os.remove(os.getcwd()+"\\data\\"+actual_file+".xlsx")
        if os.path.isfile(os.getcwd()+"\\data\\"+actual_file+".pdf"):
            os.remove(os.getcwd()+"\\data\\"+actual_file+".pdf")
        actual_file=actual_file+".xls"
        if 'pyautogui_save' in kwargs:
            pyautogui.FAILSAFE=False
            time.sleep(15)
            pyautogui.typewrite(os.getcwd()+"\\data\\"+actual_file,  interval=0.2, pause=2)
            pyautogui.press('enter',  pause=7)
        else:
            if Global_variables.browser_name in ['chrome']:
                    procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\chrome_save_window.exe \\data\\"+actual_file)
                    procesobj.wait()
                    time.sleep(3)
                    del procesobj
            if Global_variables.browser_name in ['firefox']:
                    procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\FF_save_window.exe \\data\\"+actual_file)
                    procesobj.wait()
                    time.sleep(3)
                    del procesobj
            if Global_variables.browser_name in ['ie', 'edge']:
                    procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\IE_save_window.exe \\data\\"+actual_file)
                    procesobj.wait()
                    time.sleep(3)
                    del procesobj
        keyboard.send('ctrl+j')
        time.sleep(2)
        keyboard.send('ctrl+j')
        time.sleep(2)   
            
    def create_excel(self,path1,path2, **kwargs):
        """
        This function open exiting excel sheet which has '.xls' extension and save as '.xlsx' extension 
        Usage: utillobj.create_xls('C2053851_actual_2.xls','C2053851_actual_2.xlsx')
            or
        Usage: create_excel('C2053851_actual_1_'+browser+'.xls','C2053851_actual_1_'+browser+'.xlsx', pyautogui_save=True)
        """
        pyautogui.FAILSAFE=False
        if 'msg' in kwargs:
            kwargs['msg']=kwargs['msg'].replace(" ",'_')
        if 'cancle' in kwargs:
            os.system('start excel.exe '+os.getcwd()+"\\data\\"+path1)
            time.sleep(5)
            procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\Cancel_SaveAs_Excel_Sheet.exe "+kwargs['msg'])
            procesobj.wait()
            time.sleep(3)
            del procesobj
        else:
            if 'pyautogui_save' in kwargs:
                os.system('start excel.exe '+os.getcwd()+"\\data\\"+path1)
                time.sleep(5)
                procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\SaveAs_Excel_Sheet.exe \\data\\"+path2)
                procesobj.wait()
                time.sleep(3)
                del procesobj
            else:
                if Global_variables.browser_name in ['chrome']:
                    procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\create_excel.exe \\data\\"+path1+" "+path2)
                    procesobj.wait()
                    time.sleep(3)
                    del procesobj
    
    def saveas_excel_sheet(self, filename, **kwargs):
        """
        :param msg='step_02:'
        :Usage: saveas_excel_sheet('C2053851_actual_1_'+browser+'.xlsx')
                    or
        :Usage: saveas_excel_sheet('C2053851_actual_1_'+browser+'.xlsx',cancle=True, msg='Step_02')
        """
        pyautogui.FAILSAFE=False
        if 'msg' in kwargs:
            kwargs['msg']=kwargs['msg'].replace(" ",'_')
        if 'cancle' in kwargs:
            procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\Cancel_SaveAs_Excel_Sheet.exe "+kwargs['msg'])
            procesobj.wait()
            time.sleep(3)
            del procesobj
        else:
            file=os.getcwd()+"\\data\\"+filename
            if os.path.exists(file):
                try:
                    os.remove(file)
                except:
                    print(filename+" Not exits in directory")
            time.sleep(5)
            procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\SaveAs_Excel_Sheet.exe \\data\\"+filename)
            procesobj.wait()
            time.sleep(3)
            del procesobj
    
    def saveas_excel_sheet_esrimap(self, filename, **kwargs):
        """
        :param msg='step_02:'
        :Usage: saveas_excel_sheet_esrimap('C2053851_actual_1_'+browser+'.xlsx')
                    or
        :Usage: saveas_excel_sheet_esrimap('C2053851_actual_1_'+browser+'.xlsx',cancle=True, msg='Step_02')
        """
        pyautogui.FAILSAFE=False
        if 'msg' in kwargs:
            kwargs['msg']=kwargs['msg'].replace(" ",'_')
        if 'cancle' in kwargs:
            procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\Cancel_SaveAs_Excel_Sheet.exe "+kwargs['msg'])
            procesobj.wait()
            time.sleep(3)
            del procesobj
        else:
            file=os.getcwd()+"\\data\\"+filename
            if os.path.exists(file):
                try:
                    os.remove(file)
                except:
                    print(filename+" Not exits in directory")
            time.sleep(5)
            procesobj = subprocess.Popen(os.getcwd()+"\\common\\lib\\SaveAs_Excel_Sheet_esrimap.exe \\data\\"+filename)
            procesobj.wait()
            time.sleep(3)
            del procesobj
    
    def select_combobox_item(self, combo_id, combo_item, **kwargs):
        """
        Syntax: utillobj.select_combobox_item('comboSourceFields', 'Equals to')
        """
        combo_btn=self.driver.find_element_by_css_selector("div[id*=" + combo_id + "] div[id^='BiButton']")
        core_utilobj.left_click(self, combo_btn)
#         UtillityMethods.default_left_click(self, object_locator=combo_btn, **kwargs)
        time.sleep(3)
        menu_items=self.driver.find_elements_by_css_selector("div[id^='BiPopup'][style*='inherit'] div[id^='BiComboBoxItem']")
        actual_popup_list=[el.text.strip() for el in menu_items]
        core_utilobj.left_click(self, menu_items[actual_popup_list.index(combo_item)])
#         UtillityMethods.default_click(self, menu_items[actual_popup_list.index(combo_item)])
        time.sleep(1)
    
    def select_any_combobox_item(self, combo_btn_elem, combo_item, **kwargs):
        """
        Syntax: utillobj.select_any_combobox_item(combo_btn_elem, 'Equals to')
        """
        UtillityMethods.default_click(self, combo_btn_elem)
        time.sleep(2)
        menu_items=self.driver.find_elements_by_css_selector("div[id^='BiPopup'][style*='inherit'] div[id^='BiComboBoxItem']")
        actual_popup_list=[el.text.strip() for el in menu_items]
        if 'verify' in kwargs:
            actual_popup_list=[el.text.strip() for el in menu_items  if bool(re.match('\S+', el.text.strip()))]
            UtillityMethods.as_List_equal(self,kwargs['expected_combobox_list'], actual_popup_list, kwargs['msg'])
#         core_utilobj.left_click(self, menu_items[actual_popup_list.index(combo_item)])
#         UtillityMethods.default_click(self, menu_items[actual_popup_list.index(combo_item)], mouse_duration=1)
        UtillityMethods.default_click_with_firefox_update(self, menu_items[actual_popup_list.index(combo_item)], mouse_duration=1)
        time.sleep(1)
    
    def autoit_activex(self):
       
        file = os.getcwd() + "\common\\lib\\ActiveX_OK.exe"
        proc = subprocess.Popen(file, stdout=subprocess.PIPE)
        (out, err) = proc.communicate() 
         
    def open_outlook_attachment(self):
            
        file = os.getcwd() + "\common\\lib\\open_outlook_attachment.exe"
        proc = subprocess.Popen(file, stdout=subprocess.PIPE)
        (out, err) = proc.communicate()   
      
        
    def ibfs_save_as(self, file_name, file_type=None,**kwargs):  
        
        """
        Param: file_name: '<fex_name to be saved>'
        Syntax: utillobj.ibfs_save("test1")
        @author = Niranjan
        """
        time.sleep(9)
        file_name_input_css="#IbfsOpenFileDialog7_cbFileName input"
        elem1=(By.CSS_SELECTOR, file_name_input_css)
        self._validate_page(elem1)
#         action = ActionChains(self.driver) 
        self.driver.find_element_by_css_selector(file_name_input_css).clear()
        element = self.driver.find_element_by_css_selector(file_name_input_css)
        UtillityMethods.set_text_to_textbox_using_keybord(self, file_name, text_box_elem=element)
        
#         if Global_variables.browser_name in ['firefox']:
#             time.sleep(2)
#             UtillityMethods.set_text_field_using_actionchains(self, element, file_name)
#             time.sleep(2)
#         elif Global_variables.browser_name in ['ie', 'edge']:
#             time.sleep(2)
#             UtillityMethods.set_text_to_textbox_using_keybord(self, file_name, text_box_elem=element)
#             time.sleep(2)
#         else:
#             action.send_keys_to_element(element,file_name).click(element).perform()
        #action.send_keys(file_name).perform()
        time.sleep(1)
        if file_type != None:
            UtillityMethods.select_combobox_item(self,'IbfsOpenFileDialog7_cbFilterType', file_type)
        if 'save_folder' in kwargs :
            time.sleep(3)
            folder_tree=self.driver.find_elements_by_css_selector("#paneIbfsExplorer_exTree div[class='bi-tree-view-body-content']>table>tbody>tr>td")
            for item in folder_tree :
                if item.text.strip()==kwargs['save_folder'] :
                    folder_icon=item.find_element_by_css_selector("img[src*='folder_closed']")
                    UtillityMethods.default_left_click(self,object_locator=folder_icon, **kwargs)
                    break
            time.sleep(4)
        ok_button_css = "{0} {1}".format(OpenMasterFileDialog.PARENT_CSS, OpenMasterFileDialog.OK_BUTTON_CSS)
        ok_button_elem=UtillityMethods.validate_and_get_webdriver_object(self, ok_button_css, 'Ok Button')
        core_utilobj.python_left_click(self, ok_button_elem)
        try:
            UtillityMethods.synchronize_until_element_is_visible(self, "div[id^='BiDialog'] img[src*='exclamation']", 9)
            if self.driver.find_element_by_css_selector("div[id^='BiDialog'] img[src*='exclamation']").is_displayed():
                btn_css="div[id^='BiDialog'] div[class=bi-button-label]"
                dialog_btns=self.driver.find_elements_by_css_selector(btn_css)
                btn_text_list=[el.text.strip() for el in dialog_btns]
                core_utilobj.python_left_click(self, dialog_btns[btn_text_list.index('Yes')])
        except:
            pass
        Global_variables.saved_case_flag = True
        time.sleep(2)
    
    def ibfs_save(self, file_name, file_type=None,**kwargs):  
        
        """
        Param: file_name: '<fex_name to be saved>'
        Syntax: utillobj.ibfs_save("test1")
        @author = Niranjan
        """
        time.sleep(9)
        file_name_input_css="#IbfsOpenFileDialog7_cbFileName input"
        elem1=(By.CSS_SELECTOR, file_name_input_css)
        self._validate_page(elem1)
#         action = ActionChains(self.driver) 
        self.driver.find_element_by_css_selector(file_name_input_css).clear()
        element = self.driver.find_element_by_css_selector(file_name_input_css)
        UtillityMethods.set_text_to_textbox_using_keybord(self, file_name, text_box_elem=element)
#         if Global_variables.browser_name in ['firefox']:
#             time.sleep(2)
#             UtillityMethods.set_text_field_using_actionchains(self, element, file_name)
#             time.sleep(2)
#         else:
#             action.send_keys_to_element(element,file_name).click(element).perform()
        time.sleep(1)
        if file_type != None:
            UtillityMethods.select_combobox_item(self,'IbfsOpenFileDialog7_cbFilterType', file_type)
        if 'save_folder' in kwargs :
            time.sleep(3)
            folder_tree=self.driver.find_elements_by_css_selector("#paneIbfsExplorer_exTree div[class='bi-tree-view-body-content']>table>tbody>tr>td")
            for item in folder_tree :
                if item.text.strip()==kwargs['save_folder'] :
                    folder_icon=item.find_element_by_css_selector("img[src*='folder_closed']")
                    UtillityMethods.default_left_click(self,object_locator=folder_icon, **kwargs)
                    break
            time.sleep(4)
        elem=self.driver.find_element_by_id("IbfsOpenFileDialog7_btnOK")
        UtillityMethods.default_click(self, elem)
        Global_variables.saved_case_flag = True
        time.sleep(1)
    
    def select_item_in_dialog(self, dialog_css, item_name, click_type='left'):
        """
        Utilobj.select_item_in_dialog("#pageOptionsDlg #iaPageList", "Page 1 ( Copy )")
        """
        oDialog=self.driver.find_element_by_css_selector(dialog_css)
        item_list=oDialog.find_elements_by_css_selector("table tr")
        actual_popup_list=[el.text.strip() for el in item_list if bool(re.match('\S+', el.text.strip()))]
        if click_type=='left':
            item_list[actual_popup_list.index(item_name)].click()
        else:
            UtillityMethods.click_on_screen(self, item_list[actual_popup_list.index(item_name)], 'middle', click_type=1)
        time.sleep(2)
    
    def verify_items_in_dialog(self, dialog_css, expected_item_list, msg):
        """
        Utilobj.verify_items_in_dialog("#pageOptionsDlg #iaPageList", "Page 1 ( Copy )")
        """
        oDialog=self.driver.find_element_by_css_selector(dialog_css)
        item_list=oDialog.find_elements_by_css_selector("table tr")
        actual_popup_list=[el.text.strip() for el in item_list if bool(re.match('\S+', el.text.strip()))]
        UtillityMethods.as_List_equal(self, expected_item_list, actual_popup_list, msg)
        
    def select_or_verify_bipop_menu(self, *args, sync_expected_number=1, **kwargs):
        """
        param: *args will contain the item to be clicked: item_name: 'Show Data'.
        param: kwargs['verify']: This will only verify the bipopup menu.
        param: kwargs['expected_popup_list']: This is a list to verify the menu list.
        param: kwargs['msg']: This is the message to be passed to asertion.
        Syntax: select_or_verify_bipop_menu('Show Data')
        Syntax: select_or_verify_bipop_menu('Edit',verify='true',expected_popup_list=['', 'Delete'],msg='Step 10: Verify popup menu')
        @author = Niranjan
        """
        temp_css=kwargs['custom_css'] if 'custom_css' in kwargs else "table tr"
        bipopup_css="div[id^='BiPopup'][style*='inherit']"
        time.sleep(5)
        #UtillityMethods.synchronize_with_number_of_element(self, bipopup_css, int(sync_expected_number), 4)
        bipopups=self.driver.find_elements_by_css_selector(bipopup_css)
        menu_items=bipopups[len(bipopups)-1].find_elements_by_css_selector(temp_css)
        if Global_variables.browser_name in ['firefox','ie', 'edge']:
            UtillityMethods.click_on_screen(self, bipopups[len(bipopups)-1], 'top_middle', y_offset=2)
        time.sleep(2)
        if 'verify' in kwargs:
            actual_popup_list=[el.text.strip() for el in menu_items  if bool(re.match('\S+', el.text.strip())) and el.is_displayed()] #el.is_displayed() - checking whether element is visible. Because, Edge driver return invisible element text too 
            UtillityMethods.as_List_equal(self, kwargs['expected_popup_list'], actual_popup_list, kwargs['msg'] + " Verify all items in in BiPopup.")
        if 'expected_ticked_list' in kwargs:
            actual_ticked_list=[el.text.strip() for el in menu_items if bool(re.match('.*checked$', el.find_element_by_css_selector("td:nth-child(1)").get_attribute("class")))]
            UtillityMethods.as_List_equal(self, kwargs['expected_ticked_list'], actual_ticked_list, kwargs['msg'] + " Verify ticked or selected items in in BiPopup.")
        if len(args)>0:
            actual_popup_list=[el.text.strip() for el in menu_items]
            cord_xy = core_utilobj.get_web_element_coordinate(self, menu_items[actual_popup_list.index(args[0])], element_location='middle_right', xoffset=-7)
            core_utilobj.move_to_element(self, menu_items[actual_popup_list.index(args[0])])
            core_utilobj.left_click(self, menu_items[actual_popup_list.index(args[0])])
            core_utilobj.python_move_to_offset(self, cord_xy['x'], cord_xy['y'])
        time.sleep(2)
    
    def verify_chart_color(self, parent_id, riser_class, color, msg, **kwargs):
        """
        parent_id='MAINTABLE_0'
        riser_class='riser!s1!g8!mbar!'
        color='green' OR 'text-green'(text-green means it will verify only green not the rgb of green)
        color='green'
        kwargs = attribute_type='fill' or 'stroke' (pass attribute type to get css value of color)
        kwargs = attribute='yes'(if want to verify attribute value for color instead pf css property)
        Syntax:verify_chart_color('MAINTABLE_0', 'riser!s1!g8!mbar!', 'green', 'Step 10: Verify Color')
        Syntax:verify_chart_color('MAINTABLE_0', 'riser!s1!g8!mbar!', 'green', 'Step 10: Verify Color',attribute='yes')
        Syntax:verify_chart_color('MAINTABLE_0', 'riser!s1!g8!mbar!', 'green', 'Step 10: Verify Color',attribute_type='stroke')
        Syntax:verify_chart_color('MAINTABLE_0', 'riser!s1!g8!mbar!', 'green', 'Step 10: Verify Color',attribute_type='stroke',attribute='yes')
        @author : Niranjan
        """
        attribute_type=kwargs['attribute_type'] if 'attribute_type' in kwargs else "fill"
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else "[class*='" + riser_class + "']"
        raiser_css="#"+ parent_id + " " + custom_css
        if 'attribute' in kwargs:
            #if self.browser == 'IE':
            if Global_variables.browser_name in ['ie', 'edge']:
                temp_obj=((self.driver.find_element_by_css_selector(raiser_css).get_attribute(attribute_type))[:-9]+")")[4:]
            else:
                temp_obj=((self.driver.find_element_by_css_selector(raiser_css).get_attribute(attribute_type))[:-10]+")")[4:]
            actual_color = "rgb"+temp_obj
            expected_color=UtillityMethods.color_picker(self, color, 'rgb')
        else:
            actual_color = Color.from_string(self.driver.find_element_by_css_selector(raiser_css).value_of_css_property(attribute_type)).rgba
            expected_color=UtillityMethods.color_picker(self, color, 'rgba')
        UtillityMethods.asequal(self, actual_color, expected_color, msg)
    
    def compare_excel_sheet(self, base_file, actual_file, sheet_name):
        import xlrd
        wb1 = xlrd.open_workbook(os.getcwd() + "\data\\" + base_file)
        ws1 = wb1.sheet_by_name(sheet_name)
        wb2 = xlrd.open_workbook(os.getcwd() + "\data\\" + actual_file)
        ws2 = wb2.sheet_by_name(sheet_name)
        for r in range(0, ws1.nrows):
            for c in range(0, ws1.ncols):
                if ws1.cell(r, c).value == ws2.cell(r, c).value:
                    status=[0]
                    continue
                else:
                    status=[r+1,c+1]
                    return (status)
        return (status)
    
    def verify_excel_sheet(self, base_file, actual_file, sheet_name, msg):
        """
        base_file='C2053851_Base.xlsx'
        actual_file='C2053851_actual.xlsx'
        sheet_name='AHTML'
        usage: utillobj.verify_excel_sheet('C2053851_Base.xlsx', 'C2053851_actual.xlsx', 'AHTML', 'Step 05: Verify that all the records')
        @author : Niranjan
        """
        x= self.compare_excel_sheet(base_file, actual_file, sheet_name)
        self.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))
    
    def compare_excel_sheet_dynamic_values(self, base_file, actual_file, sheet_name,**kwargs):
        import xlrd
        wb1 = xlrd.open_workbook(os.getcwd() + "\data\\" + base_file)
        ws1 = wb1.sheet_by_name(sheet_name)
        wb2 = xlrd.open_workbook(os.getcwd() + "\data\\" + actual_file)
        ws2 = wb2.sheet_by_name(sheet_name)
        if 'no_of_rows' in kwargs:
            total_rows_to_verify = kwargs['no_of_rows']
        else:
            total_rows_to_verify = ws1.nrows
        if 'starting_row' in kwargs:
            starting_rows_to_verify = kwargs['starting_row']
        else:
            starting_rows_to_verify = 0
        for r in range(starting_rows_to_verify, total_rows_to_verify):
            for c in range(0, ws1.ncols):
                print(r,c)
                if type(ws1.cell(r, c).value) == float:
                    base = round(ws1.cell(r, c).value, 2)
                    actual = round(ws2.cell(r, c).value, 2)                        
                else:
                    base = ws1.cell(r, c).value
                    actual = ws2.cell(r, c).value
                if base == actual:
                    status=[0]
                    continue
                else:
                    status=["Row: {0}".format(r+1), "Column: {0}".format(c+1), "Expected: {0}".format(base), "Actual: {0}".format(actual)]
                    return (status)
        return (status)
    
    def verify_excel_sheet_dynamic_values(self, base_file, actual_file, sheet_name, msg, **kwargs):
        """
        base_file='C2053851_Base.xlsx'
        actual_file='C2053851_actual.xlsx'
        sheet_name='AHTML'
        usage: utillobj.verify_excel_sheet_dynamic_values('C2053851_Base.xlsx', 'C2053851_actual.xlsx', 'AHTML', 'Step 05: Verify',starting_row=5,no_of_rows=10)
        """
        x= self.compare_excel_sheet_dynamic_values(base_file, actual_file, sheet_name, **kwargs)
        self.asequal(len(x),1,msg+ ' -'+ str(x))
    
    def open_file(self, file_name):
        file =os.getcwd()+"\data\\"+file_name
        subprocess.run(["start", "", "/max", ""+file+""], shell=True)
         
    def create_lasso(self, x1, y1, x2, y2, pause=0):
        """
        utillobj.create_lasso(sx, sy, tx, ty, pause=1)
        """
        x1, y1, x2, y2=int(x1), int(y1), int(x2), int(y2)
        if sys.platform == 'linux':
            mouse_.press(int(x1), int(y1))
            time.sleep(3)
            pyautogui.moveTo(int(x2), int(y2), duration=3)
            time.sleep(3)
            mouse_.release(int(x2), int(y2))
        else:
            import uiautomation as automate
            mouse_obj=uisoup.mouse
            automate.Win32API.SetCursorPos(x1-9, y1-9)
            time.sleep(pause)
            mouse_obj.move(x1, y1)
            time.sleep(pause)
            automate.Win32API.mouse_event(automate.MouseEventFlags.LeftDown | automate.MouseEventFlags.Absolute, 0, 0, 0, 0)
            time.sleep(pause)
            mouse_obj.move(x2, y2)
            time.sleep(pause)
            automate.Win32API.mouse_event(automate.MouseEventFlags.Absolute | automate.MouseEventFlags.LeftUp, 0, 0, 0, 0)
    
    def drag_drop_using_pyautogui(self,source_elem,target_elem,**kwargs):
        """
        This function to drag and drop object from source to target
        :Usage: drag_drop_using_pyautogui(source_elem,target_elem, 9, 0, 9, 0)
        """
        if 'cord_type' in kwargs:
            cord_type=kwargs['cord_type']
        else:
            cord_type='middle'  
        if 'sx_offset' in kwargs:
            source_offset_x=kwargs['sx_offset']
        else:
            source_offset_x=0
        if 'sy_offset' in kwargs:
            source_offset_y=kwargs['sy_offset']
        else:
            source_offset_y=0 
        if 'tx_offset' in kwargs:
            target_offset_x=kwargs['tx_offset']
        else:
            target_offset_x=0
        if 'ty_offset' in kwargs:
            target_offset_y=kwargs['ty_offset']
        else:
            target_offset_y=0      
        pyautogui.FAILSAFE=False  
        source_obj=UtillityMethods.get_object_screen_coordinate(self, source_elem, coordinate_type=cord_type, x_offset=source_offset_x, y_offset=source_offset_y)
        source_obj_x=source_obj['x']
        source_obj_y=source_obj['y']
        target_obj=UtillityMethods.get_object_screen_coordinate(self, target_elem, coordinate_type=cord_type, x_offset=target_offset_x, y_offset=target_offset_y)
        target_obj_x=target_obj['x']
        target_obj_y=target_obj['y']
        time.sleep(2)
        pyautogui.mouseDown(source_obj_x,source_obj_y)
        time.sleep(15)
        pyautogui.click(target_obj_x, target_obj_y, button='left')
        time.sleep(2)
        pyautogui.dragRel(target_obj_x, target_obj_y, 4, button='left')
        time.sleep(5)
        
    def drag_drop_using_uisoup(self,source_elem,target_elem,**kwargs):
        """
        This function to drag and drop object from source to target
        :Usage: drag_drop_using_pyautogui(source_elem,target_elem, sx_offset=9, sy_offset=0, tx_offset=9, ty_offset=0)
        """
        src_cord_type=kwargs['src_cord_type'] if 'src_cord_type' in kwargs else 'middle'
        trg_cord_type=kwargs['trg_cord_type'] if 'trg_cord_type' in kwargs else 'middle'
        if 'sx_offset' in kwargs:
            source_offset_x=kwargs['sx_offset']
        else:
            source_offset_x=0
        if 'sy_offset' in kwargs:
            source_offset_y=kwargs['sy_offset']
        else:
            source_offset_y=0 
        if 'tx_offset' in kwargs:
            target_offset_x=kwargs['tx_offset']
        else:
            target_offset_x=0
        if 'ty_offset' in kwargs:
            target_offset_y=kwargs['ty_offset']
        else:
            target_offset_y=0      
        pyautogui.FAILSAFE=False  
        source_obj=UtillityMethods.get_object_screen_coordinate(self, source_elem, coordinate_type=src_cord_type, x_offset=source_offset_x, y_offset=source_offset_y)
        source_obj_x=source_obj['x']
        source_obj_y=source_obj['y']
        target_obj=UtillityMethods.get_object_screen_coordinate(self, target_elem, coordinate_type=trg_cord_type, x_offset=target_offset_x, y_offset=target_offset_y)
        target_obj_x=target_obj['x']
        target_obj_y=target_obj['y']
        time.sleep(2)
        #if self.browser=='Chrome':
        if sys.platform == 'linux':
                mouse_.press(int(source_obj_x), int(source_obj_y))
                time.sleep(3)
                pyautogui.moveTo(int(target_obj_x), int(target_obj_y), duration=3)
                time.sleep(3)
                mouse_.release(int(target_obj_x), int(target_obj_y))
        else:
            if Global_variables.browser_name in ['chrome']:
                time.sleep(5)
                mouse_delay=kwargs['mouse_speed'] if 'mouse_speed' in kwargs else 25
                path_ = os.path.join(root_path.ROOT_PATH, 'drag_drop_helper.exe '+str(source_obj_x)+' '+ str(source_obj_y)+' '+ str(target_obj_x)+' '+ str(target_obj_y)+' '+str(mouse_delay))
                procesobj = subprocess.Popen(path_)
                procesobj.wait()
                time.sleep(3)
                del procesobj
            else:
                mouse_obj=uisoup.mouse
                mouse_obj.click(source_obj_x,source_obj_y)
                time.sleep(5)
                mouse_obj.drag(source_obj_x,source_obj_y, target_obj_x, target_obj_y)
                time.sleep(2)
                mouse_obj.click(target_obj_x, target_obj_y)
                time.sleep(10)
        
    def seletct_tree_view_plus_node(self, parent_table_css, prodedure_path, click_type=0, **kwargs):
        """
        :param: parent_table_css='#WFServerProcDlg_procTree2 table'
        :param: prodedure_path='ibisamp->carinst'
        :param: click_type=0 or 1 ('0' for single click and '1' for double click)
        :usage: utillobj.seletct_tree_view_plus_node(parent_table_element, 'ibisamp->carinst')
        @author : Niranjan 
        """
        elem=(By.CSS_SELECTOR, parent_table_css +' tr')
        self._validate_page(elem)
        targets=prodedure_path.split('->')
        rows=self.driver.find_elements_by_css_selector(parent_table_css + " tr")
        time.sleep(1)
        UtillityMethods.default_left_click(self, object_locator=rows[0], **kwargs)
        time.sleep(1)
        for e in targets[:-1]:
            get_folder=True
            while get_folder:
                #if self.browser == 'Firefox':
                if Global_variables.browser_name in ['firefox']:
                    pyautogui.press('pagedown', pause=2)
                else:
                    action = ActionChains(self.driver)
                    action.send_keys(keys.Keys.PAGE_DOWN).perform()
                    time.sleep(2)
                    del action
                rows=self.driver.find_elements_by_css_selector(parent_table_css + " tr")
                time.sleep(1)
                tmp_list=[el.text.strip() for el in rows]
                if e in tmp_list:
                    rows[tmp_list.index(e)].click()
                    time.sleep(2)
                    tree_img_obj=rows[tmp_list.index(e)].find_element_by_css_selector("img[src*='tree-view-plus']")
                    UtillityMethods.default_left_click(self, object_locator=tree_img_obj, **kwargs)
                    time.sleep(3)
                    get_folder=False
                    break
            time.sleep(1)
        time.sleep(1)
        if len(targets)>1:
            get_procedure=True
            while get_procedure:
                #if self.browser == 'Firefox':
                if Global_variables.browser_name in ['firefox']:
                    pyautogui.press('pagedown', pause=2)
                else:
                    action = ActionChains(self.driver)
                    action.send_keys(keys.Keys.PAGE_DOWN).perform()
                    time.sleep(2)
                    del action
                rows=self.driver.find_elements_by_css_selector(parent_table_css + " tr")
                time.sleep(1)
                tmp_list=[el.text.strip() for el in rows]
                if targets[-1] in tmp_list:
                    if click_type == 1:
                        if Global_variables.browser_name in ['firefox']:
                            UtillityMethods.click_type_using_pyautogui(self, rows[tmp_list.index(targets[-1])], doubleClick=True, **kwargs)
                        else:
                            ActionChains(self.driver).double_click(rows[tmp_list.index(targets[-1])]).perform()
                    else:
                        UtillityMethods.default_left_click(self, object_locator=rows[tmp_list.index(targets[-1])], **kwargs)
#                         rows[tmp_list.index(targets[-1])].click()
                    get_procedure=False
                    break
        time.sleep(1)   
    
    def set_text_field_using_actionchains(self, field_elem, text_string, pause_time=1, **kwargs):
        """
        Params: field_elem = This is the text field element
        Usage: text_string = Text String to be used
        Usage: set_text_field_using_actionchains(field_elem,'This is a Text string')
        """
        pyautogui.FAILSAFE=False
        UtillityMethods.click_on_screen(self, field_elem, 'middle', 0)
        time.sleep(3)
        pyautogui.tripleClick()
        time.sleep(3)
        pyautogui.hotkey('ctrl','a')
        time.sleep(3)
        if Global_variables.browser_name in ['edge']:
            if sys.platform == 'linux':
                pykeyboard.tap_key(pykeyboard.backspace_key)
            else:
                keyboard.press('BACKSPACE')
        else :
            pyautogui.hotkey('del')
        time.sleep(pause_time)
        if 'pyautogui_type' in kwargs:
            pyautogui.typewrite(text_string, interval=0.3, pause=5)
        elif 'keyboard_type' in kwargs:
            if sys.platform == 'linux':
                pykeyboard.type_string(text_string, interval=1)
            else:
                keyboard.write(text_string, delay=1)
        else:
            action = ActionChains(self.driver) 
            action.send_keys_to_element(field_elem, text_string).perform()  
            del action      
        time.sleep(2)

    def select_item_from_ibfs_explorer_list(self, file_item_name, file_type=None, step_num='Step 00.00 :', **kwargs):
        """
        :param: file_item_name: 'ENIADefault_combine.sty' or 'FTP_Dlist'
        :usage: select_item_from_ibfs_explorer_list('ENIADefault_combine.sty')
        """
        ibfs_file_list_css="#paneIbfsExplorer_exList > div.bi-tree-view-body-content > table > tbody > tr"
        elem=(By.CSS_SELECTOR,ibfs_file_list_css)
        UtillityMethods._validate_page(self, elem)
        ibfs_file=self.driver.find_elements_by_css_selector(ibfs_file_list_css)
        UtillityMethods.default_left_click(self, object_locator=ibfs_file[0])
        get_file=True
        while get_file:
            action = ActionChains(self.driver)
            time.sleep(1)
            if Global_variables.browser_name in ['firefox']:
                pyautogui.press('pagedown', pause=1)
            else:
                action.send_keys(keys.Keys.PAGE_DOWN).perform()
            ibfs_file=self.driver.find_elements_by_css_selector(ibfs_file_list_css)
            time.sleep(1)
            for i in range(len(ibfs_file)):
                ibfs_file=self.driver.find_elements_by_css_selector(ibfs_file_list_css)
                if file_item_name in ibfs_file[i].text.split(" ")[0]:
                    theme_image_css="#paneIbfsExplorer_exList>div.bi-tree-view-body-content>table>tbody>tr:nth-child(" + str(i+1) + ") img"
                    theme_obj=self.driver.find_element_by_css_selector(theme_image_css)
                    UtillityMethods.default_left_click(self, object_locator=theme_obj, **kwargs)
                    get_file=False
                    break
                else:
                    pass
            del action
        time.sleep(1)
        selected_value=self.driver.find_element_by_css_selector("#IbfsOpenFileDialog7_cbFileName input").get_attribute("value")
        sty_check=selected_value == file_item_name
        UtillityMethods.asequal(self,True, sty_check, step_num+" Verify whether the " + file_item_name + " is selected.")
        if file_type != None:
            UtillityMethods.select_combobox_item(self,'IbfsOpenFileDialog7_cbFilterType', file_type)
            time.sleep(1)
        self.driver.find_element_by_id("IbfsOpenFileDialog7_btnOK").click()
        time.sleep(1)
        
    def click_dialog_button(self, dialog, btn_name):
        """
        :param: dialog: "div[id^='BiDialog']"
        :param: btn_name: "OK"
        :Usage: utillobj.click_dialog_button("div[id^='BiDialog']", "OK")
        """        
        dialog_css=dialog + " div[id^='BiButton']"
        dialog_btns=self.driver.find_elements_by_css_selector(dialog_css)
        btn_text_list=[el.text.strip() for el in dialog_btns]
        dialog_btns[btn_text_list.index(btn_name)].click()
        time.sleep(5)
    
    def click_type_using_pyautogui(self, obj_locator,**kwargs):
        """
        :kwargs: default_move=False/True(if False - mouse not moved, if True - mouse will move)
        :param: object_locator: object location
        :param: source_x=interger value only either postive or negative
        :param: source_y=interger value only either postive or negative
        :param: leftClick=True
        :param: doubleClick=True
        :param: rightClick=True
        :Usage: click_type_using_pyautogui(object_locator, 50, 50, leftClick=True)
        """
        if 'cord_type' in kwargs:
            cord_type=kwargs['cord_type']
        else:
            cord_type='middle'
        UtillityMethods.click_on_screen(self, obj_locator, coordinate_type=cord_type, **kwargs)
        if 'move' in kwargs:
            UtillityMethods.click_on_screen(self, obj_locator, coordinate_type='start', **kwargs)
        if 'leftClick' in kwargs:
            UtillityMethods.click_on_screen(self, obj_locator, coordinate_type=cord_type, click_type=0, **kwargs)
            time.sleep(2)
        if 'doubleClick' in kwargs:
            UtillityMethods.click_on_screen(self, obj_locator, coordinate_type=cord_type, click_type=2, **kwargs)
        if 'rightClick' in kwargs:
            UtillityMethods.click_on_screen(self, obj_locator, coordinate_type=cord_type, click_type=1, **kwargs)
            time.sleep(3)
        
    
    def drag_to_using_pyautogui(self,source_elem,target_elem,**kwargs):
        """
        :param: source_elem: Where you want to drag source location
        :param: target_elem: Where you want to drag target location
        :param: source_offset_x=interger value only either postive or negative
        :param: source_offset_y=interger value only either postive or negative
        :param: target_offset_x=interger value only either postive or negative
        :param: target_offset_y=interger value only either postive or negative
        :param: target_offset_y=interger value only either postive or negative
        :Usage: drag_to_using_pyautogui(self, source, target, 0, 50, 0, 50)
        """    
        cord_type=kwargs['cord_type'] if 'cord_type' in kwargs else 'middle' 
        source_offset_x=kwargs['sx_offset'] if 'sx_offset' in kwargs else 0
        source_offset_y=kwargs['sy_offset'] if 'sy_offset' in kwargs else 0
        target_offset_x=kwargs['tx_offset'] if 'tx_offset' in kwargs else 0
        target_offset_y=kwargs['ty_offset'] if 'ty_offset' in kwargs else 0
        pyautogui.FAILSAFE=False  
        source_obj=UtillityMethods.get_object_screen_coordinate(self, source_elem, coordinate_type=cord_type, x_offset=source_offset_x, y_offset=source_offset_y)
        source_obj_x=source_obj['x']
        source_obj_y=source_obj['y']
        target_obj=UtillityMethods.get_object_screen_coordinate(self, target_elem, coordinate_type=cord_type, x_offset=target_offset_x, y_offset=target_offset_y)
        target_obj_x=target_obj['x']
        target_obj_y=target_obj['y']
        time.sleep(2)
        pyautogui.mouseDown(source_obj_x,source_obj_y)
        time.sleep(15)
        pyautogui.moveTo(target_obj_x, target_obj_y)
        time.sleep(2)
        pyautogui.mouseUp(target_obj_x, target_obj_y)
        time.sleep(5)
    
    def drag_using_uisoup(self,source_elem,target_elem,**kwargs):
        """
        :param: source_elem: Where you want to drag source location
        :param: target_elem: Where you want to drag target location
        :param: source_offset_x=interger value only either postive or negative
        :param: source_offset_y=interger value only either postive or negative
        :param: target_offset_x=interger value only either postive or negative
        :param: target_offset_y=interger value only either postive or negative
        :param: target_offset_y=interger value only either postive or negative
        :Usage: drag_using_mouse_lib(self, source, target, 0, 50, 0, 50)
        """  
        cord_type=kwargs['cord_type'] if 'cord_type' in kwargs else 'middle' 
        source_offset_x=kwargs['sx_offset'] if 'sx_offset' in kwargs else 0
        source_offset_y=kwargs['sy_offset'] if 'sy_offset' in kwargs else 0
        target_offset_x=kwargs['tx_offset'] if 'tx_offset' in kwargs else 0
        target_offset_y=kwargs['ty_offset'] if 'ty_offset' in kwargs else 0
        source_obj=UtillityMethods.get_object_screen_coordinate(self, source_elem, coordinate_type=cord_type, x_offset=source_offset_x, y_offset=source_offset_y)
        source_obj_x=source_obj['x']
        source_obj_y=source_obj['y']
        target_obj=UtillityMethods.get_object_screen_coordinate(self, target_elem, coordinate_type=cord_type, x_offset=target_offset_x, y_offset=target_offset_y)
        target_obj_x=target_obj['x']
        target_obj_y=target_obj['y']
        uisoup.mouse.drag(source_obj_x, source_obj_y, target_obj_x, target_obj_y)
       
    def mouse_action_using_pyautogui(self, object_locator,source_x=0, source_y=0, pause=15, **kwargs):
        """
        This function is helpful, if you need to use mouseUp, mouseMove, mouseDown.
        """
        pyautogui.FAILSAFE=False
        cord_type=kwargs['cord_type'] if 'cord_type' in kwargs else 'middle'
        get_loc = UtillityMethods.get_object_screen_coordinate(self, object_locator, coordinate_type=cord_type, **kwargs)
        x=get_loc['x']
        y=get_loc['y']
        if 'mouse_down' in kwargs:
            pyautogui.mouseDown(x, y)
            time.sleep(pause)
        if 'mouse_up' in kwargs:
            pyautogui.mouseUp(x + source_x, y + source_y)
            time.sleep(2)
        if 'mouse_move' in kwargs:
            pyautogui.moveTo(x + source_x, y + source_y)
            time.sleep(2)
    
    def default_left_click(self, browser_style='firefox',**kwargs):
        """
        Usage: 
        default_left_click(self, object_locator=obj_cell_css,action_move=True)
        default_left_click(self, object_locator=obj_cell_css,action_move_offset=True,ax_offset=1,ay_offset=1)
        default_left_click(self, object_locator=obj_cell_css)
        default_left_click(self, object_locator=obj_cell_css,x_offset=0,y_offset=0)
        """
        pyautogui.FAILSAFE=False
        #if self.browser.lower() in ['ie', 'firefox'] or self.browser.lower()==browser_style:
        if Global_variables.browser_name in ['ie', 'firefox', 'edge'] or Global_variables.browser_name==browser_style:
            #UtillityMethods.click_type_using_pyautogui(self, kwargs['object_locator'], leftClick=True, **kwargs)
            coordinates=UtillityMethods.get_object_screen_coordinate(self, kwargs['object_locator'], coordinate_type='middle', **kwargs)
            if sys.platform == 'linux':
                mouse_.click(int(coordinates['x']), int(coordinates['y']))
            else:
                uisoup.mouse.click(coordinates['x'], coordinates['y'])
        else:
            if 'action_move_only' in kwargs:
                ActionChains(self.driver).move_to_element(kwargs['object_locator']).perform()
            elif 'action_move' in kwargs:
                ActionChains(self.driver).move_to_element(kwargs['object_locator']).click().perform()
            elif 'active_move_offset' in kwargs:    
                ActionChains(self.driver).move_to_element_with_offset(kwargs['object_locator'],kwargs['ax_offset'],kwargs['ay_offset']).click().perform()
            else:
                kwargs['object_locator'].click()
    
    def default_click(self, obj_locator, click_option=0, **kwargs):            
        """
        Usage: 
        param: click_option=0(left Click), 1(right click), 2(double click) 
        default_click(self, object_locator=obj_cell_css,click_option=0)
        @author: Magesh
        """
        if click_option == 1:
            if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
                UtillityMethods.click_on_screen(self, obj_locator, coordinate_type='middle',click_type=1,**kwargs)
            else:
                action1 = ActionChains(self.driver)
                action1.move_to_element(obj_locator).context_click(obj_locator).perform()
        elif click_option == 2:
            if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
                UtillityMethods.click_on_screen(self, obj_locator, coordinate_type='middle',click_type=2,**kwargs)
            else:
                action1 = ActionChains(self.driver)
                action1.move_to_element(obj_locator).double_click(obj_locator).perform()
        else:
            if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
                UtillityMethods.click_on_screen(self, obj_locator, coordinate_type='right', x_offset=-7, click_type=0,**kwargs)
            else:
                action1 = ActionChains(self.driver)
                action1.move_to_element(obj_locator).click(obj_locator).perform()
            
    def default_click_with_firefox_update(self, obj_locator, click_option=0, **kwargs):            
        """
        Selenium click for Firefox left click(0)
        Usage: 
        param: click_option=0(left Click), 1(right click), 2(double click) 
        default_click_with_firefox_update(self, object_locator=obj_cell_css,click_option=0)
        
        """
        if click_option == 1:
            if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
                UtillityMethods.click_on_screen(self, obj_locator, coordinate_type='middle',click_type=1,**kwargs)
            else:
                action1 = ActionChains(self.driver)
                action1.move_to_element(obj_locator).context_click(obj_locator).perform()
        elif click_option == 2:
            if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
                UtillityMethods.click_on_screen(self, obj_locator, coordinate_type='middle',click_type=2,**kwargs)
            else:
                action1 = ActionChains(self.driver)
                action1.move_to_element(obj_locator).double_click(obj_locator).perform()
        else:
            if Global_variables.browser_name == "firefox":
                obj_locator.click()
            else:
                action1 = ActionChains(self.driver)
                action1.move_to_element(obj_locator).click(obj_locator).perform()
    
    def get_attribute_value(self, elem,*args):
        """
        This function will return only one value for one element's attribute/property.
        elem: Any web element . Eg.- self.driver.find_element_by_css_selector("#avfToValue input")
        args: 'int_value' or 'float_value' or 'text_value' or 'text' or dom_visible_text
        elem1=driver.find_element_by_css_selector("#avfToValue input")
        d=get_css_value(elem1, "int_value")
        returns: d={'int_value': '205'}
        returns: d={'float_value': '21.58'}
        returns: d={'text_value': 'Jan'}
        """ 
        dict_obj={}
        for arg in args:
            if arg == 'int_value':
                dict_obj[arg]=re.match('(\d+).*', elem.get_attribute('value')).group(1)
            elif arg == 'posneg_int_value':
                dict_obj[arg]=re.match('([+-]?\d+).*', elem.get_attribute('value')).group(1)
            elif arg == 'float_value':
                dict_obj[arg]=re.match('(\d+\.\d+).*', elem.get_attribute('value')).group(1)
            elif arg == 'posneg_float_value':
                dict_obj[arg]=re.match('([+-]?\d+\.\d+).*', elem.get_attribute('value')).group(1)
            elif arg == 'text_value': # for input tag
                dict_obj[arg]=elem.get_attribute('value')
            elif arg == 'text': #dom text
                dict_obj[arg]=elem.get_attribute('text')
            elif arg == 'dom_visible_text': #dom visible text
                dict_obj[arg]=elem.text.strip()
            else:
                dict_obj[arg]=elem.get_attribute(arg)
        return(dict_obj)
        
    def get_css_value(self, elem,*args):
        """
        This function will return a dictionary object of css properties.
        elem: Any web element . Eg.- self.driver.find_element_by_css_selector("#resultArea div[id^='BoxLayoutMiniWindow']")
        args: 'top' or 'left'
        d=get_css_value(elem, "left", "top", "width", "height")        
        returns: d={'top': '205', 'height': '586', 'left': '0'}
        """ 
        dict_obj={}
        for arg in args:
            dict_obj[arg]=re.match('(\d+).*', elem.value_of_css_property(arg)).group(1)
        return(dict_obj)
    
    def drag_drop_using_Sikuli(self, drag_image, dropimage):
        """
        This function do drag and drop operation using sikuli
        :param: drag_image='image1'
        :param: dropimage='image2'
        :Usage: drag_drop_using_Sikuli('image1', 'image2') 
        """
        siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
        src_loc=os.getcwd() +"\\sikuli\\dnd --arg " + drag_image + " " + dropimage
        os.system(siklui_loc+' -r '+ src_loc)
        time.sleep(135)

    def verify_picture_using_sikuli(self,image_name, msg, tolerance=0.70):
        """      
        Description:- In order to verify any unavoidable image comparison situation, then keep your image under data folder.
        :param image_name='image1.png'
        :param msg='your message'
        :Usage: verify_picture_using_sikuli('image1.png', 'your message') 
        """
        if sys.platform == 'linux':
            print('Sikuli Implementation TODO')
            UtillityMethods.asequal(self, False, True, msg)
        else:
            if type(tolerance) is not float:
                raise TypeError('Please pass tolerance value in floating type. Example tolerance=0.70')
            siklui_loc=os.path.join(sikuli_root.ROOT_PATH, 'runsikulix.cmd')
            image_name=os.getcwd() +"\\data\\"+image_name
            src_loc = os.path.join(sikuli_root.ROOT_PATH, 'verify_picture --arg {0} {1}'.format(image_name, tolerance))
            p=subprocess.Popen(siklui_loc+' -r '+ src_loc, shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
            result = re.match('.*:(.*\d+).*',str(p.communicate()[0]).replace('\r\n','')).group(1).replace(' ','')
            UtillityMethods.asequal(self, '123', result, msg) 
    
    def verify_regional_picture_using_sikuli(self, image_name, msg, parent_element=None, image_parent_elem_x=0, image_parent_elem_y=0, image_parent_elem_width=0, image_parent_elem_height=0):
        """      
        Description:- In order to verify any unavoidable image comparison situation, then keep your image under data folder.
        :param image_name='image1.png'
        :param msg='your message'
        :param parent_element='parent object of the Image'
        :Usage: verify_regional_picture_using_sikuli('blog.png', "Step 5.1.1: Verify", parent_element=obj)
        @author: Aftab_Alam_Khan 
        """
        if sys.platform == 'linux':
            print('Sikuli Implementation TODO')
            UtillityMethods.asequal(self, False, True, msg)
        else:
            if parent_element!=None:
                try:
                    image_elem_x=int(parent_element.location['x']) + int(image_parent_elem_x)
                    image_elem_y=int(parent_element.location['y']) + int(image_parent_elem_y)
                    image_elem_width=int(parent_element.size['width']) + int(image_parent_elem_width)
                    image_elem_height=int(parent_element.size['height']) + int(image_parent_elem_height)
                except NoSuchElementException:
                    raise AttributeError("NoSuchElement Exist for the iamge element {0}".format(image_name))
            else:
                image_elem_x=int(image_parent_elem_x)
                image_elem_y=int(image_parent_elem_y)
                image_elem_width=int(image_parent_elem_width)
                image_elem_height=int(image_parent_elem_height)
            siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
            expected_image_name=os.getcwd() +"\\data\\"+image_name
            arguments="{0} {1} {2} {3} {4}".format(expected_image_name, image_elem_x, image_elem_y, image_elem_width, image_elem_height)
            src_loc=os.getcwd() +"\\common\\lib\\sikuli\\verify_regional_picture --arg " + arguments
            p=subprocess.Popen(siklui_loc+' -r '+ src_loc, shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
            result = re.match('.*:(.*\d+).*',str(p.communicate()[0]).replace('\r\n','')).group(1).replace(' ','')
            UtillityMethods.asequal(self, '123', result, msg) 
         
    def switch_to_window_handle_test(self, wndnum, **kwargs):
        old_windows=UtillityMethods.windows
        if 'custom_windows' in kwargs:
            old_windows=kwargs['custom_windows']
        new_windows=self.driver.window_handles
        new_set=set(new_windows)
        old_set=set(old_windows)
        diff_set=new_set-old_set
        last_window=list(diff_set)
        time.sleep(2)
        if wndnum>0:
            self.driver.switch_to.window(last_window[0])
        else:
            self.driver.switch_to.window(old_windows[-1])
        time.sleep(4)

    def get_browser_height(self):
        """
        This will return the actual working area of Browser height
        :Usage browser_width, browser_height = UtillityMethods.get_browser_height(self)
        @author: AAkhan
        """
        dict_obj={}
        if sys.platform == 'linux':
            dict_obj['browser_width']=self.driver.execute_script("return window.screenX")
            browser_screeny_=self.driver.execute_script("return window.screenY")
            outer_height = self.driver.execute_script("return window.outerHeight")
            inner_height = self.driver.execute_script("return window.innerHeight")
            dict_obj["browser_height"]=int(browser_screeny_ + (outer_height - inner_height))
        else:    
            for _time in range(72):
                try:
                    screen_width = self.driver.execute_script("return screen.width")
                    screen_height = self.driver.execute_script("return screen.height")
                    outer_height = self.driver.execute_script("return window.outerHeight")
                    availWidth = self.driver.execute_script("return screen.availWidth")
                    availHeight = self.driver.execute_script("return screen.availHeight")
                    innerWidth = self.driver.execute_script("return window.innerWidth || document.documentElement.clientWidth || document.body.clientWidth;")
                    innerHeight = self.driver.execute_script("return document.body.scrollHeight || window.innerHeight;")
                except TimeoutException:
                    time.sleep(5)
                except Exception as e:
                    print("Exception occur in get_current_browser_specification- {0}".format(e))
    #         innerWidth = self.driver.execute_script("return window.innerWidth|| document.documentElement.clientWidth|| document.body.clientWidth;")
    #         innerHeight = self.driver.execute_script("return window.innerHeight|| document.documentElement.clientHeight|| document.body.clientHeight;")
            dict_obj['browser_width'] = availWidth - innerWidth
            dict_obj['browser_height'] = availHeight - innerHeight
            dict_obj['outer_height'] = screen_height - outer_height
        return (dict_obj)

    def get_object_screen_coordinate(self, elem, coordinate_type='start', **kwargs):
        """
        elem:- This is the object for which x,y coordinate to be returned.
        coordinate_type='start' OR 'top_middle' OR 'top_right' OR 'left' OR 'middle' OR 'right' OR 'bottom_left' OR 'bottom_middle' OR 'bottom_right' OR 'offset'
        The return type is a dictionary like = {'x': 524, 'y': 993}
        """
        dict_obj={}
        wait_time=kwargs['pause'] if 'pause' in kwargs else 0
        mouse_move_duration=kwargs['mouse_duration'] if 'mouse_duration' in kwargs else 0.5
        time.sleep(wait_time)
        x_offset=kwargs['x_offset'] if 'x_offset' in kwargs else 0
        y_offset=kwargs['y_offset'] if 'y_offset' in kwargs else 0
        elem_x=elem.location['x'] + Global_variables.current_working_area_browser_x
        elem_y=elem.location['y'] + Global_variables.current_working_area_browser_y
        elem_h=elem.size['height']
        elem_w=elem.size['width']
        time.sleep(wait_time)
        if coordinate_type=='start':
            dict_obj['x'] = elem_x + x_offset
            dict_obj['y'] = elem_y+ y_offset
        if coordinate_type=='top_middle':
            dict_obj['x'] = elem_x + (elem_w/2) + x_offset
            dict_obj['y'] = elem_y + y_offset
        if coordinate_type=='top_right':
            dict_obj['x'] = elem_x + elem_w + x_offset
            dict_obj['y'] = elem_y + y_offset
        if coordinate_type=='left':
            dict_obj['x'] = elem_x + x_offset
            dict_obj['y'] = elem_y + (elem_h/2) + y_offset
        if coordinate_type=='middle':
            dict_obj['x'] = elem_x + (elem_w/2) + x_offset
            dict_obj['y'] = elem_y + (elem_h/2) + y_offset
        if coordinate_type=='right':
            dict_obj['x'] = elem_x + elem_w + x_offset
            dict_obj['y'] = elem_y + (elem_h/2) + y_offset
        if coordinate_type=='bottom_left':
            dict_obj['x'] = elem_x + x_offset
            dict_obj['y'] = elem_y + elem_h + y_offset
        if coordinate_type=='bottom_middle':
            dict_obj['x'] = elem_x + (elem_w/2) + x_offset
            dict_obj['y'] = elem_y + elem_h + y_offset
        if coordinate_type=='bottom_right':
            dict_obj['x'] = elem_x + elem_w + x_offset
            dict_obj['y'] = elem_y + elem_h + y_offset
        if coordinate_type=='offset':
            dict_obj['x'] =  x_offset
            dict_obj['y'] =  y_offset
        time.sleep(wait_time)
        if 'move' in kwargs:
            pyautogui.moveTo(dict_obj['x'], dict_obj['y'], mouse_move_duration)
            time.sleep(wait_time)
        return(dict_obj)
    
    def click_on_screen(self, parent_elem, coordinate_type, click_type=3, **kwargs):
        """
        kwargs pause=0,1,2,.... (any Integer number)
        kwargs mouse_duration=0.5,1,2.2,.... (any Integer number or Float number)
        param click_type=0(left Click), 1(right click), 2(double click), 3(move)default
        :Usage click_on_screen(self, obj_locator, coordinate_type='middle', **kwargs)
            OR
        :Usage click_on_screen(self, obj_locator, coordinate_type='middle', x_offset=0, y_offset=0)
            OR
        :Usage click_on_screen(self, obj_locator, coordinate_type='middle', javascript_marker_enable=True)
        """
        before_click_time_sleep_in_seconds=1
        wait_time=kwargs['pause'] if 'pause' in kwargs else 0
        mouse_move_duration=kwargs['mouse_duration'] if 'mouse_duration' in kwargs else 0.5
        if 'javascript_marker_enable' in kwargs:
            coordinates=UtillityMethods.enable_marker_using_javascript(self, parent_elem, coordinate_type=coordinate_type, **kwargs)
        else:
            coordinates=UtillityMethods.get_object_screen_coordinate(self, parent_elem, coordinate_type=coordinate_type, **kwargs)
        x=coordinates['x']
        y=coordinates['y']
        time.sleep(wait_time)
        if Global_variables.browser_name in ['ie', 'edge']:
            uisoup.mouse.move(x, y)
        else:
            pyautogui.moveTo(x, y, mouse_move_duration)
        time.sleep(wait_time)
        if click_type==0:
            time.sleep(before_click_time_sleep_in_seconds)
            if Global_variables.browser_name in ['ie', 'edge']:
                uisoup.mouse.click(x, y)
            else :
                pyautogui.click(x,y, button='left')
            time.sleep(wait_time)
        if click_type==1:
            time.sleep(before_click_time_sleep_in_seconds)
            if Global_variables.browser_name in ['ie', 'edge']:
                uisoup.mouse.click(x, y, button_name=uisoup.mouse.RIGHT_BUTTON)
            else:
                pyautogui.click(x,y, button='left')
                time.sleep(wait_time)
                time.sleep(before_click_time_sleep_in_seconds)
                pyautogui.click(x,y, button='right')
            time.sleep(wait_time)
        if click_type==2:
            time.sleep(before_click_time_sleep_in_seconds)
            if Global_variables.browser_name in ['ie', 'edge']:
                uisoup.mouse.double_click(x, y)
            else :
                pyautogui.doubleClick(x,y, button='left')
            time.sleep(wait_time)
            
    def enable_marker_using_javascript(self, parent_elem, coordinate_type='middle', **kwargs):
        """
        :param parent_elem=driver.find_element_by_css_selector("#MAINTABLE_wbody0 [class*='riser!s0!g0!mwedge!r0!c1'])
        :param coordinate_type="start"
        :usage enable_marker_using_javascript(self, parent_elem, coordinate_type=coordinate_type)
        """
        self.driver.execute_script("arguments[0].style.display='inline';",parent_elem)
        time.sleep(5)
        coordinate=UtillityMethods.get_object_screen_coordinate(self, parent_elem, coordinate_type=coordinate_type, **kwargs)
        self.driver.execute_script("arguments[0].style.display='';",parent_elem)
        '''self.driver.set_window_size(1600, 1000)
        time.sleep(3)
        self.driver.maximize_window()
        time.sleep(7)'''
        return(coordinate)
        
    def drag_drop_on_screen(self, **kwargs):
        """drag_drop_on_screen(sx_offset=10,sy_offset=0,tx_offset=0,ty_offset=0)
           @author: AAkhan
        """
        source_obj_x=kwargs['sx_offset'] if 'sx_offset' in kwargs else 0
        source_obj_y=kwargs['sy_offset'] if 'sy_offset' in kwargs else 0
        target_obj_x=kwargs['tx_offset'] if 'tx_offset' in kwargs else 0
        target_obj_y=kwargs['ty_offset'] if 'ty_offset' in kwargs else 0
        time.sleep(2)
        if sys.platform == 'linux':
                mouse_.press(int(source_obj_x), int(source_obj_y))
                time.sleep(3)
                pyautogui.moveTo(int(target_obj_x),int(target_obj_y), duration=3)
                time.sleep(3)
                mouse_.release(int(target_obj_x), int(target_obj_y))
        else:
            if Global_variables.browser_name=='chrome':
                time.sleep(5)
                mouse_delay=kwargs['mouse_speed'] if 'mouse_speed' in kwargs else 25
                path_ = os.path.join(root_path.ROOT_PATH, 'drag_drop_helper.exe '+str(source_obj_x)+' '+ str(source_obj_y)+' '+ str(target_obj_x)+' '+ str(target_obj_y)+' '+str(mouse_delay))
                procesobj = subprocess.Popen(path_)
                procesobj.wait()
                time.sleep(3)
                del procesobj
            else:
                mouse_obj=uisoup.mouse
                mouse_obj.click(source_obj_x,source_obj_y)
                time.sleep(5)
                mouse_obj.drag(source_obj_x,source_obj_y, target_obj_x, target_obj_y)
                mouse_obj.release_button()
#                 time.sleep(2)
#                 mouse_obj.click(target_obj_x, target_obj_y)
#                 time.sleep(10)
#             
    def verify_print_dialog_and_click(self, msg, image='cancel', title_name='Print'):
        status=False
        for x in range(0,6):
            try:
                val=pywinauto.findwindows.find_window(title_re=".*"+title_name+"*.")
                os.system(os.getcwd()+"\\common\\lib\\Verify_print_dialog_open.exe "+ self.browser)
                if val >0:
                    status=True
                break
            except:
                pass
            time.sleep(2)
        UtillityMethods.asequal(self, status, True, msg)
    
    def verify_object_visible(self, css, visible, msg, **kwargs):
        """
        :params css="#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']"
        :params visible=True Or False
        :params msg='Step 5: Filter Button Visible'
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", True, 'Step 5: Filter Button Visible')
                OR
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", False, 'Step 5: Filter Button Removed')
        """
        try:
            if 'elem_obj' in kwargs:
                status = kwargs['elem_obj'].is_displayed()
            else:
                status = self.driver.find_element_by_css_selector(css).is_displayed()
        except:
            status = False
        UtillityMethods.asequal(self, status, visible, msg)
        
    def verify_object_visible_by_xpath(self, xpath, visible, msg, **kwargs):
        """
        :params css="#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']"
        :params visible=True Or False
        :params msg='Step 5: Filter Button Visible'
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", True, 'Step 5: Filter Button Visible')
                OR
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", False, 'Step 5: Filter Button Removed')
        """
        try:
            if 'elem_obj' in kwargs:
                status = kwargs['elem_obj'].is_displayed()
            else:
                status = self.driver.find_element_by_xpath(xpath).is_displayed()
        except:
            status = False
        UtillityMethods.asequal(self, status, visible, msg)
    

    def get_frame_height(self, **kwargs):
        """
        This function return back frame height in IA
        """
        path1=kwargs['css_path1'] if 'css_path1' in kwargs else "#windowOptionsPane"
        path2=kwargs['css_path2'] if 'css_path2' in kwargs else "#sbMain"
        dict_obj={}
        ia_frame_height1=self.driver.find_element_by_css_selector(path1)
        ia_frame_height2=self.driver.find_element_by_css_selector(path2)
        dict_obj['height1'] = ia_frame_height1.size['height']
        dict_obj['height2'] = ia_frame_height2.size['height']
        dict_obj['ia_frame_height'] = dict_obj['height1'] + dict_obj['height2']
        return (dict_obj)
    

    def switch_to_frame(self, pause=0, **kwargs):
        """
        Switches focus to the specified frame
        :param pause=0,1,2...        (pass Interger/Float number)
        :kwargs frame_css="[id^='ReportIframe']"   (pass frame css)
        :kwargs frame_height_value=0        (Interger number)
        :Usage switch_to_frame(pause=2, frame_css='frame[src]')
            OR
        :Usage switch_to_frame(pause=2, frame_css='frame[src]', frame_height_value=0)
        """
        if 'frame_css' in kwargs:
            path_css=kwargs['frame_css']
#             excess_outer_grayed_width=0
#             excess_outer_grayed_height=0
        else:
            path_css="[id^='ReportIframe']"
        frame_element_obj = self.driver.find_element_by_css_selector(path_css)
        frame_actual_location = UtillityMethods.get_object_screen_coordinate(self, frame_element_obj, coordinate_type='start')
        WebDriverWait(self.driver, 100).until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, path_css)))
        time.sleep(pause)
        Global_variables.current_working_area_browser_x=frame_actual_location['x']
        Global_variables.current_working_area_browser_y=frame_actual_location['y']
        time.sleep(pause)
#             screen_width=self.driver.execute_script("return window.innerWidth|| document.documentElement.clientWidth|| document.body.clientWidth;")
#             ia_tool_width=self.driver.execute_script("return document.getElementById('queryViewPane').clientWidth;")
#             excess_outer_grayed_width=screen_width-ia_tool_width
#             innerHeight = self.driver.execute_script("return window.innerHeight|| document.documentElement.clientHeight|| document.body.clientHeight;")
#             ia_tool_toptoolbar_height=self.driver.execute_script("return document.getElementById('applicationToolBarBox').clientHeight;")
#             ia_tool_ribbon_height=self.driver.execute_script("return document.getElementById('HomeTab').offsetHeight;")
#             ia_tool_query_height=self.driver.execute_script("return document.getElementById('queryViewPane').clientHeight;")
#             ia_tool_status_height=self.driver.execute_script("return document.getElementById('sbMain').offsetHeight;")
#             excess_outer_grayed_height=innerHeight-(ia_tool_toptoolbar_height+ia_tool_ribbon_height+ia_tool_query_height+ia_tool_status_height)
#         if 'frame_height_value' in kwargs:
#             frame_height=kwargs['frame_height_value']
#         else:
#             get_frame_elem=UtillityMethods.get_frame_height(self)
#             frame_height=get_frame_elem['ia_frame_height']
#         WebDriverWait(self.driver, 100).until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, path_css)))
#         time.sleep(pause)
#         get_browser_height = UtillityMethods.get_browser_height(self)
#         position_x = get_browser_height['browser_width'] - excess_outer_grayed_width
#         position_y = get_browser_height['browser_height'] - (frame_height + excess_outer_grayed_height)
#         UtillityMethods.browser_x=position_x
#         UtillityMethods.browser_y=position_y
#         time.sleep(pause)
        
    def switch_to_default_content(self, pause=0):
        """
        Switches focus to the default content
        :param pause=0,1,2...        (pass Interger/Float number)
        :Usage switch_to_default_content(pause=1)
        """
        self.driver.switch_to.default_content()
        time.sleep(pause)
        core_utilobj.update_current_working_area_browser_specification(self)
        time.sleep(pause)
        
    def verify_popup(self, css, msg, **kwargs):
        """
        params: css,msg
        kwargs: caption_css="div[id^='QbDialog'][style*='inherit'] [class*='active'] [class*='caption'] [class*='bi-label']"
        kwargs: caption_text="Location Type"
        :Usage verify_popup(css, "Step04: Verify dialog box displayed", caption_css=cap)
        """        
        UtillityMethods.verify_object_visible(self, css, True, msg)
        if 'caption_css' in kwargs:
            elem1=self.driver.find_element_by_css_selector(kwargs['caption_css'])
            d=UtillityMethods.get_attribute_value(self,elem1, "dom_visible_text")
            UtillityMethods.asequal(self, d['dom_visible_text'], kwargs['caption_text'], "StepX: Verify Caption text in popup")
        if 'popup_text_css' in kwargs:
            elem1=self.driver.find_element_by_css_selector(kwargs['popup_text_css'])
            d=UtillityMethods.get_attribute_value(self,elem1, "dom_visible_text")
            UtillityMethods.asin(self, kwargs['popup_text'], d['dom_visible_text'], "StepX: Verify text in popup")
        if 'popup_text_elem' in kwargs:
            elem1=kwargs['popup_text_elem']
            d=UtillityMethods.get_attribute_value(self,elem1, "dom_visible_text")
            UtillityMethods.asin(self, kwargs['popup_text'], d['dom_visible_text'], "StepX: Verify text in popup")
            
    def verify_notify_popup(self, notify_css ="div[class^='notify-pop'][class*='pop-top']", notify_text="Favorites added", notify_bg_color='fruit_salad', notify_opacity="0.8", msg="StepX"):
        '''
        Desc: This function will verify notify popup text, background transparency and background color
        usage: verify_favorites_notify_popup("Step21")
        '''
        UtillityMethods.synchronize_with_number_of_element(self, notify_css, 1, 15)
        actual_password_changed_text = self.driver.find_element_by_css_selector(notify_css).text.strip()
        actual_bg_transparent = self.driver.find_element_by_css_selector(notify_css).value_of_css_property('opacity')
        bg_color = self.driver.find_element_by_css_selector(notify_css).value_of_css_property('background-color')
        actual_password_changed_background_color = Color.from_string(bg_color).rgba
        expected_color=UtillityMethods.color_picker(self, notify_bg_color, 'rgba')
        UtillityMethods.asequal(self, expected_color, actual_password_changed_background_color, msg+": -Verify notify popup background color")
        UtillityMethods.asequal(self, notify_text, actual_password_changed_text, msg+": -Verify notify Popup text") 
        UtillityMethods.asequal(self, notify_opacity, actual_bg_transparent, msg+": -Verify notify Popup background transparency")
    
#     def expand_domain_folders_in_open_dialog(self, target_table_path):
#         """
#         syntax:- expand_folder_in_open_dialog("new_retail->dimensions")
#         """
#         targets=target_table_path.split('->')
#         open_dialog_file_name="#IbfsOpenFileDialog7_cbFileName"
#         elem1=(By.CSS_SELECTOR, open_dialog_file_name)
#         self._validate_page(elem1)
#         time.sleep(4)  
#         if len(targets)>0:
#             apps_css="#paneIbfsExplorer_exTree > div.bi-tree-view-body-content table tr"
#             for app in targets[:-1]:
#                 x=[el.text.strip() for el in self.driver.find_elements_by_css_selector(apps_css)]
#                 apps=self.driver.find_elements_by_css_selector(apps_css)
#                 apps[x.index(app)].find_element_by_css_selector("img[src*='triangle']").click()
#             time.sleep(5)
#             get_folder=True
#             counter=0
#             while get_folder:
#                 x=[el.text.strip() for el in self.driver.find_elements_by_css_selector(apps_css)]
#                 if targets[-1] in x:
#                     apps=self.driver.find_elements_by_css_selector(apps_css)
#                     apps[x.index(targets[-1])].find_element_by_css_selector("img[src*='folder']").click()
#                     time.sleep(5)
#                     get_folder=False
#                     break
#                 else:
#                     counter+=1
#                     action = ActionChains(self.driver)
#                     action.send_keys(keys.Keys.PAGE_DOWN).perform()
#                     time.sleep(5)
#                     del action
#                     if counter==10:
#                         break
#                 time.sleep(1)
#             time.sleep(1)
            
    def expand_domain_folders_in_open_dialog(self, target_table_path):
        """
        syntax:- expand_folder_in_open_dialog("new_retail->dimensions")
        #paneIbfsExplorer_exTree > div.bi-tree-view-body-content table tr img[src*='folder']
        """
        elem=self.driver.find_element_by_css_selector("#paneIbfsExplorer_exTree")
        UtillityMethods.click_on_screen(self, elem, 'middle')
        time.sleep(2)
        folders_css = "#paneIbfsExplorer_exTree > div.bi-tree-view-body-content table tr"
        selected_folder_css = folders_css + "[class*='selected']"
        folder_expand_css = selected_folder_css + " img[src*='collapsed']"
        folder_img_css = "img[src]:nth-child(2)"
        folders=target_table_path.split('->')
        folders[0] = 'Workspaces' if folders[0] == 'Domains' else folders[0]
        for folder in folders :
            UtillityMethods.scroll_down_and_find_item_using_mouse(self, folders_css, folder, pause_time=3)
            time.sleep(2)
            folders_obj_list = self.driver.find_elements_by_css_selector(folders_css)
            for folder_obj in folders_obj_list :
                if folder_obj.text.strip() == folder :
                    new_folders_obj_list = self.driver.find_elements_by_css_selector(folders_css)
                    elem=new_folders_obj_list[[el.text.strip() for el in new_folders_obj_list].index(folder)].find_element_by_css_selector(folder_img_css)
                    UtillityMethods.default_click(self, elem)
                    time.sleep(2)
                    UtillityMethods.synchronize_with_number_of_element(self, selected_folder_css, 1, 6)
                    time.sleep(2)
                    expand_icon_objs = self.driver.find_elements_by_css_selector(folder_expand_css)
                    if len(expand_icon_objs) != 0 :
                        expand_icon_objs[0].click()
                        time.sleep(5)
                    break
            else :
                error_msg = "Unable to find [{0}] folder in master file dialog window"
                raise FileNotFoundError(error_msg)
            
#     def select_masterfile_in_open_dialog(self, target_table_path, master_file_name, file_type=None):
#         """
#         syntax:- select_masterfile_in_open_dialog("new_retail->dimensions", "wf_retail_vendor")
#         """
#         UtillityMethods.expand_domain_folders_in_open_dialog(self, target_table_path)
#         '''
#         targets=target_table_path.split('->')
#         open_dialog_file_name="#IbfsOpenFileDialog7_cbFileName"
#         elem1=(By.CSS_SELECTOR, open_dialog_file_name)
#         self._validate_page(elem1)
#         time.sleep(4)  
#         if len(targets)>0:
#             apps_css="#paneIbfsExplorer_exTree > div.bi-tree-view-body-content table tr"
#             for app in targets[:-1]:
#                 x=[el.text.strip() for el in self.driver.find_elements_by_css_selector(apps_css)]
#                 apps=self.driver.find_elements_by_css_selector(apps_css)
#                 apps[x.index(app)].find_element_by_css_selector("img[src*='triangle']").click()
#             time.sleep(5)
#             get_folder=True
#             counter=0
#             while get_folder:
#                 x=[el.text.strip() for el in self.driver.find_elements_by_css_selector(apps_css)]
#                 if targets[-1] in x:
#                     apps=self.driver.find_elements_by_css_selector(apps_css)
#                     apps[x.index(targets[-1])].find_element_by_css_selector("img[src*='folder']").click()
#                     time.sleep(5)
#                     get_folder=False
#                     break
#                 else:
#                     counter+=1
#                     action = ActionChains(self.driver)
#                     action.send_keys(keys.Keys.PAGE_DOWN).perform()
#                     time.sleep(5)
#                     del action
#                     if counter==10:
#                         break
#                 time.sleep(1)
#             time.sleep(1)'''
#         file_name_input_css="#IbfsOpenFileDialog7_cbFileName input" 
#         self.driver.find_element_by_css_selector(file_name_input_css).click()
#         time.sleep(2)    
#         UtillityMethods.ibfs_save_as(self, master_file_name, file_type=file_type)
#         time.sleep(1)
    
    def select_masterfile_in_open_dialog(self, target_table_path, master_file_name, file_type=None):
        """
        syntax:- select_masterfile_in_open_dialog("new_retail->dimensions", "wf_retail_vendor")
        """
        UtillityMethods.synchronize_with_visble_text(self, OpenMasterFileDialog.OK_BUTTON_CSS, "Open", 120)
        UtillityMethods.expand_domain_folders_in_open_dialog(self, target_table_path)
        wait_logo_css = '{0} {1}'.format(OpenMasterFileDialog.FILE_EXPLORER_CSS, OpenMasterFileDialog.FILE_EXPLORER_WAIT_LOGO_CSS)
        UtillityMethods.synchronize_until_element_disappear(self, wait_logo_css, 45)
        elem = UtillityMethods.validate_and_get_webdriver_object(self, '#paneIbfsExplorer_exList', 'Explorer Panel')
        core_utilobj.python_move_to_element(self, elem)
        time.sleep(2) 
        master_file_list_css="{0} {1}".format(OpenMasterFileDialog.FILE_EXPLORER_CSS, OpenMasterFileDialog.FILE_EXPLORER_TABLE_ROW_OF_COLUMN_CSS)
        UtillityMethods.scroll_down_and_find_item_using_mouse(self, master_file_list_css, master_file_name, pause_time=1)
        time.sleep(2)
        required_master_file_list=UtillityMethods.validate_and_get_webdriver_objects(self, master_file_list_css, 'Exlorer panel table')
        gotit=False
        for master_file in required_master_file_list:
            if master_file_name == master_file.text:
                core_utilobj.left_click(self, master_file)
                gotit=True
                break
        if gotit == False:
            error_msg='Required [' + master_file_name + '] file is not available under [' + target_table_path + '].'
            raise FileNotFoundError(error_msg)
        ok_button_css = "{0} {1}".format(OpenMasterFileDialog.PARENT_CSS, OpenMasterFileDialog.OK_BUTTON_CSS)
        ok_button_elem=UtillityMethods.validate_and_get_webdriver_object(self, ok_button_css, 'Ok Button')
        core_utilobj.left_click(self, ok_button_elem)
        time.sleep(1)
    
    def verify_item_in_open_dialog(self, target_table_path, item_name, item_exist, msg):
        """
        syntax:- verify_item_in_open_dialog("new_retail->dimensions", "wf_retail_vendor", True, "Step X: Verify master/fex file listed")
        """
        UtillityMethods.expand_domain_folders_in_open_dialog(self, target_table_path)
        time.sleep(2)
        item_list=[el.text.strip() for el in UtillityMethods.validate_and_get_webdriver_objects(self, "#paneIbfsExplorer_exList div[class$='content']>table>tbody>tr>td:nth-child(1)", 'Exlorer panel table')]
        status=bool(item_name in item_list)
        UtillityMethods.asequal(self, item_exist, status, msg)
    
    def get_rest_test_page(self, mrid, mrpass):
        """
        Usage: active_run_fex_api_login('Active_report.fex','S7067','mrid','mrpass')
        """        
        node = UtillityMethods.parseinitfile(self, 'nodeid')
        port = UtillityMethods.parseinitfile(self, 'httpport')
        context = UtillityMethods.parseinitfile(self, 'wfcontext')
        setup_url = 'http://' + node + ':' + port + context + '/home'
        self.driver.get(setup_url)
        UtillityMethods.login_wf(self, mrid,mrpass)        
        api_url = setup_url.replace('home','') + 'wfirs?IBFS_action=TEST'
        self.driver.get(api_url)
        time.sleep(5)
    
    def verify_stack_chart(self, elem_obj1, elem_obj2, msg):
        """
        """
        temp_obj1=float(elem_obj1.get_attribute('x'))
        temp_obj2=float(elem_obj2.get_attribute('x'))
        obj_status = False
        if temp_obj1 == temp_obj2:
            obj_status = True
        UtillityMethods.asequal(self, True, obj_status, msg)
    
    def verify_cluster_chart(self, elem_obj1, elem_obj2, msg):
        """
        """
        temp_obj1=float(elem_obj1.get_attribute('x'))
        temp_obj2=float(elem_obj2.get_attribute('x'))
        obj_status = False
        if temp_obj1 < temp_obj2:
            obj_status = True
        UtillityMethods.asequal(self, True, obj_status, msg)
        
    def verify_run_time_title(self, table_id, expected_title, msg, **kwargs):
        """
        To verify run time title in Active Report.
        :param table_id='ITableData0'
        :param expected_title=['OrderNumberINTEGER', 'PackedOrderP9']
        :param msg="step 1'
        :Usage verify_run_time_title('ITableData0', ['OrderNumberINTEGER', 'PackedOrderP9'], "Step 1")
        :Usage verify_run_time_title('ITableData0', ['OrderNumberINTEGER', 'PackedOrderP9'], "Step 1", custom_css="[id^='THEAD']")
        """
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else "tr.arGridColumnHeading td[align='RIGHT']"
        parent="#" + table_id + " " + custom_css
        soup = UtillityMethods.beautifulsoup_object_creation(self)
        temp_obj = soup.select(parent)
        actual_title=(re.sub(" ",'',i.get_text().replace('\n','')) for i in temp_obj)
        for expt_label in expected_title:
            if expt_label == next(actual_title):
                stat= True
            else:
                stat=False
                break
        del actual_title
        UtillityMethods.asequal(self, stat, True, msg + " Verify the Run time Titles.")
    
    def mouse_scroll(self, scroll_type, number_of_times, option='autohotkey', pause=1):
        """
        :param : scroll_type ='up' (if you want scroll page towards to top
        :param : scroll_type ='down' (if you want scroll page towards to down)
        :param : number_of_times=3 (How many times to scroll mouse)
        :usage : mouse_scroll('up',5)
        """
        if sys.platform == 'linux':
            if scroll_type == 'up':
                mouse_.scroll(vertical=+int(number_of_times))
            elif scroll_type == 'down':
                mouse_.scroll(vertical=-int(number_of_times))
        else:
            if option == 'autohotkey':
                if scroll_type == 'up':
                    procesobj = subprocess.Popen(os.getcwd() + '\\common\\lib\\mouse_scroll.exe WheelUp '+str(number_of_times))
                    procesobj.wait()
                    del procesobj
                elif scroll_type == 'down':
                    procesobj = subprocess.Popen(os.getcwd() + '\\common\\lib\\mouse_scroll.exe WheelDown '+str(number_of_times))
                    procesobj.wait()
                    del procesobj
            elif option == 'uiautomation':
                if scroll_type == 'up':
                    automation.WheelUp(int(number_of_times), waitTime=pause)
                elif scroll_type == 'down':
                    automation.WheelDown(int(number_of_times), waitTime=pause)
                else:
                    print("Please Select Scroll 'up' or 'down' option.")
    
    def verify_checked_class_property(self, web_element, msg, check_property=True):
        '''
        Desc: This function is to verify whether an element's class property is checked.
        '''
        class_value=web_element.get_attribute("class")
        UtillityMethods.asequal(self, check_property, bool(re.match('.*checked\s*', class_value)) , msg)
    
    def verify_element_text(self,element_css,expected_text,msg):
        
        '''
        Desc: This function is to verify any given element text.
        '''
        actual_text=self.driver.find_element_by_css_selector(element_css).text.strip()
        
        UtillityMethods.asequal(self,actual_text,expected_text,msg)
        
    def verify_element_textin(self,element_css,expected_text,msg):
        
        '''
        Desc: This function is to verify any given element text.
        '''
        actual_text=self.driver.find_element_by_tag_name(element_css).text.strip()
        print(actual_text)
        UtillityMethods.asequal(self,actual_text,expected_text,msg)
    
    def select_or_verify_html_drop_down_option(self,dropdown_css,option,**kwargs):
        '''
        Desc: This function is to select or verify pure html drop down options.
        :param : dropdown_css="#IBILAYOUTDIV0TABS select[class='arDashboardMergeDropdown']"
        :param : option=United State
        :Usage : select_or_verify_html_drop_down_option("#IBILAYOUTDIV0TABS select[class='arDashboardMergeDropdown']",'North America and South America',verify_default_option='EMEA',verify_options=expected_options,msg='Step 18.1 : Verify BUSINESS_REGION_1 options')
        '''
        drop_down_items=Select(self.driver.find_element_by_css_selector(dropdown_css))
        if 'verify_default_option' in kwargs :
            default_item=drop_down_items.first_selected_option.text.strip()
            print(default_item)
            self.asequal(default_item,kwargs['verify_default_option'],'Step X : Verify '+kwargs['verify_default_option']+' is selected as default')
        if 'verify_options' in kwargs :
            actual_options=[value.text.strip() for value in drop_down_items.options]
            print(actual_options)
            self.asequal(actual_options,kwargs['verify_options'],kwargs['msg'])
        if option!=None :
            drop_down_items.select_by_visible_text(option)
            
    def save_file_from_browser(self,filename,**kwargs):
        '''
        Desc: This function is to save files from the browser . Example Excel or pptx files
        :param : filename= the file name to save to with the extension
        :kwargs: custom_ie_re = if more than one IE window the custom title name for a new IE window which contains the save dialog.
        :Usage : save_file_from_browser('one.xlsx', custom_ie_re='.*BIP_RUN.*')
        '''
        custom_ie_re=kwargs['custom_ie_re'] if 'custom_ie_re' in kwargs else ".*Internet Explore.*"
        custom_cr_re=kwargs['custom_cr_re'] if 'custom_cr_re' in kwargs else ".*Chrome"
        br = Global_variables.browser_name
        if br=='chrome':
            time.sleep(5)
            automation.WindowControl(RegexName="Save As.*").SetFocus()
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").Click()
            time.sleep(2)
            pyautogui.hotkey('ctrl','a')
            time.sleep(2)
            pyautogui.hotkey('del')
            time.sleep(2)
#             automation.EditControl(RegexName="File name.*").SendKeys(os.getcwd() + "\data\\" + filename)
            path = os.getcwd() + "\data\\" + filename
            print(path)
            pyautogui.typewrite(path)
            automation.ButtonControl(Name="Save").SetFocus()
            automation.ButtonControl(Name="Save").Click(10, 10, False, 10)
            time.sleep(2)
            if automation.WindowControl(RegexName="Confirm Save.*").Exists():
                automation.WindowControl(RegexName="Confirm Save.*").SetFocus()
                automation.ButtonControl(Name="Yes").Click(10, 10, False, 10)
            time.sleep(6)   
        elif br=='firefox':
            automation.WindowControl(RegexName="Opening.*").Exists(10, 1)
            automation.WindowControl(RegexName="Opening.*").SetFocus()
            time.sleep(2)
            automation.RadioButtonControl(Name="Save File").Click(10, 0, False, 0)
            time.sleep(2)
            automation.ButtonControl(Name="OK").SetFocus()
            automation.ButtonControl(Name="OK").Click(10, 10, False, 10)
            #automation.ButtonControl(Name="OK").Click()
            time.sleep(5)
            automation.WindowControl(RegexName="Enter name of file.*").SetFocus()
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").Click()
            time.sleep(2)
            pyautogui.hotkey('ctrl','a')
            time.sleep(2)
            pyautogui.hotkey('del')
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").SendKeys(os.getcwd() + "\data\\" + filename)
            automation.ButtonControl(Name="Save").SetFocus()
            automation.ButtonControl(Name="Save").Click(10, 10, False, 10)
            time.sleep(2)
            if automation.WindowControl(RegexName="Confirm Save.*").Exists():
                automation.WindowControl(RegexName="Confirm Save.*").SetFocus()
                automation.ButtonControl(Name="Yes").Click(10, 10, False, 10)
            time.sleep(6) 
        elif br=="edge":
            win_control=automation.WindowControl(ClassName="ApplicationFrameWindow").GroupControl(Name="Save")
            automation.WaitForExist(win_control, timeout=90)
            win_control.SetFocus()
            win_control.ButtonControl(Name="Split button, more options").Click()
            automation.WindowControl(ClassName="ApplicationFrameWindow").ButtonControl(Name="Save as").Click()
            save_as_win=automation.WindowControl(Name="Save As")
            automation.WaitForExist(save_as_win, timeout=30)
            save_as_win.SetFocus()
            save_as_win.EditControl(Name="File name:").SetFocus()
            save_as_win.EditControl(Name="File name:").Click()
            time.sleep(2)
#             pyautogui.hotkey('ctrl','a')
#             time.sleep(2)
#             pyautogui.hotkey('del')
#             time.sleep(2)
            save_as_win.EditControl(Name="File name:").SendKeys(os.getcwd() + "\data\\" + filename)
            time.sleep(2)
            save_as_win.ButtonControl(Name="Save").Click()
            time.sleep(2)
            if automation.WindowControl(RegexName="Confirm Save.*").Exists():
                automation.WindowControl(RegexName="Confirm Save.*").SetFocus()
                automation.ButtonControl(Name="Yes").Click(10, 10, False, 10)
            time.sleep(6)   
            d_button=automation.WindowControl(ClassName="ApplicationFrameWindow").ButtonControl(Name="Dismiss")
            automation.WaitForExist(d_button, timeout=60)
            d_button.MoveCursorToMyCenter()
            d_button.Click()
            automation.WindowControl(ClassName="ApplicationFrameWindow").SetFocus()
        else:
                app = pywinauto.application.Application().connect(class_name="IEFrame", title_re=custom_ie_re)
                ieframe = app.IEFrame
                ieframe.Wait('ready')
                time.sleep(8)
                if bool(ieframe[u'Frame Notification Bar'].Exists()):
                    
                    framenotificationbar = ieframe[u'Frame Notification Bar']
                    framenotificationbar.Maximize()
                    framenotificationbar.Restore()
                    time.sleep(3)
                    pyautogui.hotkey('alt','n')
                    time.sleep(2)
                    pyautogui.press('up')
                    time.sleep(2)
                    pyautogui.press('a')
                    time.sleep(2)
#                     pyautogui.press('down')
#                     time.sleep(2)
#                     pyautogui.press('enter')
#                     time.sleep(2)
                    dlg=app.window(title = 'Save As')
                    
                    dlg[u'Edit'].ClickInput()
                    time.sleep(2)
                    pyautogui.hotkey('ctrl','a')
                    time.sleep(2)
                    pyautogui.hotkey('del')
                    time.sleep(2)
                    #dlg[u'Edit'].set_text(os.getcwd() + "\data\\" + filename)
                    pyautogui.typewrite(os.getcwd() + "\data\\" + filename, interval=0.3, pause=3)
                    time.sleep(2)
                    dlg['Button1'].ClickInput()
                    time.sleep(2)
                       
                    if app.window(title='Confirm Save As').Exists()==True:
                        window1 = app.window(title='Confirm Save As')
                        window1.Maximize()
                        window1.Restore()
                        button = window1[u'Yes']
                        button.ClickInput()
                    
                    time.sleep(6) 
    
    
    def compare_using_openpyxl(self, base_file, base_file_sheet_name, actual_file, actual_file_sheet_name, rows, columns):
        '''
        Desc:- This function is used to compare two xlsx files using openpyxl. Keep your base file under data folder. Also your run time 
        xlsx under data folder.
        '''
        workbook1 = load_workbook(os.getcwd() + "\\data\\" + base_file)
        sheet1 = workbook1.get_sheet_by_name(base_file_sheet_name)
        workbook2 = load_workbook(os.getcwd() + "\\data\\" + actual_file)
        sheet2 = workbook2.get_sheet_by_name(actual_file_sheet_name)
        status={'status':'pass'}
        for r in range(1, rows+1):
            
            for c in range(1, columns+1):
                
                if sheet1.cell(row=r, column=c).value != sheet2.cell(row=r, column=c).value:
                    status['status']= 'fail at row number [' + str(r) + '] and column number [' + str(c) + '].'
                    return(status)
        return(status)
    
    def verify_excel_files(self, base_file, base_file_sheet_name, actual_file, actual_file_sheet_name, rows, columns, msg, tool='openpyxl'):
        '''
        Desc:- 
        toor=openpyxl or xlrd
        '''
        if tool=='openpyxl':
            status=UtillityMethods.compare_using_openpyxl(self, base_file, base_file_sheet_name, actual_file, actual_file_sheet_name, rows, columns)
            print ('main status', status)
        if tool=='xlrd':
            status=UtillityMethods.compare_using_xlrd(self, base_file, base_file_sheet_name, actual_file, actual_file_sheet_name, rows, columns)
        a=status['status']
        print (str(a))
        UtillityMethods.asequal(self,'pass', str(a), msg)  
        
    def compare_using_xlrd(self, base_file, base_file_sheet_name, actual_file, actual_file_sheet_name, rows, columns):
        import xlrd
        workbook1 = xlrd.open_workbook(os.getcwd() + "\data\\" + base_file)
        sheet1 = workbook1.sheet_by_name(base_file_sheet_name)
        workbook2 = xlrd.open_workbook(os.getcwd() + "\data\\" + actual_file)
        sheet2 = workbook2.sheet_by_name(actual_file_sheet_name)
        status={'status':'pass'}
        for r in range(0, rows):
            for c in range(0, columns):
                
                if sheet1.cell(r, c).value != sheet2.cell(r, c).value:
                    status['status']='fail at row number [' + str(r+1) + '] and column number [' + str(c+1) + '].'
                    return(status)
        return (status)
    
    
    def wait_for_object_not_visible(self, css, timeout, msg):
        """
        :params css="div[id^='ibi'][class=tdgchart-tooltip] .tdgchart-tooltip-pad"
        :params timeout=30
        :params msg='Step 5: Verify tooltip menu is not visible'
        :Usage wait_for_object_not_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", 30, 'Step 5: Filter Button Visible')
                OR
        :Usage wait_for_object_not_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", 15, 'Step 5: Filter Button Removed')
        """
        not_visible=True
        try:
            wait = WebDriverWait(self.driver, timeout)
            status = bool(wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR,css))));
        except:
            status = False
        UtillityMethods.asequal(self, status, not_visible, msg)
        

    def get_actual_tooltip_list(self, raw_tooltip_list):
        '''
        Desc:- This function will return the actual tooltip in a list format, where input will be a raw list.
                This will remove all unwanted items like '>', blank space, empty items, etc.
        '''
        actual_list=[]      
        for raw_tooltip_item in raw_tooltip_list:
            if bool(re.match(r'.*:\s.*', raw_tooltip_item)):
                reqobj = re.match('(.*):\s{1,}(.*)', raw_tooltip_item)
                new_element = str(reqobj.group(1)) + ":" + str(reqobj.group(2))
            elif bool(re.match(r'^>', raw_tooltip_item)):
                new_element = re.sub('>', '', raw_tooltip_item)
            elif bool(raw_tooltip_item==''):
                continue
            else:
                new_element=raw_tooltip_item
            actual_list.append(new_element)
        actual_list=[list_value for list_value in actual_list if list_value not in '']
        return(actual_list)
    
    def save_as_excel_sheet (self, filename):
        '''
        Use this to save the current open Excel workbook as a new filename in data folder
        '''
        
        try:
            
            if bool(Application().Connect(class_name='XLMAIN')):
                pyautogui.hotkey('alt','f')
                time.sleep(4)
                pyautogui.hotkey('a')
                filebool=True
                time.sleep(10)
                app = Application().Connect(class_name='XLMAIN')
                mainapp=app.XLMAIN
                mainapp.Maximize()
                window = app.Dialog
                window.SetFocus()
                window.Maximize()
                time.sleep(10)
                edit = window[u'Edit']
                edit.ClickInput()
                pyautogui.hotkey('ctrl','a')
                time.sleep(2)
                pyautogui.hotkey('del')
                time.sleep(2)
                pyautogui.typewrite(os.getcwd() + "\data\\" + filename, interval=0.3, pause=3)
                time.sleep(2)
                button = window[u'&Save']
                button.ClickInput()
                app.Kill_()
        except pywinauto.findwindows.ElementNotFoundError:
            filebool=False
        UtillityMethods.asequal(self,filebool,True,'Step X. Excel File available to save')
        
    
    def save_as_powerpoint_document (self, filename):
        '''
        Use this to save the current open Powerpoint doc as a new filename in data folder
        '''
        
        try:
            
            if bool(Application().Connect(class_name='PPTFrameClass')):
                pyautogui.hotkey('alt','f')
                time.sleep(4)
                pyautogui.hotkey('a')
                filebool=True
                time.sleep(10)
                app = Application().Connect(class_name='PPTFrameClass')
                mainapp=app.PPTFrameClass
                mainapp.Maximize()
                window = app.Dialog
                window.SetFocus()
                window.Maximize()
                time.sleep(10)
                edit = window[u'Edit']
                edit.ClickInput()
                pyautogui.hotkey('ctrl','a')
                time.sleep(2)
                pyautogui.hotkey('del')
                time.sleep(2)
                pyautogui.typewrite(os.getcwd() + "\data\\" + filename, interval=0.3, pause=3)
                time.sleep(2)
                button = window[u'&Save']
                button.ClickInput()
                app.Kill_()
        except pywinauto.findwindows.ElementNotFoundError:
            filebool=False
        UtillityMethods.asequal(self,filebool,True,'Step X. Powerpoint File available to save')
        
    def save_as_word_document (self, filename):
        '''
        Use this to save the current open word doc as a new filename in data folder
        '''
        
        try:
            
            if bool(Application().Connect(class_name='OpusApp')):
                pyautogui.hotkey('alt','f')
                time.sleep(4)
                pyautogui.hotkey('a')
                filebool=True
                time.sleep(10)
                app = Application().Connect(class_name='OpusApp')
                mainapp=app.OpusApp
                mainapp.Maximize()
                window = app.Dialog
                window.SetFocus()
                window.Maximize()
                time.sleep(10)
                edit = window[u'Edit']
                edit.ClickInput()
                pyautogui.hotkey('ctrl','a')
                time.sleep(2)
                pyautogui.hotkey('del')
                time.sleep(2)
                pyautogui.typewrite(os.getcwd() + "\data\\" + filename, interval=0.3, pause=3)
                time.sleep(2)
                button = window[u'&Save']
                button.ClickInput()
                app.Kill_()
        except pywinauto.findwindows.ElementNotFoundError:
            filebool=False
        UtillityMethods.asequal(self,filebool,True,'Step X. Word document File available to save')
        
    def zoom_powerpoint_document(self, zoomlevel):
        try:
            
            if bool(Application().Connect(class_name='PPTFrameClass')):
                
                app = Application().Connect(class_name='PPTFrameClass')
                pptframeclass = app.PPTFrameClass
                pptframeclass.SetFocus()
                pptframeclass.Maximize()
                filebool=True
                time.sleep(10)
                pyautogui.hotkey('alt')
                time.sleep(1)
                pyautogui.hotkey('w')
                time.sleep(1)
                pyautogui.hotkey('q')
                time.sleep(3)
                window = app.Zoom
                if app.zoom.Exists() ==True:
                    pyautogui.hotkey('p')
                    time.sleep(2)
                    pyautogui.hotkey('ctrl','a')
                    time.sleep(2)
                    pyautogui.hotkey('del')
                    time.sleep(2)
                    pyautogui.typewrite(str(zoomlevel), interval=0.3, pause=3)
                    button = window.OK
                    #button.ClickInput()
                    pyautogui.hotkey('enter')
                    time.sleep(6)
        except pywinauto.findwindows.ElementNotFoundError:
            filebool=False
        UtillityMethods.asequal(self,filebool,True,'Step X. Powerpoint file available to zoom')
        
    def zoom_excel_document(self, zoomlevel):
        try:
            
            if bool(Application().Connect(class_name='XLMAIN')):
                
                app = Application().Connect(class_name='XLMAIN')
                pptframeclass = app.XLMAIN
                pptframeclass.SetFocus()
                pptframeclass.Maximize()
                filebool=True
                time.sleep(10)
                pyautogui.hotkey('alt')
                time.sleep(1)
                pyautogui.hotkey('w')
                time.sleep(1)
                pyautogui.hotkey('q')
                time.sleep(3)
                window = app.Zoom
                if app.zoom.Exists() ==True:
                    pyautogui.hotkey('c')
                    time.sleep(2)
                    pyautogui.hotkey('ctrl','a')
                    time.sleep(2)
                    pyautogui.hotkey('del')
                    time.sleep(2)
                    pyautogui.typewrite(str(zoomlevel), interval=0.3, pause=3)
                    button = window.OK
                    #button.ClickInput()
                    pyautogui.hotkey('enter')
                    time.sleep(6)
        except pywinauto.findwindows.ElementNotFoundError:
            filebool=False
        UtillityMethods.asequal(self,filebool,True,'Step X. Excel file available to zoom')
        
    def zoom_word_document(self, zoomlevel):
        try:
            
            if bool(Application().Connect(class_name='OpusApp')):
                
                app = Application().Connect(class_name='OpusApp')
                pptframeclass = app.OpusApp
                pptframeclass.SetFocus()
                pptframeclass.Maximize()
                filebool=True
                time.sleep(10)
                pyautogui.click(1280, 650, button='left')
                time.sleep(10)
                pyautogui.hotkey('alt')
                time.sleep(1)
                pyautogui.hotkey('w')
                time.sleep(1)
                pyautogui.hotkey('q')
                time.sleep(3)
                window = app.Zoom
                if app.zoom.Exists() ==True:
                    pyautogui.hotkey('e')
                    time.sleep(2)
                    pyautogui.hotkey('ctrl','a')
                    time.sleep(2)
                    pyautogui.hotkey('del')
                    time.sleep(2)
                    pyautogui.typewrite(str(zoomlevel), interval=0.3, pause=3)
                    button = window.OK
                    #button.ClickInput()
                    pyautogui.hotkey('enter')
                    time.sleep(10)
        except pywinauto.findwindows.ElementNotFoundError:
            filebool=False
        UtillityMethods.asequal(self,filebool,True,'Step X.Word file available to zoom')
        
    def ie_security_dialog_close(self):
        '''
        use this action to allow and close the IE security dialog 
        '''
        
        try:
            app = Application().Connect(class_name='IEFrame')
            if bool(app[u'Internet Explorer Security'].Exists()):
                window=app[u'Internet Explorer Security']
                button = window[u'&Allow']
                button.Click()
        except pywinauto.findwindows.ElementNotFoundError:
            pass
    
        
    def verify_and_save_outlook_attachment_type(self, windowtitle, attachment_name,savename,**kwargs):
        '''
        windowtitle= The window title of the outlook msg window. Ex. "Active Report - Message (HTML) "
        attachment_name=The attachment name that you want to verify exists in the attachment. Ex. ActiveReport1.htm
        savename=The filename that you want to save the attachment as from the outlook widnow. Ex. newsave.htm
        '''
        
        try:
            if bool(Application().Connect(class_name='rctrl_renwnd32')):
                app = Application().Connect(class_name='rctrl_renwnd32')
                outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
                app[windowtitle].Exists()
                window=app[windowtitle]
                window.SetFocus()
                window.Maximize()
                
                inbox=outlook.GetDefaultFolder(6)
                messages=inbox.Items
                message = messages.GetFirst()
                for attachment in message.Attachments:
                    if attachment.DisplayName == attachment_name:
                        status=True
                    else:
                        status=False
                    UtillityMethods.asequal(self,True,status,'Step X. Verify attachment name')
                    
                    attachment.SaveAsFile(os.path.join(os.getcwd() + "\data\\" + savename))
                filebool=True
                if 'close' in kwargs:
                    app.Kill_()
#                   
        except pywinauto.findwindows.ElementNotFoundError:
            filebool=False
            if bool(Application().Connect(class_name='rctrl_renwnd32')):
                app = Application().Connect(class_name='rctrl_renwnd32')
                app.Kill_()
        UtillityMethods.asequal(self,filebool,True,'Step X. Attachment saved and name Verified')
        
    def select_dropdown(self, select_css, select_type, value):    
        '''
        select_css: #combobox_dsPROMPT_1
        type: index or value or visible_text 
        value: 'Jan, 1987
        Usage: select_dropdown('#combobox_dsPROMPT_1', 'visible_text', 'Jan, 1987')
        '''
        select=Select(self.driver.find_element_by_css_selector(select_css))
        if select_type=='index':
            select.select_by_index(value)
        elif select_type=='value':
            select.select_by_value(value)
        elif select_type=='visible_text':
            select.select_by_visible_text(value)
        
    def verify_dropdown_value(self, select_css, value_list=None, msg=None, expected_default_selected_value=None, default_selection_msg=None): 
        '''
        select_css: #combobox_dsPROMPT_1
        value_list: ['[All]','ITALY,COUPE','ITALY,ROADSTER']
        msg: "Step02: Verify list"
        expected_default_selected_value: "ITALY,COUPE"
        default_selection_msg: "Step02: Verify selected value"
        Usage: verify_dropdown_value("#combobox_dsPROMPT_1",expected_default_selected_value="ITALY,COUPE", default_selection_msg="Step1.2: Verify")
        '''
        if expected_default_selected_value != None:
            select_obj=Select(self.driver.find_element_by_css_selector(select_css))
            actual_default_selected_value = select_obj.first_selected_option.text
            UtillityMethods.asequal(self, expected_default_selected_value, actual_default_selected_value, default_selection_msg)
        if value_list != None:
            select_values_obj=self.driver.find_elements_by_css_selector(select_css+ "  option" )
            total_list = (el.text for el in select_values_obj)
            verify_status=True
            for value in value_list:
                if value == next(total_list):
                    verify_status= True
                else:
                    verify_status=False
                    break
            del total_list
            UtillityMethods.asequal(self, verify_status, True, msg)
            
    def compare_two_file_contents(self, base_file, actual_file, msg):
        '''
        use this action to compare any two file contents saved onto the data folder.
        Usage: compare_two_file_contents('one.csv','two.csv','Step x. Verify the csv contents')
        '''
        
        with open(os.getcwd()+"\data\\"+base_file) as f1, open(os.getcwd()+"\data\\"+actual_file) as f2:
            for l1, l2 in zip(f1, f2):
                if l1 == l2:
                    status_bool=True
                else:
                    status_bool=False
        UtillityMethods.asequal(self, status_bool, True, msg)
        
    def drag_slider_using_pageup_or_pagedown(self, parent_css, slider_button_css="[class^='ui-slider-handle']", pageup_or_pagedown='pageup', number_of_time=1, span_index=0, pause_time=3):
        '''
        This function is used to move the slider span button using page up or page down keys.
        Sometimes we face one slider button or two slider button. So accordingly user has to send span_index as 0 or 1.
        '''
        counter=0
        slider_point_css = parent_css + " " + slider_button_css
        start_point=self.driver.find_elements_by_css_selector(slider_point_css)[span_index]
        pyautogui.moveTo(5, 5, 4)
        UtillityMethods.click_on_screen(self,start_point,'middle',click_type=0)
        if Global_variables.browser_name == 'firefox':
            time.sleep(8)
        else:
            time.sleep(3)
        while counter < number_of_time:
            if pageup_or_pagedown == 'pageup':
                start_point.send_keys(keys.Keys.PAGE_UP)
            if pageup_or_pagedown == 'pagedown':
                start_point.send_keys(keys.Keys.PAGE_DOWN)
            if pageup_or_pagedown == 'down':
                start_point.send_keys(keys.Keys.DOWN)
            if pageup_or_pagedown == 'up':
                start_point.send_keys(keys.Keys.UP) 
            time.sleep(pause_time) 
            counter = counter+1
    
    def get_configparser_object(self, file_path):
        """
        This method used to read given file path using ConfigParser package and return object 
        """
        Error_Msg="[{0}] path not exists in file system".format(file_path)
        if os.path.exists(file_path) :
            parser=ConfigParser()
            parser.optionxform = str
            parser.read(file_path)
            return parser
        else :
            raise FileNotFoundError(Error_Msg)
    
    def read_configparser_key_value(self, file_path, section, parser_key):
        """
        This method used to read specific section key value
        """
        SECTION_MISSING_ERROR=" [{0}] section is not listed in [{1}] file ".format(section, file_path)
        SECTION_KEY_MISSING_ERROR=" [{0}] key is not listed under the [{1}] section in [{2}] file".format(parser_key, section, file_path)
        parser_obj=UtillityMethods.get_configparser_object(self, file_path)
        if  parser_obj.has_section(section) :
            if parser_obj.has_option(section, parser_key):
                return parser_obj[section][parser_key]
            else :
                raise KeyError(SECTION_KEY_MISSING_ERROR)  
        else :
            raise KeyError(SECTION_MISSING_ERROR)
    
    def scroll_and_select_combobox_item(self, combobox_btn_css, combobox_item, **kwargs):
        """
        This method used to scroll and click on combobox item.
        Example Usage : scroll_and_select_combobox_item(self, "#dlgSecurityUserEdit #SecurityUserEditDialog_cbDefGroup div[id^='BiButton']", 'P116')
        """
        Error_Msg="[{0}] item not listed in comboxbox".format(combobox_item)
        combobox_items_css="div[id^='BiPopup'][style*='inherit'] div[id^='BiComboBoxItem']"
        combobox_scrollable_css="div[id^='BiPopup'][style*='inherit']>div[class*='bi-combo-box-list']"
        combo_btn=self.driver.find_element_by_css_selector(combobox_btn_css)
        UtillityMethods.default_click(self, combo_btn, **kwargs)
        UtillityMethods.synchronize_with_number_of_element(self, "div[id^='BiPopup'][style*='inherit']", 1, expire_time=4)
        combobox_items_obj=self.driver.find_elements_by_css_selector(combobox_items_css)
        combobox_item_obj=JavaScript.find_elements_by_text(self, combobox_items_obj, combobox_item)
        if len(combobox_item_obj) == 0 :
            raise NoSuchElementException(Error_Msg)
        scroll_offset=JavaScript.get_scroll_offsetTop(self, combobox_item_obj[0])
        JavaScript.scroll_element(self, combobox_scrollable_css, scroll_offset-20, wait_time=1)
        combobox_item_obj[0].click()
    
    def scroll_down_and_find_item_using_page_down(self, item_list_css, item_name_to_find, pause_time=1):
        """
        This method used to scroll down and find item using keybord page down key
        Example usage : scroll_down_and_find_item_using_page_down("#SecurityManagerDialog_treeUsers table[class='bi-tree-view-table']>tbody>tr>td:nth-child(1), 'Admin')
        """
        UtillityMethods.__scroll_down_and_find_item(self, item_list_css, item_name_to_find, scroll_type='page_down', pause_time=pause_time)
    
    def scroll_down_and_find_item_using_mouse(self, item_list_css, item_name_to_find, pause_time=1, number_of_times=1):
        """
        This method used to scroll down and find item using mouse down wheel.
        Example usage : scroll_down_and_find_item_using_mouse("#SecurityManagerDialog_treeUsers table[class='bi-tree-view-table']>tbody>tr>td:nth-child(1), 'Admin')
        """
        UtillityMethods.__scroll_down_and_find_item(self, item_list_css, item_name_to_find, scroll_type='mouse', pause_time=pause_time, number_of_times=number_of_times)
    
    def __scroll_down_and_find_item(self, item_list_css, item_name_to_find, scroll_type, pause_time=1, number_of_times=1):
        """
        This method used to scroll element and find any specific item. It will bring given item_name_to_find element to top visible area if item name found.
        example usage : __scroll_down_and_find_item(self, item_list_css, item_name_to_find, scroll_type='mouse', pause_time=pause_time)
        """
        previouse_item_names_list=[]
        while True :
            item_list_object=self.driver.find_elements_by_css_selector(item_list_css)
            current_item_names_list=JavaScript.get_elements_text(self, item_list_object)
            if item_name_to_find not in current_item_names_list and previouse_item_names_list != current_item_names_list :
                if scroll_type == 'page_down' :
                    keyboard.send('page down')
                if scroll_type == 'mouse' :
                    UtillityMethods.mouse_scroll(self, 'down', number_of_times, option='uiautomation',  pause=0)
            else :
                if item_name_to_find==current_item_names_list[-1] :
                    UtillityMethods.mouse_scroll(self, 'down', number_of_times, option='uiautomation',  pause=0)
                break
            time.sleep(pause_time) 
            previouse_item_names_list=current_item_names_list.copy()
            
    def verify_xml_xls(self, base_file_name, actual_file_name, msg):
        base_file_name = os.getcwd() +"\\data\\" + base_file_name
        actual_file_name = os.getcwd() +"\\data\\" + actual_file_name
        base_file_name = base_file_name
        print(base_file_name)
        path = os.path.join(root_path.ROOT_PATH, 'verify_xml_xls.vbs '+ actual_file_name + " " + base_file_name)
        print(path)
        verify_xml_xls_cmd="cscript.exe " +path
        print(verify_xml_xls_cmd)
#         try:
#             if bool(pywinauto.findwindows.find_window(title_re='Windows Script Host')):
#                 print ("VBA error message")
#                 app = pywinauto.application.Application().Connect(title_re='Windows Script Host', class_name='#32770')
#                 window = app.Dialog
#                 button = window.OK
#                 button.Click()
#         except:
#             pass
        resp=subprocess.call(verify_xml_xls_cmd, shell=True)
        UtillityMethods.asequal(self, 9, resp, msg)
        
    def get_masterfile_field_path(self, master_file_name, filed_path_key):
        """
        This method used to get data field path from master_file.data
        """
        MASTER_FILE_NAME='master_file.data'
        MASTER_FILE_PATH=os.path.join(root_path.ROOT_PATH, MASTER_FILE_NAME)
        field_path=UtillityMethods.read_configparser_key_value(self, MASTER_FILE_PATH, master_file_name, filed_path_key)
        return field_path
    
    def left_click_with_offset(self, element_object, x_offset=0, y_offset=0, **kwargs):
        """
        This method used to left click on element with offset
        Example Usage : left_click_with_offset(element_object, 10,20)
        """
        if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
            UtillityMethods.click_on_screen(self, element_object, coordinate_type='middle',click_type=0, x_offset=x_offset, y_offset=y_offset, **kwargs)
        else:
            ActionChains(self.driver).move_to_element_with_offset(element_object, x_offset, y_offset).click().perform()
    
    def right_click_with_offset(self, element_object, x_offset=0, y_offset=0, **kwargs):
        """
        This method used to right click on element with offset
        Example Usage : left_click_with_offset(element_object, 10,20)
        """
        if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
            UtillityMethods.click_on_screen(self, element_object, coordinate_type='middle',click_type=1, x_offset=x_offset, y_offset=y_offset, **kwargs)
        else:
            ActionChains(self.driver).move_to_element_with_offset(element_object, x_offset, y_offset).context_click().perform()
    
    def double_click_with_offset(self, element_object, x_offset=0, y_offset=0, **kwargs):
        """
        This method used to double click on element with offset
        Example Usage : left_click_with_offset(element_object, 10,20)
        """
        if Global_variables.browser_name in ['ie', 'firefox', 'edge']:
            UtillityMethods.click_on_screen(self, element_object, coordinate_type='middle',click_type=2, x_offset=x_offset, y_offset=y_offset, **kwargs)
        else:
            ActionChains(self.driver).move_to_element_with_offset(element_object, x_offset, y_offset).double_click().perform()
            
    def save_as_file_from_browser(self,filename,**kwargs):
        '''
        Desc: This function is to save files from the browser . Example Excel or pptx files
        :param : filename= the file name to save to with the extension
        :kwargs: custom_ie_re = if more than one IE window the custom title name for a new IE window which contains the save dialog.
        :Usage : save_file_from_browser('one.xlsx', custom_ie_re='.*BIP_RUN.*')
        '''
        
        custom_ie_re=kwargs['custom_ie_re'] if 'custom_ie_re' in kwargs else ".*Internet Explore.*"
        custom_cr_re=kwargs['custom_cr_re'] if 'custom_cr_re' in kwargs else ".*Chrome"
        
        br = Global_variables.browser_name
        if os.path.isfile(os.getcwd()+"\data\\"+filename+".htm"):
            os.remove(os.getcwd()+"\data\\"+filename+".htm")
        if os.path.isfile(os.getcwd()+"\data\\"+filename+".html"):
            os.remove(os.getcwd()+"\data\\"+filename+".html")
        
        if br=='firefox':
            automation.WindowControl(RegexName="Save As.*").SetFocus()
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").Click()
            time.sleep(2)
            pyautogui.hotkey('ctrl','a')
            time.sleep(2)
            pyautogui.hotkey('del')
            time.sleep(2)
            #automation.EditControl(RegexName="File name.*").SendKeys(keys, interval, waitTime)
            automation.EditControl(RegexName="File name.*").SetValue(os.getcwd() + "\data\\" + filename)
            automation.ButtonControl(Name="Save").SetFocus()
            automation.ButtonControl(Name="Save").Click(10, 10, False, 10)
            time.sleep(2)
            if automation.WindowControl(RegexName="Confirm Save.*").Exists():
                automation.WindowControl(RegexName="Confirm Save.*").SetFocus()
                automation.ButtonControl(Name="Yes").Click(10, 10, False, 10)
                 
        else:
            automation.WindowControl(RegexName="Save Webpage.*").SetFocus()
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").Click()
            time.sleep(2)
            pyautogui.hotkey('ctrl','a')
            time.sleep(2)
            pyautogui.hotkey('del')
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").SetValue(os.getcwd() + "\data\\" + filename)
            automation.ButtonControl(Name="Save").SetFocus()
            automation.ButtonControl(Name="Save").Click(10, 10, False, 10)
            time.sleep(2)
            if automation.WindowControl(RegexName="Confirm Save.*").Exists():
                automation.WindowControl(RegexName="Confirm Save.*").SetFocus()
                automation.ButtonControl(Name="Yes").Click(10, 10, False, 10)
    
    def verify_visible_elements(self, elements_obj, expected_total_visible_elements, msg):
        """
        Descriptions : This method used to get visible element count
        """
        actaul_visible_elements=0
        for element in elements_obj :
            if element.is_displayed() == True :
                actaul_visible_elements+=1
        UtillityMethods.asequal(self, expected_total_visible_elements, actaul_visible_elements, msg)
    
    def get_element_border_property_value(self, element_obj, border_property_name):
        """
        Descriptions : This method used to get element border property value such as border-color, border-width, border-style.
        It will return given border property value if all border directions(left, top, right, bottom) have same value else None
        border_property_name = 'color, 'width', 'style'
        """
        top=element_obj.value_of_css_property('border-top-' + border_property_name)
        right=element_obj.value_of_css_property('border-right-' + border_property_name)
        bottom=element_obj.value_of_css_property('border-bottom-' + border_property_name)
        left=element_obj.value_of_css_property('border-left-' + border_property_name)
        actual_value=None
        if top == right == bottom == left :
            actual_value = top
            return actual_value
    
    def scroll_down_on_element(self, scrollable_element_obj, number_of_scroll=None):
        """
        Descriptions : This method used get to scroll down on scrollable element
        :arg - scrollable_element_obj = object of scrollable element.
        :arg - number_of_scroll = we can pass number to how many times to scroll down. 
        It will keep on scroll down until scroll bar reach bottom of the given scrollable_element if number_of_scroll==None
        example usage1 : scroll_down_on_element()
        example usage2 : scroll_down_on_element(3)
        """
        scroll_count = 0
        max_scroll_count = 250
        core_utilobj.python_move_to_element(self, scrollable_element_obj, element_location='top_middle', yoffset=10)
        time.sleep(2)
        while True :
            is_scroll_reached_bottom = JavaScript.check_vertical_scroll_reached_bottom(self, scrollable_element_obj)
            if scroll_count == number_of_scroll or max_scroll_count == scroll_count or is_scroll_reached_bottom == True :
                break
            else :
                UtillityMethods.mouse_scroll(self, 'down', 1, option='uiautomation',  pause=0)
                scroll_count+= 1

    def verify_element_has_vertical_scrollbar(self, scroll_element_ob, msg):
        """
        Descriptions : This method used to verify whether specific element object has vertical scroll bar
        :arg - scroll_element_ob = object of scrollable element.   
        """
        scroll_status = JavaScript.check_element_has_vertical_scrollbar(self, scroll_element_ob)
        UtillityMethods.asequal(self, True, scroll_status, msg)
    
    def verify_element_has_horizontal_scrollbar(self, scroll_element_ob,  msg):
        """
        Descriptions : This method used to verify whether specific element object has horizontal scroll bar
        :arg - scroll_element_ob = object of scrollable element.   
        """
        scroll_status = JavaScript.check_element_has_horizontal_scrollbar(self, scroll_element_ob)
        UtillityMethods.asequal(self, True, scroll_status, msg)
    
    def verify_element_does_not_have_vertical_scrollbar(self, scroll_element_ob,  msg):
        """
        Descriptions : This method used to verify whether specific element object does not have vertical scroll bar
        :arg - scroll_element_ob = object of scrollable element.   
        """
        scroll_status = JavaScript.check_element_has_vertical_scrollbar(self, scroll_element_ob)
        UtillityMethods.asequal(self, False, scroll_status, msg)
    
    def verify_element_does_not_have_horizontal_scrollbar(self, scroll_element_ob,  msg):
        """
        Descriptions : This method used to verify whether specific element object does not have horizontal scroll bar
        :arg - scroll_element_ob = object of scrollable element.   
        """
        scroll_status = JavaScript.check_element_has_horizontal_scrollbar(self, scroll_element_ob)
        UtillityMethods.asequal(self, False, scroll_status, msg)
    
    def verify_vertical_scrollbar_reached_bottom(self, scroll_element_ob, msg):
        """
        Descriptions : This method used to verify vertical scroll bar reached bottom of the given element object
        :arg - scroll_element_ob = object of scrollable element.   
        """
        scroll_status = JavaScript.check_vertical_scroll_reached_bottom(self, scroll_element_ob)
        UtillityMethods.asequal(self, True, scroll_status, msg)
        
    def verify_tab_item_using_uiautomation(self, new_tab_name):
        """
        Descriptions : This method used to verify link opened as a tab item
        :arg - new_tab_name = .*Technical Content.*   
        """
        expected_control_type_search=automation.TabItemControl(RegexName=new_tab_name).LocalizedControlType      
        UtillityMethods.asequal(self, expected_control_type_search, 'tab item', "StepX: Verify tab item appears")
        
    def set_browser_window_size(self, x='998', y='768'):
        '''
        Description:-This function used to set the browser size to the given width and height
        report_obj.set_browser_window_size()
        '''
        self.driver.set_window_size(x, y)
    
    def click_on_edge_browser_popup_button_using_uiautomation(self, button_name='Leave'):
        """
        Description : This method will click any types button based on button name.
        """
        try :
            button = automation.ButtonControl(Name=button_name)
            button.Click()
        except :
            pass
    
    def reset_edge_browser_zoom(self):
        """
        Description : This method will rest the default(100%) zoom size
        """
        try :
            ActionChains(self.driver).send_keys(keys.Keys.CONTROL, 0).perform()
            ActionChains(self.driver).send_keys(keys.Keys.CONTROL, 0).perform()
        except :
            print("Something went wrong")
        
    def get_console_log(self, log_type='browser'):
        '''
        Description : This method will read console log and return as a list.
        log_type : ['browser', 'driver']
        '''
        log_value = self.driver.get_log(log_type)
        log_value_list = []
        for value_ in log_value:
            for obj_ in value_.keys():
                log_value_list.append(obj_)
            for obj1_ in value_.values():
                log_value_list.append(obj1_)
        return log_value_list
    
    def find_element_and_scroll_into_view(self, element_text, elements_object_list, element_container_object, element_index=1):
        """
        Description : Scroll mosuse to bring element to visible area.
        :arg1 - element_container_object - Parent or container element of elemnt object
        Note - This method won't work for dynamic objects. Exp - Data filed tree. I will work for new home page domains tree and ect. 
        """
        element_container_bottom_y = int(core_utilobj.get_web_element_coordinate(self, element_container_object, 'bottom_middle')['y'])
        element_container_top_y = int(core_utilobj.get_web_element_coordinate(self, element_container_object, 'bottom_middle')['y'])
        item_index_list = JavaScript.find_all_index_of_element_by_text(self, elements_object_list, element_text)
        max_scroll = len(elements_object_list)
        scroll_count = 0
        if len(item_index_list)>0 :
            core_utilobj.python_move_to_element(self, element_container_object)
            item_obj = elements_object_list[item_index_list[element_index-1]]    
            while True :
                item_bottom_y = int(core_utilobj.get_web_element_coordinate(self, item_obj, 'bottom_middle')['y'])
                item_top_y = int(core_utilobj.get_web_element_coordinate(self, item_obj, 'top_middle')['y'])
                if element_container_bottom_y < item_bottom_y and scroll_count < max_scroll:
                    UtillityMethods.mouse_scroll(self, 'down', '1', option='uiautomation')
                    scroll_count+=1
                elif item_top_y < element_container_top_y and scroll_count < max_scroll:
                    UtillityMethods.mouse_scroll(self, 'up', '1', option='uiautomation')
                    scroll_count+=1
                else : 
                    break
            return item_obj
        else :
            return None
        
    def verify_placeholder(self, place_holder_css, place_holder_text, msg):
        """
        Description : This method it used to verify placeholder text value
        usage : verify_placeholder(".main-box", "30px", "Step 01.01 : verify placeholder value")
        """
        place_holder_obj = UtillityMethods.validate_and_get_webdriver_object(self, place_holder_css, "place holder css")
        actual_place_holder_value = place_holder_obj.get_attribute('placeholder')
        UtillityMethods.asequal(self, place_holder_text, actual_place_holder_value, msg)
        
    def verify_number_of_browser_windows(self, total_windows, step_num):
        """
        Descriptiion : Verify the number of browser windows
        :Usage - verify_number_of_browser_windows(2, '02.01')
        """
        total_windows = int(total_windows)
        actual_total_windows = len(self.driver.window_handles)
        msg = "Step {0} : Verify {1} windows opened in current browser".format(step_num, total_windows)
        UtillityMethods.asequal(self, total_windows, actual_total_windows, msg)
    
    def list_values_verification(self, expected_list, acutal_list, step_num, values_name, assert_type, value_len=None, slicing=(None, None)):
        """
        Description: Verify the given list values with difference types of assert type.
        :Parameters:
            expected_list:list = List which contains expected values to verify.
            acutal_list:list = List which contains actual values to verify.
            step_num:str = Verification msg step number. Exmp: "03.01".
            values_name:str = Name of the values to compose verification message. Exmp: ContextMenu, X-Axisi Labels ans etc.
            assert_type:str: Method of assert. Exmp: "eqaul" , "notin", "in"
            value_len:int: Length of the list values. List should contains only string values.
            slicing:tuple: Slicing index values to slice given list before verify
        """
        assert_types = ("equal", "in", "notin")
        if assert_type not in assert_types:
            raise TypeError("[{}] is invalid assert type. Valid types are {}".format(assert_type, assert_types))
        acutal_list = acutal_list.copy()[slicing[0]:slicing[1]]
        if value_len:
            expected_list = [value[:value_len] for value in expected_list]
            acutal_list = [value[:value_len] for value in acutal_list]
        if assert_type == assert_types[0]:
            msg = "Step {0} : Verify {1}".format(step_num, values_name)
            self.as_List_equal(expected_list, acutal_list, msg)
        elif assert_type == assert_types[1]:
            missing_values = set(expected_list) - set(acutal_list)
            msg = "Step {0} : Verify {1} in {2}".format(step_num, expected_list, values_name)
            if len(expected_list) == 0:
                self.asin(expected_list, acutal_list, msg)
            elif missing_values:
                self.asin(list(missing_values), acutal_list, msg)
            else:
                self.asequal(True, True, msg)
        elif assert_type == assert_types[2]:
            msg = "Step {0} : Verify {1} not in {2}".format(step_num, expected_list, values_name)
            not_missing_values = set(expected_list).intersection(acutal_list)
            if not_missing_values:
                self.asin(list(not_missing_values), acutal_list, msg)
            else:
                self.asequal(True, True, msg)
        else:
            raise NotImplemented
        
    def verify_element_enabled(self, css_locator, element_name, step_num, parent_obj=None):
        """
        Description: Verify function will verify object is enabled based on css property opacity and pointer event
        :Usage - verify_element_enabled(css_selector, 'Drill Anywhere', '01')
        """
        element_obj = self.validate_and_get_webdriver_object_using_locator(css_locator, element_name, parent_obj)
        msg = "Step {0}: Verify [{1}] is enabled".format(step_num, element_name)
        state = True if element_obj.value_of_css_property('opacity') == '1' and element_obj.value_of_css_property('auto') == 'none' else False
        self.asequal(False, state, msg)
        
    def verify_element_disabled(self, css_locator, element_name, step_num, parent_obj=None):
        """
        Description: Verify function will verify object is disabled based on css property opacity and pointer event
        :Usage - verify_element_disabled(css_selector, 'Drill Anywhere', '01')
        """
        element_obj = self.validate_and_get_webdriver_object_using_locator(css_locator, element_name, parent_obj)
        msg = "Step {0}: Verify [{1}] is disabled".format(step_num, element_name)
        state = True if element_obj.value_of_css_property('opacity') == '0.5' and element_obj.value_of_css_property('pointer-events') == 'none' else False
        self.asequal(True, state, msg)
        